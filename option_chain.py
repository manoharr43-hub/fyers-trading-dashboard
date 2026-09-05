"""
option_chain.py (ENHANCED VERSION)
==================================
NSE India Options Chain Dashboard with AI-Powered Price Action Signals
+ Institutional Buy/Sell Pressure Analysis

Data Source: FYERS (Primary) → NSE (Fallback for option chain only)
Live Signals: MSS, HH/HL/LH/LL, BOS, CHoCH, VWAP, EMA, RSI, MACD, Volume, RVOL
Confirmation: 5M, 15M, 30M, 1H, 1D multi-timeframe analysis
Trade Signal Output: BUY/SELL/HOLD with Entry, SL, T1, T2, T3, Probability, Confidence
Pressure Signals: Buy Pressure, Sell Pressure, Net Bias, Aggression, Volume/OI Anomalies

This version includes institutional buy/sell pressure indicators
while preserving all original functionality.
"""

from __future__ import annotations

import io
import logging
import math
import time
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from typing import Any, Optional
from collections import deque

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import requests
import streamlit as st
from plotly.subplots import make_subplots
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════
# 1. LOGGING (ORIGINAL)
# ══════════════════════════════════════════════════════════════════════════

logger = logging.getLogger("option_chain_dashboard")
if not logger.handlers:
    _handler = logging.StreamHandler()
    _handler.setFormatter(
        logging.Formatter("%(asctime)s [%(levelname)s] %(name)s: %(message)s")
    )
    logger.addHandler(_handler)
logger.setLevel(logging.INFO)


# ══════════════════════════════════════════════════════════════════════════
# 2. CONSTANTS (ORIGINAL - UNMODIFIED)
# ══════════════════════════════════════════════════════════════════════════

NSE_BASE_URL = "https://www.nseindia.com"
NSE_INDEX_CHAIN_URL = f"{NSE_BASE_URL}/api/option-chain-indices"
NSE_EQUITY_CHAIN_URL = f"{NSE_BASE_URL}/api/option-chain-equities"

INDEX_SYMBOLS: dict[str, str] = {
    "NIFTY": "NIFTY",
    "BANKNIFTY": "BANKNIFTY",
    "FINNIFTY": "FINNIFTY",
    "MIDCPNIFTY": "MIDCPNIFTY",
    "SENSEX": "SENSEX",
}

NSE_UNSUPPORTED_INDICES: set[str] = {"SENSEX", "BANKEX"}

DEFAULT_LOT_SIZES: dict[str, int] = {
    "NIFTY": 25,
    "BANKNIFTY": 15,
    "FINNIFTY": 25,
    "MIDCPNIFTY": 50,
    "SENSEX": 10,
    "BANKEX": 15,
    "_STOCK_DEFAULT": 1,
}

RISK_FREE_RATE = 0.07
MIN_SIGMA = 0.01
MAX_SIGMA = 5.0
TRADING_DAYS_MIN_T = 0.25

REQUEST_TIMEOUT = 10
MAX_RETRIES = 3
RETRY_BACKOFF_SECONDS = 1.5

_NSE_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    ),
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Referer": f"{NSE_BASE_URL}/option-chain",
    "Connection": "keep-alive",
}

REQUIRED_CHAIN_COLUMNS = ["strike_price", "ce_ltp", "ce_oi", "pe_ltp", "pe_oi"]

DARK_BG = "#0d1117"
PANEL_BG = "#161b22"
BORDER_COLOR = "#30363d"
TEXT_MAIN = "#e6edf3"
TEXT_MUTED = "#8b949e"
GREEN = "#3fb950"
RED = "#f85149"
AMBER = "#d29922"
BLUE = "#58a6ff"

TIMEFRAMES = {
    "5M": 5 * 60,
    "15M": 15 * 60,
    "30M": 30 * 60,
    "1H": 60 * 60,
    "1D": 24 * 60 * 60,
}

# ══════════════════════════════════════════════════════════════════════════
# SCALPING / BIG-MOVEMENT EARLY WARNING (ADDITIVE - ORIGINAL LOGIC UNCHANGED)
# ══════════════════════════════════════════════════════════════════════════
SCALPING_TIMEFRAMES = {
    "1M": 1,
    "3M": 3,
    "5M": 5,
    "15M": 15,
}
SCALPING_MIN_CANDLES = 30
SCALP_RVOL_THRESHOLD = 1.30
SCALP_ATR_PERIOD = 14
SCALP_BREAKOUT_LOOKBACK = 12
SCALP_EARLY_SCORE_THRESHOLD = 70.0

# MOVEMENT-BEFORE-IT-HAPPENS HISTORY
MOVEMENT_HISTORY_KEY = "oc_movement_history"
MOVEMENT_HISTORY_MAX = 60
MOVEMENT_EARLY_THRESHOLD = 65.0
MOVEMENT_STRONG_THRESHOLD = 78.0
MOVEMENT_MIN_RISING_SCANS = 2

DEFAULT_RSI_PERIOD = 14
DEFAULT_EMA_PERIODS = {"fast": 9, "slow": 21}
DEFAULT_MACD_PARAMS = {"fast": 12, "slow": 26, "signal": 9}
DEFAULT_VWAP_PERIOD = 20

MSS_MIN_STRENGTH = 1.0
BOS_CONFIRMATION_BARS = 1
CHOCH_CONFIRMATION_BARS = 2

FYERS_INDEX_SYMBOL_CANDIDATES: dict[str, list[str]] = {
    "NIFTY": ["NSE:NIFTY50-INDEX"],
    "BANKNIFTY": ["NSE:NIFTYBANK-INDEX", "NSE:BANKNIFTY-INDEX"],
    "FINNIFTY": ["NSE:FINNIFTY-INDEX"],
    "MIDCPNIFTY": ["NSE:MIDCPNIFTY-INDEX", "NSE:MIDCAPNIFTY-INDEX"],
    "SENSEX": ["BSE:SENSEX-INDEX", "BSE:SENSEX-INDEX50"],
    "BANKEX": ["BSE:BANKEX-INDEX"],
}


# ══════════════════════════════════════════════════════════════════════════
# 0. BUY/SELL PRESSURE DATATYPES (NEW - Non-intrusive)
# ══════════════════════════════════════════════════════════════════════════

@dataclass
class MarketPressure:
    """Market-wide pressure summary."""
    total_call_pressure: float = 50.0
    total_put_pressure: float = 50.0
    net_market_bias: float = 0.0
    market_sentiment: str = "NEUTRAL"
    pcr_vs_pressure: str = "N/A"
    volume_surge_detected: bool = False
    oi_accumulation_detected: bool = False
    itm_pressure: float = 50.0
    atm_pressure: float = 50.0
    otm_pressure: float = 50.0


# ══════════════════════════════════════════════════════════════════════════
# 3. HTTP / SESSION LAYER (ORIGINAL - UNMODIFIED)
# ══════════════════════════════════════════════════════════════════════════

def _build_retrying_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(_NSE_HEADERS)
    retry_cfg = Retry(
        total=MAX_RETRIES,
        backoff_factor=RETRY_BACKOFF_SECONDS,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=("GET",),
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry_cfg)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session


@st.cache_resource(show_spinner=False)
def get_nse_session() -> requests.Session:
    session = _build_retrying_session()
    _warm_up_session(session)
    return session


def _warm_up_session(session: requests.Session) -> bool:
    try:
        session.get(NSE_BASE_URL, timeout=REQUEST_TIMEOUT)
        session.get(f"{NSE_BASE_URL}/option-chain", timeout=REQUEST_TIMEOUT)
        return True
    except requests.exceptions.RequestException as e:
        logger.warning("NSE session warm-up failed: %s", e)
        return False


def fetch_json_with_retry(
    session: requests.Session, url: str, params: Optional[dict] = None,
    max_retries: int = MAX_RETRIES,
) -> tuple[Optional[dict], Optional[str]]:
    last_error = "Unknown error"
    for attempt in range(1, max_retries + 1):
        try:
            resp = session.get(url, params=params, timeout=REQUEST_TIMEOUT)
        except requests.exceptions.Timeout:
            last_error = f"Timeout on attempt {attempt}/{max_retries}"
            logger.warning("%s for %s", last_error, url)
            time.sleep(RETRY_BACKOFF_SECONDS * attempt)
            continue
        except requests.exceptions.ConnectionError as e:
            last_error = f"Connection error on attempt {attempt}/{max_retries}: {e}"
            logger.warning(last_error)
            time.sleep(RETRY_BACKOFF_SECONDS * attempt)
            continue
        except requests.exceptions.RequestException as e:
            last_error = f"Request exception on attempt {attempt}/{max_retries}: {e}"
            logger.warning(last_error)
            time.sleep(RETRY_BACKOFF_SECONDS * attempt)
            continue

        if resp.status_code in (401, 403):
            last_error = f"HTTP {resp.status_code} (stale session) on attempt {attempt}/{max_retries}"
            logger.warning("%s — re-warming NSE session and retrying", last_error)
            _warm_up_session(session)
            time.sleep(RETRY_BACKOFF_SECONDS * attempt)
            continue

        if resp.status_code != 200:
            last_error = f"HTTP {resp.status_code} on attempt {attempt}/{max_retries}"
            logger.warning(last_error)
            time.sleep(RETRY_BACKOFF_SECONDS * attempt)
            continue

        try:
            payload = resp.json()
        except ValueError as e:
            last_error = f"Invalid JSON on attempt {attempt}/{max_retries}: {e}"
            logger.warning(last_error)
            time.sleep(RETRY_BACKOFF_SECONDS * attempt)
            continue

        if not payload:
            last_error = f"Empty JSON payload on attempt {attempt}/{max_retries}"
            logger.warning(last_error)
            time.sleep(RETRY_BACKOFF_SECONDS * attempt)
            continue

        return payload, None

    logger.error("fetch_json_with_retry exhausted all retries for %s: %s", url, last_error)
    return None, last_error


# ══════════════════════════════════════════════════════════════════════════
# 4. FYERS LIVE DATA FUNCTIONS (ORIGINAL - UNMODIFIED)
# ══════════════════════════════════════════════════════════════════════════

def _fyers_field(d: dict, *aliases: str, default: Any = None) -> Any:
    for alias in aliases:
        if alias in d and d[alias] is not None:
            return d[alias]
    return default


def _safe_num(val: Any, default: float = 0.0) -> float:
    try:
        if val is None:
            return default
        f = float(val)
        if math.isnan(f) or math.isinf(f):
            return default
        return f
    except (TypeError, ValueError):
        return default


def normalize_stock_symbol(raw: str) -> str:
    s = (raw or "").strip().upper()
    if s.endswith("-EQ"):
        s = s[:-3]
    if ":" in s:
        s = s.split(":")[-1]
    return s


def fyers_stock_symbol_candidates(stock: str) -> list[str]:
    base = normalize_stock_symbol(stock)
    return [f"NSE:{base}-EQ", f"NSE:{base}"]


def _fyers_index_candidates(symbol_key: str) -> list[str]:
    return FYERS_INDEX_SYMBOL_CANDIDATES.get(symbol_key, [f"NSE:{symbol_key}-INDEX"])


def _fyers_call_optionchain(fyers: Any, symbol: str, strikecount: int, timestamp: str = "") -> Optional[dict]:
    req: dict[str, Any] = {"symbol": symbol, "strikecount": int(strikecount)}
    if timestamp:
        req["timestamp"] = str(timestamp)
    try:
        return fyers.optionchain(data=req)
    except Exception as e:
        logger.warning("FYERS optionchain() call raised for %s: %s", symbol, e)
        return None


def _fyers_call_history(fyers: Any, symbol: str, resolution: str, count: int = 100) -> Optional[dict]:
    """Fetch OHLCV candles from FYERS using a valid historical date range."""
    try:
        end_dt = datetime.now(timezone.utc)
        start_dt = end_dt - timedelta(days=7)

        req = {
            "symbol": symbol,
            "resolution": str(resolution),
            "date_format": "1",
            "range_from": str(int(start_dt.timestamp())),
            "range_to": str(int(end_dt.timestamp())),
            "cont_flag": "1",
        }
        return fyers.history(data=req)
    except Exception as e:
        logger.warning("FYERS history() call raised for %s (res %s): %s", symbol, resolution, e)
        return None


def fetch_fyers_candles(fyers: Any, symbol: str, timeframe_minutes: int, count: int = 100) -> Optional[pd.DataFrame]:
    """Fetches OHLCV candles from FYERS for a given timeframe."""
    if fyers is None:
        return None

    resolution_map = {
        1: "1",
        3: "3",
        5: "5",
        15: "15",
        30: "30",
        60: "60",
        1440: "1D",
    }
    resolution = resolution_map.get(timeframe_minutes, str(timeframe_minutes))

    resp = _fyers_call_history(fyers, symbol, resolution, count)
    if not isinstance(resp, dict) or resp.get("s") != "ok":
        logger.warning("FYERS history returned non-ok status for %s: %s", symbol, resp.get("s") if resp else None)
        return None

    data = resp.get("candles", []) if isinstance(resp.get("candles"), list) else []
    if not data:
        logger.warning("FYERS history returned empty candles for %s", symbol)
        return None

    rows = []
    for candle in data:
        if not isinstance(candle, list) or len(candle) < 5:
            continue
        try:
            rows.append({
                "timestamp": int(candle[0]) if len(candle) > 0 else 0,
                "open": float(candle[1]) if len(candle) > 1 else 0.0,
                "high": float(candle[2]) if len(candle) > 2 else 0.0,
                "low": float(candle[3]) if len(candle) > 3 else 0.0,
                "close": float(candle[4]) if len(candle) > 4 else 0.0,
                "volume": float(candle[5]) if len(candle) > 5 else 0.0,
            })
        except (TypeError, ValueError, IndexError):
            continue

    if not rows:
        return None

    df = pd.DataFrame(rows)
    df = df.sort_values("timestamp").reset_index(drop=True)
    return df.tail(max(1, int(count))).reset_index(drop=True)


# ══════════════════════════════════════════════════════════════════════════
# 5. TECHNICAL INDICATOR FUNCTIONS (ORIGINAL - UNMODIFIED)
# ══════════════════════════════════════════════════════════════════════════

def calculate_rsi(df: pd.DataFrame, period: int = DEFAULT_RSI_PERIOD, col: str = "close") -> pd.Series:
    """Calculate RSI (Relative Strength Index)."""
    if df.empty or col not in df.columns:
        return pd.Series(index=df.index, dtype=float)
    
    delta = df[col].diff()
    gain = delta.where(delta > 0, 0.0)
    loss = -delta.where(delta < 0, 0.0)
    
    avg_gain = gain.rolling(window=period, min_periods=1).mean()
    avg_loss = loss.rolling(window=period, min_periods=1).mean()
    
    rs = avg_gain / avg_loss.replace(0, 1e-10)
    rsi = 100 - (100 / (1 + rs))
    return rsi.fillna(50.0)


def calculate_ema(df: pd.DataFrame, period: int, col: str = "close") -> pd.Series:
    """Calculate Exponential Moving Average."""
    if df.empty or col not in df.columns:
        return pd.Series(index=df.index, dtype=float)
    return df[col].ewm(span=period, adjust=False).mean()


def calculate_macd(df: pd.DataFrame, fast: int = DEFAULT_MACD_PARAMS["fast"],
                   slow: int = DEFAULT_MACD_PARAMS["slow"],
                   signal: int = DEFAULT_MACD_PARAMS["signal"],
                   col: str = "close") -> tuple[pd.Series, pd.Series, pd.Series]:
    """Calculate MACD, Signal line, and Histogram."""
    if df.empty or col not in df.columns:
        return pd.Series(0, index=df.index), pd.Series(0, index=df.index), pd.Series(0, index=df.index)
    
    ema_fast = calculate_ema(df, fast, col)
    ema_slow = calculate_ema(df, slow, col)
    macd = ema_fast - ema_slow
    signal_line = macd.ewm(span=signal, adjust=False).mean()
    histogram = macd - signal_line
    
    return macd, signal_line, histogram


def calculate_vwap(df: pd.DataFrame) -> pd.Series:
    """Calculate Volume Weighted Average Price."""
    if df.empty or not all(c in df.columns for c in ["high", "low", "close", "volume"]):
        return pd.Series(index=df.index, dtype=float)
    
    typical_price = (df["high"] + df["low"] + df["close"]) / 3
    vwap = (typical_price * df["volume"]).cumsum() / df["volume"].cumsum()
    return vwap.fillna(df["close"])


def calculate_rvol(df: pd.DataFrame, period: int = 20) -> pd.Series:
    """Calculate Relative Volume (current volume vs average)."""
    if df.empty or "volume" not in df.columns:
        return pd.Series(1.0, index=df.index)
    
    avg_vol = df["volume"].rolling(window=period, min_periods=1).mean()
    rvol = df["volume"] / avg_vol.replace(0, 1.0)
    return rvol.fillna(1.0)


def add_technical_indicators(df: pd.DataFrame) -> pd.DataFrame:
    """Add all technical indicators to DataFrame."""
    if df.empty:
        return df
    
    d = df.copy()
    
    d["rsi"] = calculate_rsi(d)
    d["ema_9"] = calculate_ema(d, 9)
    d["ema_21"] = calculate_ema(d, 21)
    d["macd"], d["macd_signal"], d["macd_hist"] = calculate_macd(d)
    d["vwap"] = calculate_vwap(d)
    d["rvol"] = calculate_rvol(d)
    
    return d


# ══════════════════════════════════════════════════════════════════════════
# 6. MARKET STRUCTURE DETECTION (ORIGINAL - UNMODIFIED)
# ══════════════════════════════════════════════════════════════════════════

def detect_hh_ll(df: pd.DataFrame) -> tuple[pd.Series, pd.Series]:
    """Detect Higher High (HH) and Lower Low (LL) relative to previous candle."""
    if df.empty or len(df) < 2:
        return pd.Series(False, index=df.index), pd.Series(False, index=df.index)
    
    hh = pd.Series(False, index=df.index)
    ll = pd.Series(False, index=df.index)
    
    for i in range(1, len(df)):
        hh.iloc[i] = df["high"].iloc[i] > df["high"].iloc[i-1]
        ll.iloc[i] = df["low"].iloc[i] < df["low"].iloc[i-1]
    
    return hh, ll


def detect_structure_levels(df: pd.DataFrame, lookback: int = 5) -> dict[str, float]:
    """Detect major support and resistance levels (highs and lows)."""
    if df.empty or len(df) < lookback:
        return {"resistance": 0.0, "support": 0.0, "recent_high": 0.0, "recent_low": 0.0}
    
    recent = df.tail(lookback)
    return {
        "resistance": float(recent["high"].max()),
        "support": float(recent["low"].min()),
        "recent_high": float(df["high"].iloc[-1]),
        "recent_low": float(df["low"].iloc[-1]),
    }


def detect_bos(df: pd.DataFrame, structure_levels: dict) -> bool:
    """Detect Break of Structure (BOS)."""
    if df.empty or len(df) < 2:
        return False
    
    resistance = structure_levels.get("resistance", 0.0)
    support = structure_levels.get("support", 0.0)
    current_high = df["high"].iloc[-1]
    current_low = df["low"].iloc[-1]
    
    bos_up = current_high > resistance and resistance > 0
    bos_down = current_low < support and support > 0
    
    return bos_up or bos_down


def detect_choch(df: pd.DataFrame, lookback: int = 10) -> bool:
    """Detect Change of Character (CHoCH)."""
    if df.empty or len(df) < lookback:
        return False
    
    recent = df.tail(lookback)
    lows = recent["low"].values
    highs = recent["high"].values
    
    lower_lows = sum(1 for i in range(1, len(lows)) if lows[i] < lows[i-1])
    lower_highs = sum(1 for i in range(1, len(highs)) if highs[i] < highs[i-1])
    bearish_shift = (lower_lows >= lookback - 2) and (lower_highs >= lookback - 2)
    
    higher_lows = sum(1 for i in range(1, len(lows)) if lows[i] > lows[i-1])
    higher_highs = sum(1 for i in range(1, len(highs)) if highs[i] > highs[i-1])
    bullish_shift = (higher_lows >= lookback - 2) and (higher_highs >= lookback - 2)
    
    return bearish_shift or bullish_shift


def detect_mss(df_list: dict[str, pd.DataFrame]) -> dict[str, dict]:
    """Detect Market Structure Shift (MSS) across timeframes."""
    result = {tf: {"mss": False, "direction": "NONE", "strength": 0.0} for tf in df_list.keys()}
    
    if not df_list or not all(df_list.values()):
        return result
    
    if "5M" in df_list and not df_list["5M"].empty:
        df_5m = df_list["5M"]
        levels_5m = detect_structure_levels(df_5m)
        bos_5m = detect_bos(df_5m, levels_5m)
        
        if bos_5m and len(df_5m) >= 2:
            close_5m = df_5m["close"].iloc[-1]
            open_5m = df_5m["open"].iloc[-1]
            
            if close_5m > open_5m:
                direction = "UP"
                strength = abs((close_5m - levels_5m.get("support", close_5m)) / levels_5m.get("support", 1)) * 100
            else:
                direction = "DOWN"
                strength = abs((levels_5m.get("resistance", close_5m) - close_5m) / levels_5m.get("resistance", 1)) * 100
            
            result["5M"]["mss"] = strength >= MSS_MIN_STRENGTH
            result["5M"]["direction"] = direction if strength >= MSS_MIN_STRENGTH else "NONE"
            result["5M"]["strength"] = min(strength, 100.0)
    
    for tf in ["15M", "30M", "1H", "1D"]:
        if tf not in df_list or df_list[tf] is None or df_list[tf].empty:
            continue
        
        df = df_list[tf]
        choch = detect_choch(df)
        
        if choch:
            close = df["close"].iloc[-1]
            open_ = df["open"].iloc[-1]
            direction = "UP" if close > open_ else "DOWN"
            result[tf]["mss"] = True
            result[tf]["direction"] = direction
            result[tf]["strength"] = 75.0
    
    return result


# ══════════════════════════════════════════════════════════════════════════
# 7. TRADE SIGNAL GENERATION (ORIGINAL - UNMODIFIED)
# ══════════════════════════════════════════════════════════════════════════


def calculate_atr(df: pd.DataFrame, period: int = SCALP_ATR_PERIOD) -> pd.Series:
    """ATR for short-term volatility/range detection."""
    if df.empty or not all(c in df.columns for c in ("high", "low", "close")):
        return pd.Series(index=df.index, dtype=float)
    prev_close = df["close"].shift(1)
    tr = pd.concat(
        [
            df["high"] - df["low"],
            (df["high"] - prev_close).abs(),
            (df["low"] - prev_close).abs(),
        ],
        axis=1,
    ).max(axis=1)
    return tr.rolling(period, min_periods=1).mean()


def compute_scalping_early_warning(df_dict: dict[str, pd.DataFrame], spot: float) -> dict[str, Any]:
    """
    Additive early-warning model for a possible large short-term move.
    It ranks compression/range expansion, RVOL, EMA, VWAP and breakout proximity.
    It is an alert model, not a guaranteed prediction.
    """
    result = {
        "enabled": True, "status": "WAIT", "direction": "NEUTRAL",
        "score": 0.0, "confidence": 0.0, "trigger": "No setup",
        "reasons": [], "timeframe": "1M/3M/5M",
        "entry": 0.0, "stop_loss": 0.0, "target_1": 0.0,
        "target_2": 0.0, "target_3": 0.0,
    }
    if not df_dict:
        result.update(enabled=False, trigger="No FYERS candle data")
        return result

    df = next(
        (df_dict.get(tf) for tf in ("1M", "3M", "5M")
         if isinstance(df_dict.get(tf), pd.DataFrame) and not df_dict[tf].empty),
        None,
    )
    if df is None or len(df) < SCALPING_MIN_CANDLES:
        result.update(enabled=False, trigger="Insufficient scalping candles")
        return result

    d = df.copy()
    d["atr"] = calculate_atr(d)
    d["rvol"] = calculate_rvol(d, 20)
    d["ema_9"] = calculate_ema(d, 9)
    d["ema_21"] = calculate_ema(d, 21)
    d["vwap"] = calculate_vwap(d)

    last, prev = d.iloc[-1], d.iloc[-2]
    recent = d.iloc[-(SCALP_BREAKOUT_LOOKBACK + 1):-1]
    if recent.empty:
        recent = d.iloc[:-1].tail(SCALP_BREAKOUT_LOOKBACK)

    close = float(last["close"])
    atr = max(float(last["atr"]), 1e-9)
    rvol = float(last["rvol"])
    candle_range = float(last["high"] - last["low"])
    body_ratio = abs(float(last["close"] - last["open"])) / max(candle_range, 1e-9)
    ema9, ema21, vwap = float(last["ema_9"]), float(last["ema_21"]), float(last["vwap"])
    recent_high = float(recent["high"].max()) if not recent.empty else float(last["high"])
    recent_low = float(recent["low"].min()) if not recent.empty else float(last["low"])

    score_up = score_down = 0.0
    up_reasons, down_reasons = [], []

    if candle_range >= atr * 1.15 and body_ratio >= 0.55:
        if close > float(last["open"]):
            score_up += 20; up_reasons.append("range expansion")
        else:
            score_down += 20; down_reasons.append("range expansion")

    if rvol >= SCALP_RVOL_THRESHOLD:
        if close >= float(last["open"]):
            score_up += 20; up_reasons.append(f"RVOL {rvol:.1f}x")
        else:
            score_down += 20; down_reasons.append(f"RVOL {rvol:.1f}x")

    if ema9 > ema21:
        score_up += 15; up_reasons.append("EMA 9 > EMA 21")
    elif ema9 < ema21:
        score_down += 15; down_reasons.append("EMA 9 < EMA 21")

    if close > vwap:
        score_up += 10; up_reasons.append("price above VWAP")
    elif close < vwap:
        score_down += 10; down_reasons.append("price below VWAP")

    distance_up_atr = (recent_high - close) / atr
    distance_down_atr = (close - recent_low) / atr
    if close > recent_high:
        score_up += 25; up_reasons.append("recent high breakout")
    elif close < recent_low:
        score_down += 25; down_reasons.append("recent low breakdown")
    elif 0 <= distance_up_atr <= 0.35:
        score_up += 10; up_reasons.append("near resistance")
    elif 0 <= distance_down_atr <= 0.35:
        score_down += 10; down_reasons.append("near support")

    if close > float(last["open"]) and float(prev["close"]) >= float(prev["open"]):
        score_up += 10; up_reasons.append("2-candle momentum")
    elif close < float(last["open"]) and float(prev["close"]) <= float(prev["open"]):
        score_down += 10; down_reasons.append("2-candle momentum")

    if score_up >= score_down and score_up >= SCALP_EARLY_SCORE_THRESHOLD:
        direction, score, reasons = "UP", score_up, up_reasons
    elif score_down > score_up and score_down >= SCALP_EARLY_SCORE_THRESHOLD:
        direction, score, reasons = "DOWN", score_down, down_reasons
    else:
        direction, score, reasons = "NEUTRAL", max(score_up, score_down), (
            up_reasons if score_up >= score_down else down_reasons
        )

    tf5 = df_dict.get("5M")
    if isinstance(tf5, pd.DataFrame) and len(tf5) >= 21 and direction in ("UP", "DOWN"):
        htf = tf5.copy()
        htf["ema_9"] = calculate_ema(htf, 9)
        htf["ema_21"] = calculate_ema(htf, 21)
        htf_up = float(htf["ema_9"].iloc[-1]) > float(htf["ema_21"].iloc[-1])
        htf_down = float(htf["ema_9"].iloc[-1]) < float(htf["ema_21"].iloc[-1])
        aligned = (direction == "UP" and htf_up) or (direction == "DOWN" and htf_down)
        if aligned:
            score = min(100.0, score + 5); reasons.append("5M trend aligned")
        else:
            score = max(0.0, score - 5); reasons.append("5M trend not aligned")

    if direction in ("UP", "DOWN"):
        risk = max(atr * 0.8, close * 0.001)
        if direction == "UP":
            entry, sl = close, close - risk
            t1, t2, t3 = close + risk, close + 1.5 * risk, close + 2.0 * risk
        else:
            entry, sl = close, close + risk
            t1, t2, t3 = close - risk, close - 1.5 * risk, close - 2.0 * risk
        result.update(
            status="EARLY BUY" if direction == "UP" else "EARLY SELL",
            direction=direction, score=round(min(score, 100.0), 1),
            confidence=round(min(max(score, 0.0), 95.0), 1),
            trigger=" + ".join(dict.fromkeys(reasons)) or "Short-term momentum",
            reasons=list(dict.fromkeys(reasons)), entry=round(entry, 2),
            stop_loss=round(sl, 2), target_1=round(t1, 2),
            target_2=round(t2, 2), target_3=round(t3, 2),
        )
    else:
        result.update(
            status="WATCH", score=round(min(score, 100.0), 1),
            confidence=round(min(score, 95.0), 1),
            trigger=" + ".join(dict.fromkeys(reasons)) or "Waiting for expansion",
            reasons=list(dict.fromkeys(reasons)),
        )
    return result


def _render_scalping_panel(scalp: dict[str, Any]) -> None:
    st.markdown(
        '<div class="block-title">⚡ SCALPING MODE — BIG-MOVEMENT EARLY WARNING</div>',
        unsafe_allow_html=True,
    )
    if not scalp or not scalp.get("enabled"):
        st.info("Scalping mode is waiting for FYERS 1M/3M/5M/15M candle data.")
        return

    status, direction = scalp.get("status", "WAIT"), scalp.get("direction", "NEUTRAL")
    score = float(scalp.get("score", 0) or 0)
    cols = st.columns(5)
    cols[0].metric("SCALP STATUS", status)
    cols[1].metric("DIRECTION", direction)
    cols[2].metric("EARLY SCORE", f"{score:.0f}/100")
    cols[3].metric("ENTRY", f"₹{float(scalp.get('entry', 0) or 0):,.2f}" if scalp.get("entry") else "—")
    cols[4].metric("SL", f"₹{float(scalp.get('stop_loss', 0) or 0):,.2f}" if scalp.get("stop_loss") else "—")

    if status == "EARLY BUY":
        st.success("🟢 Early BUY setup — wait for breakout/volume confirmation before acting.")
    elif status == "EARLY SELL":
        st.error("🔴 Early SELL setup — wait for breakdown/volume confirmation before acting.")
    else:
        st.warning("🟡 WATCH — no strong short-term expansion setup yet.")

    st.caption("Reasons: " + (", ".join(scalp.get("reasons", [])) or "Waiting for confirmation"))
    if scalp.get("entry"):
        st.caption(
            f"Targets: T1 ₹{scalp['target_1']:,.2f} | "
            f"T2 ₹{scalp['target_2']:,.2f} | T3 ₹{scalp['target_3']:,.2f}"
        )


@dataclass
class TradeSignal:
    """Represents a single trade signal with all relevant details."""
    signal: str
    entry: float
    stop_loss: float
    target_1: float
    target_2: float
    target_3: float
    risk_reward_ratio: float
    probability: float
    confidence: float
    confirmation_timeframes: list[str] = field(default_factory=list)
    technical_reasons: list[str] = field(default_factory=list)
    timestamp: datetime = field(default_factory=datetime.now)


def generate_trade_signal(df_dict: dict[str, pd.DataFrame], spot: float, mss: dict[str, dict],
                          fyers_available: bool) -> Optional[TradeSignal]:
    """Generate a multi-timeframe confirmed trade signal."""
    
    if not fyers_available:
        logger.warning("FYERS not available - cannot generate trade signal")
        return None
    
    if not df_dict or not any(df_dict.values()):
        return None
    
    df_5m = df_dict.get("5M")
    if df_5m is None or df_5m.empty or len(df_5m) < 10:
        return None
    
    current_close = float(df_5m["close"].iloc[-1])
    current_high = float(df_5m["high"].iloc[-1])
    current_low = float(df_5m["low"].iloc[-1])
    current_rsi = float(df_5m["rsi"].iloc[-1]) if "rsi" in df_5m.columns else 50.0
    current_ema_9 = float(df_5m["ema_9"].iloc[-1]) if "ema_9" in df_5m.columns else current_close
    current_ema_21 = float(df_5m["ema_21"].iloc[-1]) if "ema_21" in df_5m.columns else current_close
    current_macd = float(df_5m["macd"].iloc[-1]) if "macd" in df_5m.columns else 0.0
    current_macd_hist = float(df_5m["macd_hist"].iloc[-1]) if "macd_hist" in df_5m.columns else 0.0
    current_rvol = float(df_5m["rvol"].iloc[-1]) if "rvol" in df_5m.columns else 1.0
    
    levels_5m = detect_structure_levels(df_5m)
    resistance = levels_5m.get("resistance", current_close)
    support = levels_5m.get("support", current_close)
    
    signal_type = "HOLD"
    confidence_score = 0.0
    probability_score = 50.0
    technical_reasons = []
    confirmed_tfs = []
    
    buy_score = 0.0
    
    if current_close > current_ema_9:
        buy_score += 25
        technical_reasons.append("Price > EMA 9")
    
    if current_ema_9 > current_ema_21:
        buy_score += 20
        technical_reasons.append("EMA 9 > EMA 21")
    
    if current_macd > 0 and current_macd_hist > 0:
        buy_score += 20
        technical_reasons.append("MACD bullish")
    
    if 40 <= current_rsi <= 70:
        buy_score += 15
        technical_reasons.append(f"RSI {current_rsi:.0f} (bullish zone)")
    
    if current_rvol > 1.2:
        buy_score += 10
        technical_reasons.append("High volume")
    
    if mss.get("5M", {}).get("mss") and mss["5M"].get("direction") == "UP":
        buy_score += 15
        technical_reasons.append("MSS confirmed (UP)")
        confirmed_tfs.append("5M")
    
    sell_score = 0.0
    
    if current_close < current_ema_9:
        sell_score += 25
        technical_reasons.append("Price < EMA 9")
    
    if current_ema_9 < current_ema_21:
        sell_score += 20
        technical_reasons.append("EMA 9 < EMA 21")
    
    if current_macd < 0 and current_macd_hist < 0:
        sell_score += 20
        technical_reasons.append("MACD bearish")
    
    if 30 <= current_rsi <= 60:
        sell_score += 15
        technical_reasons.append(f"RSI {current_rsi:.0f} (bearish zone)")
    
    if current_rvol > 1.2:
        sell_score += 10
        technical_reasons.append("High volume")
    
    if mss.get("5M", {}).get("mss") and mss["5M"].get("direction") == "DOWN":
        sell_score += 15
        technical_reasons.append("MSS confirmed (DOWN)")
        confirmed_tfs.append("5M")
    
    if buy_score > sell_score and buy_score >= 60:
        signal_type = "BUY"
        confidence_score = min(buy_score, 100.0)
        probability_score = 50.0 + (buy_score / 2)
    elif sell_score > buy_score and sell_score >= 60:
        signal_type = "SELL"
        confidence_score = min(sell_score, 100.0)
        probability_score = 50.0 + (sell_score / 2)
    else:
        signal_type = "HOLD"
        confidence_score = max(buy_score, sell_score)
        probability_score = 50.0
    
    if signal_type == "BUY":
        entry = current_close
        stop_loss = support * 0.995
        range_val = entry - stop_loss
        target_1 = entry + range_val
        target_2 = entry + (range_val * 1.5)
        target_3 = entry + (range_val * 2.0)
    elif signal_type == "SELL":
        entry = current_close
        stop_loss = resistance * 1.005
        range_val = stop_loss - entry
        target_1 = entry - range_val
        target_2 = entry - (range_val * 1.5)
        target_3 = entry - (range_val * 2.0)
    else:
        entry = current_close
        stop_loss = support
        target_1 = (resistance + entry) / 2
        target_2 = resistance
        target_3 = resistance * 1.01
    
    risk_reward = abs(entry - target_1) / abs(entry - stop_loss) if entry != stop_loss else 1.0
    
    return TradeSignal(
        signal=signal_type,
        entry=entry,
        stop_loss=stop_loss,
        target_1=target_1,
        target_2=target_2,
        target_3=target_3,
        risk_reward_ratio=risk_reward,
        probability=min(probability_score, 100.0),
        confidence=confidence_score,
        confirmation_timeframes=confirmed_tfs if confirmed_tfs else ["5M"],
        technical_reasons=technical_reasons if technical_reasons else ["Neutral"],
    )


# ══════════════════════════════════════════════════════════════════════════
# 8. DATA FETCH + PARSE LAYER (ORIGINAL - UNMODIFIED)
# ══════════════════════════════════════════════════════════════════════════

@st.cache_data(ttl=15, show_spinner=False)
def fetch_option_chain_raw(symbol: str, is_index: bool) -> dict:
    """Cached (15s TTL) raw NSE option-chain JSON fetch."""
    session = get_nse_session()
    url = NSE_INDEX_CHAIN_URL if is_index else NSE_EQUITY_CHAIN_URL
    payload, error = fetch_json_with_retry(session, url, params={"symbol": symbol})
    if payload is None:
        return {"ok": False, "payload": None, "error": error or "No data returned."}
    records = payload.get("records") if isinstance(payload, dict) else None
    if not isinstance(records, dict) or not records.get("data"):
        return {"ok": False, "payload": payload, "error": "Response had no option-chain records."}
    return {"ok": True, "payload": payload, "error": None}


def parse_option_chain(payload: dict, preferred_expiry: str = "") -> tuple[pd.DataFrame, dict]:
    """Parses NSE's raw option-chain payload."""
    meta = {
        "spot_price": 0.0, "expiry_dates": [], "selected_expiry": "",
        "fetched_at": datetime.now(), "total_rows_seen": 0, "rows_parsed": 0,
    }
    records = payload.get("records", {}) if isinstance(payload, dict) else {}
    chain = records.get("data", []) if isinstance(records, dict) else []
    meta["spot_price"] = _safe_num(records.get("underlyingValue"))
    expiry_dates = records.get("expiryDates", []) or []
    meta["expiry_dates"] = expiry_dates

    if not chain:
        return pd.DataFrame(), meta

    selected_expiry = preferred_expiry if preferred_expiry in expiry_dates else (
        expiry_dates[0] if expiry_dates else ""
    )
    meta["selected_expiry"] = selected_expiry

    rows = []
    meta["total_rows_seen"] = len(chain)
    for item in chain:
        if not isinstance(item, dict):
            continue
        if selected_expiry and item.get("expiryDate") != selected_expiry:
            continue
        strike = item.get("strikePrice")
        if strike is None:
            continue
        ce, pe = item.get("CE") or {}, item.get("PE") or {}
        rows.append({
            "strike_price": _safe_num(strike),
            "ce_ltp": _safe_num(ce.get("lastPrice")),
            "ce_change": _safe_num(ce.get("change")),
            "ce_bid": _safe_num(ce.get("bidprice")),
            "ce_bid_qty": _safe_num(ce.get("bidQty")),
            "ce_ask": _safe_num(ce.get("askPrice")),
            "ce_ask_qty": _safe_num(ce.get("askQty")),
            "ce_volume": _safe_num(ce.get("totalTradedVolume")),
            "ce_oi": _safe_num(ce.get("openInterest")),
            "ce_chng_oi": _safe_num(ce.get("changeinOpenInterest")),
            "ce_oi_change_pct": _safe_num(ce.get("pchangeinOpenInterest")),
            "ce_iv": _safe_num(ce.get("impliedVolatility")),
            "pe_ltp": _safe_num(pe.get("lastPrice")),
            "pe_change": _safe_num(pe.get("change")),
            "pe_bid": _safe_num(pe.get("bidprice")),
            "pe_bid_qty": _safe_num(pe.get("bidQty")),
            "pe_ask": _safe_num(pe.get("askPrice")),
            "pe_ask_qty": _safe_num(pe.get("askQty")),
            "pe_volume": _safe_num(pe.get("totalTradedVolume")),
            "pe_oi": _safe_num(pe.get("openInterest")),
            "pe_chng_oi": _safe_num(pe.get("changeinOpenInterest")),
            "pe_oi_change_pct": _safe_num(pe.get("pchangeinOpenInterest")),
            "pe_iv": _safe_num(pe.get("impliedVolatility")),
        })

    meta["rows_parsed"] = len(rows)
    if not rows:
        return pd.DataFrame(), meta

    df = pd.DataFrame(rows)
    df = df.groupby("strike_price", as_index=False).first()
    df.sort_values("strike_price", inplace=True)
    df.reset_index(drop=True, inplace=True)
    return df, meta


def validate_chain_df(df: pd.DataFrame) -> bool:
    try:
        if df is None or not isinstance(df, pd.DataFrame) or df.empty:
            return False
        if not all(c in df.columns for c in REQUIRED_CHAIN_COLUMNS):
            return False
        strikes = pd.to_numeric(df["strike_price"], errors="coerce").dropna()
        return bool((strikes > 0).sum() > 0)
    except Exception as e:
        logger.error("validate_chain_df raised an exception: %s", e)
        return False


def filter_strikes_around_atm(df: pd.DataFrame, spot: float, n_each_side: int) -> pd.DataFrame:
    if df is None or df.empty or n_each_side <= 0:
        return df
    d = df.sort_values("strike_price").reset_index(drop=True)
    ref = spot if spot else float(d["strike_price"].median())
    atm_idx = int((d["strike_price"] - ref).abs().idxmin())
    lo = max(0, atm_idx - n_each_side)
    hi = min(len(d), atm_idx + n_each_side + 1)
    return d.iloc[lo:hi].reset_index(drop=True)


def parse_days_to_expiry(expiry_label: str) -> float:
    if not expiry_label:
        return 7.0
    for fmt in ("%d-%b-%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            exp_dt = datetime.strptime(expiry_label, fmt)
            delta_days = (exp_dt.replace(hour=15, minute=30) - datetime.now()).total_seconds() / 86400
            return max(delta_days, TRADING_DAYS_MIN_T)
        except ValueError:
            continue
    return 7.0


# ══════════════════════════════════════════════════════════════════════════
# 9. GREEKS ENGINE (ORIGINAL - UNMODIFIED)
# ══════════════════════════════════════════════════════════════════════════

def _norm_cdf(x: float) -> float:
    return 0.5 * (1.0 + math.erf(x / math.sqrt(2.0)))


def _norm_pdf(x: float) -> float:
    return math.exp(-0.5 * x * x) / math.sqrt(2.0 * math.pi)


def bs_greeks(spot: float, strike: float, t_years: float, r: float, sigma: float,
              is_call: bool) -> dict[str, float]:
    if spot <= 0 or strike <= 0 or t_years <= 0 or sigma <= 0:
        return {"delta": 0.0, "gamma": 0.0, "theta": 0.0, "vega": 0.0}
    sigma = min(max(sigma, MIN_SIGMA), MAX_SIGMA)
    sqrt_t = math.sqrt(t_years)
    try:
        d1 = (math.log(spot / strike) + (r + 0.5 * sigma ** 2) * t_years) / (sigma * sqrt_t)
        d2 = d1 - sigma * sqrt_t
    except (ValueError, ZeroDivisionError):
        return {"delta": 0.0, "gamma": 0.0, "theta": 0.0, "vega": 0.0}

    pdf_d1 = _norm_pdf(d1)
    gamma = pdf_d1 / (spot * sigma * sqrt_t)
    vega = spot * pdf_d1 * sqrt_t / 100.0

    if is_call:
        delta = _norm_cdf(d1)
        theta = (
            -(spot * pdf_d1 * sigma) / (2 * sqrt_t)
            - r * strike * math.exp(-r * t_years) * _norm_cdf(d2)
        ) / 365.0
    else:
        delta = _norm_cdf(d1) - 1.0
        theta = (
            -(spot * pdf_d1 * sigma) / (2 * sqrt_t)
            + r * strike * math.exp(-r * t_years) * _norm_cdf(-d2)
        ) / 365.0

    return {
        "delta": round(delta, 4), "gamma": round(gamma, 6),
        "theta": round(theta, 4), "vega": round(vega, 4),
    }


def add_greeks_columns(df: pd.DataFrame, spot: float, expiry_label: str,
                        r: float = RISK_FREE_RATE) -> pd.DataFrame:
    d = df.copy()
    if d.empty:
        for col in ("ce_delta", "ce_gamma", "ce_theta", "ce_vega",
                    "pe_delta", "pe_gamma", "pe_theta", "pe_vega"):
            d[col] = 0.0
        return d

    t_years = parse_days_to_expiry(expiry_label) / 365.0

    ce_g = d.apply(
        lambda row: bs_greeks(spot, row["strike_price"], t_years, r, row["ce_iv"] / 100.0, True),
        axis=1,
    )
    pe_g = d.apply(
        lambda row: bs_greeks(spot, row["strike_price"], t_years, r, row["pe_iv"] / 100.0, False),
        axis=1,
    )
    for key in ("delta", "gamma", "theta", "vega"):
        d[f"ce_{key}"] = ce_g.apply(lambda x: x[key])
        d[f"pe_{key}"] = pe_g.apply(lambda x: x[key])
    return d


# ══════════════════════════════════════════════════════════════════════════
# 10. IV RANK / IV PERCENTILE (ORIGINAL - UNMODIFIED)
# ══════════════════════════════════════════════════════════════════════════

IV_HISTORY_KEY = "oc_atm_iv_history"
IV_HISTORY_MAX_POINTS = 500


def _atm_iv(df: pd.DataFrame, spot: float) -> float:
    if df.empty or not spot:
        return 0.0
    idx = (df["strike_price"] - spot).abs().idxmin()
    row = df.loc[idx]
    ivs = [v for v in (row.get("ce_iv", 0), row.get("pe_iv", 0)) if v and v > 0]
    return float(np.mean(ivs)) if ivs else 0.0


def update_iv_history(symbol: str, expiry_label: str, atm_iv: float) -> None:
    if atm_iv <= 0:
        return
    history = st.session_state.setdefault(IV_HISTORY_KEY, {})
    key = f"{symbol}|{expiry_label}"
    series = history.get(key, [])
    series.append(atm_iv)
    if len(series) > IV_HISTORY_MAX_POINTS:
        series = series[-IV_HISTORY_MAX_POINTS:]
    history[key] = series
    st.session_state[IV_HISTORY_KEY] = history


def compute_iv_rank_percentile(symbol: str, expiry_label: str, current_iv: float) -> tuple[float, float]:
    history = st.session_state.get(IV_HISTORY_KEY, {})
    series = history.get(f"{symbol}|{expiry_label}", [])
    if len(series) < 2 or current_iv <= 0:
        return 0.0, 0.0
    lo, hi = min(series), max(series)
    iv_rank = ((current_iv - lo) / (hi - lo)) * 100 if hi > lo else 50.0
    iv_percentile = (sum(1 for v in series if v <= current_iv) / len(series)) * 100
    return round(float(np.clip(iv_rank, 0, 100)), 1), round(iv_percentile, 1)


# ══════════════════════════════════════════════════════════════════════════
# 11. GEX / DEX (ORIGINAL - UNMODIFIED)
# ══════════════════════════════════════════════════════════════════════════

def compute_gex_dex(df: pd.DataFrame, spot: float, lot_size: int) -> dict[str, Any]:
    if df.empty or not spot:
        return {"total_gex": 0.0, "total_dex": 0.0, "by_strike": pd.DataFrame(),
                "max_gex_strike": None, "min_gex_strike": None, "gamma_flip": None}

    d = df.copy()
    d["gex"] = (
        (d.get("ce_gamma", 0) * d.get("ce_oi", 0)) - (d.get("pe_gamma", 0) * d.get("pe_oi", 0))
    ) * (spot ** 2) * 0.01 * lot_size
    d["dex"] = (
        (d.get("ce_delta", 0) * d.get("ce_oi", 0)) + (d.get("pe_delta", 0) * d.get("pe_oi", 0))
    ) * spot * lot_size

    total_gex = float(d["gex"].sum())
    total_dex = float(d["dex"].sum())
    max_gex_row = d.loc[d["gex"].idxmax()] if len(d) else None
    min_gex_row = d.loc[d["gex"].idxmin()] if len(d) else None

    d_sorted = d.sort_values("strike_price").reset_index(drop=True)
    cum_gex = d_sorted["gex"].cumsum()
    gamma_flip = None
    sign_changes = np.where(np.diff(np.sign(cum_gex.replace(0, np.nan).ffill().fillna(0))) != 0)[0]
    if len(sign_changes) > 0:
        idx = int(sign_changes[0])
        gamma_flip = float(d_sorted.loc[idx, "strike_price"])

    return {
        "total_gex": total_gex, "total_dex": total_dex,
        "by_strike": d[["strike_price", "gex", "dex"]],
        "max_gex_strike": float(max_gex_row["strike_price"]) if max_gex_row is not None else None,
        "min_gex_strike": float(min_gex_row["strike_price"]) if min_gex_row is not None else None,
        "gamma_flip": gamma_flip,
    }


# ══════════════════════════════════════════════════════════════════════════
# 12. FYERS OPTION CHAIN PARSING (ORIGINAL - UNMODIFIED)
# ══════════════════════════════════════════════════════════════════════════

def _fyers_extract_expiry_list(response: dict) -> list[tuple[str, str]]:
    data = response.get("data", {}) if isinstance(response, dict) else {}
    raw = data.get("expiryData") or data.get("expirydata") or []
    out = []
    for item in raw if isinstance(raw, list) else []:
        if not isinstance(item, dict):
            continue
        ts = item.get("expiry") or item.get("timestamp")
        if ts is None:
            continue
        try:
            label = datetime.fromtimestamp(int(float(ts))).strftime("%d-%b-%Y")
        except (TypeError, ValueError, OSError):
            label = str(ts)
        out.append((label, str(ts)))
    seen: set = set()
    deduped = []
    for label, ts in out:
        if ts not in seen:
            seen.add(ts)
            deduped.append((label, ts))

    def _ts_key(pair: tuple[str, str]) -> float:
        try:
            return float(pair[1])
        except (TypeError, ValueError):
            return 0.0

    deduped.sort(key=_ts_key)
    return deduped


def _fyers_extract_chain_rows(response: dict) -> tuple[list[dict], dict]:
    data = response.get("data", {}) if isinstance(response, dict) else {}
    for key in ("optionsChain", "options", "optionschain"):
        candidate = data.get(key)
        if isinstance(candidate, list) and candidate:
            return candidate, data
    return [], data


def _fyers_extract_spot(response: dict, data: dict) -> float:
    for src in (data, response if isinstance(response, dict) else {}):
        if not isinstance(src, dict):
            continue
        for key in ("ltp", "spot_price", "spotPrice", "underlyingValue", "underlying_value"):
            f = _safe_num(src.get(key), 0.0)
            if f > 0:
                return f
    return 0.0


def _bs_price(spot: float, strike: float, t: float, r: float, sigma: float, is_call: bool) -> float:
    if t <= 0 or sigma <= 0 or spot <= 0 or strike <= 0:
        return max(0.0, (spot - strike) if is_call else (strike - spot))
    d1 = (math.log(spot / strike) + (r + 0.5 * sigma ** 2) * t) / (sigma * math.sqrt(t))
    d2 = d1 - sigma * math.sqrt(t)
    if is_call:
        return spot * _norm_cdf(d1) - strike * math.exp(-r * t) * _norm_cdf(d2)
    return strike * math.exp(-r * t) * _norm_cdf(-d2) - spot * _norm_cdf(-d1)


def implied_volatility(price: float, spot: float, strike: float, t_years: float,
                        is_call: bool, r: float = RISK_FREE_RATE) -> float:
    if price <= 0 or spot <= 0 or strike <= 0 or t_years <= 0:
        return 0.0
    sigma = 0.30
    for _ in range(50):
        model_price = _bs_price(spot, strike, t_years, r, sigma, is_call)
        try:
            d1 = (math.log(spot / strike) + (r + 0.5 * sigma ** 2) * t_years) / (sigma * math.sqrt(t_years))
        except (ValueError, ZeroDivisionError):
            break
        vega_raw = spot * _norm_pdf(d1) * math.sqrt(t_years)
        diff = model_price - price
        if abs(diff) < 1e-4:
            break
        if vega_raw < 1e-8:
            break
        sigma -= diff / vega_raw
        sigma = max(MIN_SIGMA, min(sigma, MAX_SIGMA))
    return round(sigma * 100, 2)


def parse_fyers_chain(rows: list[dict], spot: float, expiry_label: str) -> pd.DataFrame:
    ce_rows: dict[float, dict] = {}
    pe_rows: dict[float, dict] = {}
    for item in rows:
        if not isinstance(item, dict):
            continue
        opt_type = str(_fyers_field(item, "option_type", "optionType", "type", default="")).upper()
        strike = _safe_num(_fyers_field(item, "strike_price", "strikePrice"))
        if strike <= 0 or opt_type not in ("CE", "PE"):
            continue
        rec = {
            "ltp": _safe_num(_fyers_field(item, "ltp", "last_price")),
            "change": _safe_num(_fyers_field(item, "ltpch", "change")),
            "bid": _safe_num(_fyers_field(item, "bid", "bidprice", "bidPrice")),
            "ask": _safe_num(_fyers_field(item, "ask", "askprice", "askPrice")),
            "volume": _safe_num(_fyers_field(item, "volume", "vol")),
            "oi": _safe_num(_fyers_field(item, "oi", "openInterest")),
            "chng_oi": _safe_num(_fyers_field(item, "oich", "chng_oi", "change_oi")),
            "oi_change_pct": _safe_num(_fyers_field(item, "oichp", "pchangeinOpenInterest")),
            "iv": _safe_num(_fyers_field(item, "iv", "impliedVolatility")),
        }
        (ce_rows if opt_type == "CE" else pe_rows)[strike] = rec

    strikes = sorted(set(ce_rows) | set(pe_rows))
    if not strikes:
        return pd.DataFrame()

    t_years = parse_days_to_expiry(expiry_label) / 365.0
    out_rows = []
    for strike in strikes:
        ce, pe = ce_rows.get(strike, {}), pe_rows.get(strike, {})
        ce_iv = ce.get("iv", 0.0)
        if ce_iv <= 0 and ce.get("ltp", 0) > 0 and spot > 0:
            ce_iv = implied_volatility(ce["ltp"], spot, strike, t_years, True)
        pe_iv = pe.get("iv", 0.0)
        if pe_iv <= 0 and pe.get("ltp", 0) > 0 and spot > 0:
            pe_iv = implied_volatility(pe["ltp"], spot, strike, t_years, False)
        out_rows.append({
            "strike_price": strike,
            "ce_ltp": ce.get("ltp", 0.0), "ce_change": ce.get("change", 0.0),
            "ce_bid": ce.get("bid", 0.0), "ce_bid_qty": 0.0,
            "ce_ask": ce.get("ask", 0.0), "ce_ask_qty": 0.0,
            "ce_volume": ce.get("volume", 0.0), "ce_oi": ce.get("oi", 0.0),
            "ce_chng_oi": ce.get("chng_oi", 0.0), "ce_oi_change_pct": ce.get("oi_change_pct", 0.0),
            "ce_iv": round(ce_iv, 2),
            "pe_ltp": pe.get("ltp", 0.0), "pe_change": pe.get("change", 0.0),
            "pe_bid": pe.get("bid", 0.0), "pe_bid_qty": 0.0,
            "pe_ask": pe.get("ask", 0.0), "pe_ask_qty": 0.0,
            "pe_volume": pe.get("volume", 0.0), "pe_oi": pe.get("oi", 0.0),
            "pe_chng_oi": pe.get("chng_oi", 0.0), "pe_oi_change_pct": pe.get("oi_change_pct", 0.0),
            "pe_iv": round(pe_iv, 2),
        })
    df = pd.DataFrame(out_rows)
    df.sort_values("strike_price", inplace=True)
    df.reset_index(drop=True, inplace=True)
    return df


def fetch_via_fyers(fyers: Any, symbol_key: str, is_index: bool, stock_name: str,
                     preferred_expiry: str, strike_count: int) -> dict:
    symbol_candidates = (
        _fyers_index_candidates(symbol_key) if is_index else fyers_stock_symbol_candidates(stock_name)
    )
    if not symbol_candidates:
        return {"ok": False, "df": pd.DataFrame(), "meta": {}, "error": "No FYERS symbol candidates resolved."}

    expiry_resp, used_symbol = None, symbol_candidates[0]
    for sym in symbol_candidates:
        resp = _fyers_call_optionchain(fyers, sym, strikecount=2)
        if isinstance(resp, dict) and resp.get("s") == "ok":
            expiry_resp, used_symbol = resp, sym
            break
    if expiry_resp is None:
        return {"ok": False, "df": pd.DataFrame(), "meta": {},
                "error": "FYERS returned no usable response for any symbol variant tried."}

    expiry_list = _fyers_extract_expiry_list(expiry_resp)
    if not expiry_list:
        return {"ok": False, "df": pd.DataFrame(), "meta": {}, "error": "FYERS returned no expiry dates."}

    selected_label, selected_ts = expiry_list[0]
    for label, ts in expiry_list:
        if label == preferred_expiry:
            selected_label, selected_ts = label, ts
            break

    chain_resp = _fyers_call_optionchain(fyers, used_symbol, strikecount=strike_count, timestamp=selected_ts)
    if not isinstance(chain_resp, dict) or chain_resp.get("s") != "ok":
        code = chain_resp.get("code") if isinstance(chain_resp, dict) else "—"
        return {"ok": False, "df": pd.DataFrame(), "meta": {}, "error": f"FYERS chain fetch failed (code {code})."}

    rows, data = _fyers_extract_chain_rows(chain_resp)
    if not rows:
        return {"ok": False, "df": pd.DataFrame(), "meta": {}, "error": "FYERS returned an empty options chain."}

    spot = _fyers_extract_spot(chain_resp, data)
    if not spot:
        try:
            q = fyers.quotes(data={"symbols": used_symbol})
            qv = q.get("d", [{}])[0].get("v", {}) if isinstance(q, dict) else {}
            spot = _safe_num(qv.get("lp"), 0.0)
        except Exception as e:
            logger.warning("FYERS quotes() spot-price fallback raised: %s", e)

    df = parse_fyers_chain(rows, spot, selected_label)
    if not validate_chain_df(df):
        return {"ok": False, "df": pd.DataFrame(), "meta": {}, "error": "FYERS chain failed schema validation after parsing."}

    meta = {
        "spot_price": spot, "expiry_dates": [lbl for lbl, _ in expiry_list],
        "selected_expiry": selected_label, "fetched_at": datetime.now(),
        "total_rows_seen": len(rows), "rows_parsed": len(df),
    }
    return {"ok": True, "df": df, "meta": meta, "error": None}


def fetch_chain_unified(fyers: Any, symbol_key: str, is_index: bool, stock_name: str,
                         preferred_expiry: str, strike_count: int) -> dict:
    """FYERS-first, NSE-fallback."""
    fyers_error = None
    if fyers is not None:
        result = fetch_via_fyers(fyers, symbol_key, is_index, stock_name, preferred_expiry, strike_count)
        if result["ok"]:
            result["source"] = "FYERS"
            return result
        fyers_error = result.get("error")
        logger.warning("FYERS fetch failed, falling back to NSE: %s", fyers_error)

    if is_index and symbol_key in NSE_UNSUPPORTED_INDICES:
        error = (
            f"{symbol_key} is BSE-listed and NSE's public option-chain API does not serve it — "
            "a FYERS (or other BSE-capable) client is required for this index."
        )
        combined = f"FYERS: {fyers_error} | {error}" if fyers_error else error
        return {"ok": False, "df": pd.DataFrame(), "meta": {}, "error": combined, "source": "NONE"}

    nse_symbol = symbol_key if is_index else normalize_stock_symbol(stock_name)
    raw_result = fetch_option_chain_raw(nse_symbol, is_index)
    if not raw_result.get("ok"):
        error = raw_result.get("error") or "NSE fetch failed."
        combined = f"FYERS: {fyers_error} | NSE: {error}" if fyers_error else error
        return {"ok": False, "df": pd.DataFrame(), "meta": {}, "error": combined, "source": "NONE"}

    df, meta = parse_option_chain(raw_result["payload"], preferred_expiry=preferred_expiry)
    if not validate_chain_df(df):
        error = "NSE response did not contain a usable option chain."
        combined = f"FYERS: {fyers_error} | NSE: {error}" if fyers_error else error
        return {"ok": False, "df": pd.DataFrame(), "meta": meta, "error": combined, "source": "NONE"}

    return {"ok": True, "df": df, "meta": meta, "error": None, "source": "NSE"}


# ══════════════════════════════════════════════════════════════════════════
# 13. ANALYTICS (ORIGINAL - UNMODIFIED)
# ══════════════════════════════════════════════════════════════════════════

def calc_pcr(df: pd.DataFrame) -> float:
    if df.empty:
        return 0.0
    total_ce = df["ce_oi"].sum()
    total_pe = df["pe_oi"].sum()
    return round(float(total_pe / total_ce), 3) if total_ce > 0 else 0.0


def calc_max_pain(df: pd.DataFrame) -> float:
    if df.empty:
        return 0.0
    strikes = df["strike_price"].values
    ce_oi = df["ce_oi"].values
    pe_oi = df["pe_oi"].values
    pain = [
        float(np.sum(np.maximum(s - strikes, 0) * ce_oi) + np.sum(np.maximum(strikes - s, 0) * pe_oi))
        for s in strikes
    ]
    return float(strikes[int(np.argmin(pain))]) if pain else 0.0


def calc_max_oi(df: pd.DataFrame) -> dict[str, Optional[float]]:
    if df.empty:
        return {"max_ce_oi_strike": None, "max_pe_oi_strike": None}
    return {
        "max_ce_oi_strike": float(df.loc[df["ce_oi"].idxmax(), "strike_price"]),
        "max_pe_oi_strike": float(df.loc[df["pe_oi"].idxmax(), "strike_price"]),
    }


def calc_support_resistance(df: pd.DataFrame) -> tuple[Optional[float], Optional[float]]:
    if df.empty:
        return None, None
    support = float(df.loc[df["pe_oi"].idxmax(), "strike_price"])
    resistance = float(df.loc[df["ce_oi"].idxmax(), "strike_price"])
    return support, resistance


def classify_buildup(df: pd.DataFrame) -> pd.DataFrame:
    d = df.copy()

    def _matrix(price_chg: float, oi_chg: float) -> str:
        if price_chg > 0 and oi_chg > 0:
            return "Long Buildup"
        if price_chg > 0 and oi_chg < 0:
            return "Short Covering"
        if price_chg < 0 and oi_chg > 0:
            return "Short Buildup"
        if price_chg < 0 and oi_chg < 0:
            return "Long Unwinding"
        return "Flat"

    d["CE Buildup"] = d.apply(lambda r: _matrix(r.get("ce_change", 0), r.get("ce_chng_oi", 0)), axis=1)
    d["PE Buildup"] = d.apply(lambda r: _matrix(r.get("pe_change", 0), r.get("pe_chng_oi", 0)), axis=1)

    d["Call Writing"] = d["ce_chng_oi"] > 0
    d["Call Unwinding"] = d["ce_chng_oi"] < 0
    d["Put Writing"] = d["pe_chng_oi"] > 0
    d["Put Unwinding"] = d["pe_chng_oi"] < 0
    return d


def classify_moneyness(df: pd.DataFrame, spot: float) -> pd.DataFrame:
    d = df.copy()
    if d.empty:
        d["ATM"] = False
        d["CE Moneyness"] = ""
        d["PE Moneyness"] = ""
        return d
    ref = spot if spot else float(d["strike_price"].median())
    atm_idx = (d["strike_price"] - ref).abs().idxmin()
    d["ATM"] = d.index == atm_idx
    d["CE Moneyness"] = np.where(
        d["ATM"], "ATM", np.where(d["strike_price"] < ref, "ITM", "OTM")
    )
    d["PE Moneyness"] = np.where(
        d["ATM"], "ATM", np.where(d["strike_price"] > ref, "ITM", "OTM")
    )
    return d


OI_SHIFT_HISTORY_KEY = "oc_prev_support_resistance"


def detect_oi_shift(symbol: str, expiry_label: str, support: Optional[float],
                     resistance: Optional[float]) -> list[str]:
    notes = []
    history = st.session_state.setdefault(OI_SHIFT_HISTORY_KEY, {})
    key = f"{symbol}|{expiry_label}"
    prev = history.get(key)
    if prev:
        if prev.get("support") is not None and support is not None and support != prev["support"]:
            direction = "up" if support > prev["support"] else "down"
            notes.append(f"Support shifted {direction}: {prev['support']:,.0f} -> {support:,.0f}")
        if prev.get("resistance") is not None and resistance is not None and resistance != prev["resistance"]:
            direction = "up" if resistance > prev["resistance"] else "down"
            notes.append(f"Resistance shifted {direction}: {prev['resistance']:,.0f} -> {resistance:,.0f}")
    history[key] = {"support": support, "resistance": resistance}
    st.session_state[OI_SHIFT_HISTORY_KEY] = history
    return notes


def _normalize_series(series: pd.Series) -> pd.Series:
    s = series.astype(float)
    if s.empty:
        return s
    if s.max() == s.min():
        return pd.Series(0.5, index=s.index)
    return (s - s.min()) / (s.max() - s.min())


# ══════════════════════════════════════════════════════════════════════════
# 13B. PO3 + OPTIONS INTELLIGENCE (ADDITIONAL / NON-INTRUSIVE)
# ══════════════════════════════════════════════════════════════════════════

PO3_HISTORY_KEY = "oc_po3_spot_history"
PO3_HISTORY_MAX = 120


def _po3_phase_from_price(df_price: Optional[pd.DataFrame], spot: float) -> tuple[str, str, float]:
    """Heuristic Power-of-3 phase classifier.

    Uses available underlying candles when FYERS is connected. If candles are
    unavailable, callers can fall back to the short spot-history classifier.
    This is an analytical heuristic, not a prediction or guaranteed PO3 model.
    """
    if df_price is None or df_price.empty or len(df_price) < 8:
        return "UNKNOWN", "NEUTRAL", 0.0

    d = df_price.tail(min(len(df_price), 30)).copy()
    close = pd.to_numeric(d.get("close"), errors="coerce").dropna()
    high = pd.to_numeric(d.get("high"), errors="coerce").dropna()
    low = pd.to_numeric(d.get("low"), errors="coerce").dropna()
    open_ = pd.to_numeric(d.get("open"), errors="coerce").dropna()
    if len(close) < 8 or len(high) < 8 or len(low) < 8:
        return "UNKNOWN", "NEUTRAL", 0.0

    recent = d.tail(8)
    prev = d.iloc[:-8] if len(d) > 8 else d.head(max(1, len(d) - 8))
    recent_high = float(pd.to_numeric(recent["high"], errors="coerce").max())
    recent_low = float(pd.to_numeric(recent["low"], errors="coerce").min())
    prior_high = float(pd.to_numeric(prev["high"], errors="coerce").max()) if not prev.empty else recent_high
    prior_low = float(pd.to_numeric(prev["low"], errors="coerce").min()) if not prev.empty else recent_low
    last = float(close.iloc[-1])
    prev_close = float(close.iloc[-8])
    move_pct = ((last - prev_close) / prev_close * 100.0) if prev_close else 0.0
    recent_range = ((recent_high - recent_low) / last * 100.0) if last else 0.0

    wick_up = recent_high > prior_high and last < recent_high * 0.9985
    wick_down = recent_low < prior_low and last > recent_low * 1.0015
    directional_up = move_pct > 0.35 and last >= float(close.tail(4).mean())
    directional_down = move_pct < -0.35 and last <= float(close.tail(4).mean())
    compressed = recent_range < 1.0

    if wick_up or wick_down:
        direction = "BEARISH" if wick_up and not wick_down else ("BULLISH" if wick_down and not wick_up else "NEUTRAL")
        return "MANIPULATION", direction, min(100.0, 55.0 + abs(move_pct) * 8.0)
    if directional_up:
        return "DISTRIBUTION", "BULLISH", min(100.0, 55.0 + abs(move_pct) * 10.0)
    if directional_down:
        return "DISTRIBUTION", "BEARISH", min(100.0, 55.0 + abs(move_pct) * 10.0)
    if compressed:
        return "ACCUMULATION", "NEUTRAL", 65.0
    return "ACCUMULATION", "NEUTRAL", 55.0


def _po3_phase_from_spot_history(symbol: str, spot: float) -> tuple[str, str, float]:
    """Fallback PO3 phase using recent live spot snapshots."""
    if not spot or spot <= 0:
        return "UNKNOWN", "NEUTRAL", 0.0
    history = st.session_state.setdefault(PO3_HISTORY_KEY, {})
    series = history.get(symbol, [])
    series.append(float(spot))
    history[symbol] = series[-PO3_HISTORY_MAX:]
    st.session_state[PO3_HISTORY_KEY] = history

    if len(series) < 8:
        return "ACCUMULATION", "NEUTRAL", 45.0
    window = np.asarray(series[-8:], dtype=float)
    move_pct = (window[-1] - window[0]) / window[0] * 100.0 if window[0] else 0.0
    rng_pct = (window.max() - window.min()) / window[-1] * 100.0 if window[-1] else 0.0
    if rng_pct < 0.15:
        return "ACCUMULATION", "NEUTRAL", 60.0
    if abs(move_pct) > 0.25:
        direction = "BULLISH" if move_pct > 0 else "BEARISH"
        return "DISTRIBUTION", direction, min(90.0, 55.0 + abs(move_pct) * 12.0)
    return "MANIPULATION", "BULLISH" if move_pct > 0 else ("BEARISH" if move_pct < 0 else "NEUTRAL"), 55.0


def compute_po3_options_intelligence(
    df: pd.DataFrame,
    spot: float,
    symbol: str,
    price_df: Optional[pd.DataFrame] = None,
) -> dict[str, Any]:
    """Build the requested 11-point PO3 + options confirmation panel."""
    if df is None or df.empty or not spot:
        return {
            "spot": spot, "atm_strike": 0.0, "ce_oi": 0.0, "pe_oi": 0.0,
            "ce_oi_change": 0.0, "pe_oi_change": 0.0, "pcr": 0.0,
            "call_writing": "N/A", "put_writing": "N/A", "call_unwinding": "N/A",
            "put_unwinding": "N/A", "spot_po3_phase": "UNKNOWN", "po3_phase": "UNKNOWN",
            "po3_direction": "NEUTRAL", "po3_options_confirmation": "NO DATA",
            "final_ce_bias": "NEUTRAL", "final_pe_bias": "NEUTRAL", "confidence": 0.0,
            "reason": "No option-chain data available."
        }

    atm_idx = (df["strike_price"] - float(spot)).abs().idxmin()
    atm = df.loc[atm_idx]
    atm_strike = float(atm["strike_price"])

    # Aggregate a tight ATM neighbourhood so a single strike does not dominate.
    strikes = sorted(pd.to_numeric(df["strike_price"], errors="coerce").dropna().unique())
    if len(strikes) > 1:
        step = float(np.median(np.diff(strikes)))
    else:
        step = max(abs(atm_strike) * 0.01, 1.0)
    near = df[(df["strike_price"] - atm_strike).abs() <= step * 2.0].copy()
    if near.empty:
        near = atm.to_frame().T

    ce_oi = float(near["ce_oi"].sum())
    pe_oi = float(near["pe_oi"].sum())
    ce_doi = float(near["ce_chng_oi"].sum())
    pe_doi = float(near["pe_chng_oi"].sum())
    pcr = round(pe_oi / ce_oi, 3) if ce_oi > 0 else 0.0

    def _writing(change: float) -> str:
        return "YES" if change > 0 else "NO"
    def _unwind(change: float) -> str:
        return "YES" if change < 0 else "NO"

    spot_phase, candle_direction, phase_conf = _po3_phase_from_price(price_df, spot) if price_df is not None else _po3_phase_from_spot_history(symbol, spot)

    # Options directional votes: CE-side bullish = put writing + call unwinding;
    # PE-side bearish = call writing + put unwinding. PCR is supporting context.
    ce_votes = 0
    pe_votes = 0
    reasons = []
    if pe_doi > 0:
        ce_votes += 1; reasons.append("Put writing")
    if ce_doi < 0:
        ce_votes += 1; reasons.append("Call unwinding")
    if ce_doi > 0:
        pe_votes += 1; reasons.append("Call writing")
    if pe_doi < 0:
        pe_votes += 1; reasons.append("Put unwinding")
    if pcr >= 1.20:
        ce_votes += 1; reasons.append(f"PCR {pcr:.2f} bullish")
    elif 0 < pcr <= 0.80:
        pe_votes += 1; reasons.append(f"PCR {pcr:.2f} bearish")

    if candle_direction == "BULLISH":
        ce_votes += 1
    elif candle_direction == "BEARISH":
        pe_votes += 1

    if ce_votes > pe_votes:
        option_dir = "BULLISH"
    elif pe_votes > ce_votes:
        option_dir = "BEARISH"
    else:
        option_dir = "NEUTRAL"

    if spot_phase == "MANIPULATION":
        phase_label = "MANIPULATION"
    elif spot_phase == "DISTRIBUTION":
        phase_label = "DISTRIBUTION"
    else:
        phase_label = spot_phase

    aligned = candle_direction != "NEUTRAL" and candle_direction == option_dir
    if aligned:
        confirmation = f"CONFIRMED {option_dir}"
        final_dir = option_dir
        conf = min(95.0, 55.0 + abs(ce_votes - pe_votes) * 8.0 + phase_conf * 0.20)
    elif option_dir != "NEUTRAL" and candle_direction == "NEUTRAL":
        confirmation = f"OPTIONS {option_dir} / PO3 NEUTRAL"
        final_dir = option_dir
        conf = min(78.0, 50.0 + max(ce_votes, pe_votes) * 7.0)
    elif candle_direction != "NEUTRAL" and option_dir == "NEUTRAL":
        confirmation = f"PO3 {candle_direction} / OPTIONS NEUTRAL"
        final_dir = candle_direction
        conf = min(72.0, 48.0 + phase_conf * 0.25)
    else:
        confirmation = "NO CONFIRMATION / CHOP"
        final_dir = "NEUTRAL"
        conf = 35.0

    return {
        "spot": float(spot), "atm_strike": atm_strike,
        "ce_oi": ce_oi, "pe_oi": pe_oi,
        "ce_oi_change": ce_doi, "pe_oi_change": pe_doi, "pcr": pcr,
        "call_writing": _writing(ce_doi), "put_writing": _writing(pe_doi),
        "call_unwinding": _unwind(ce_doi), "put_unwinding": _unwind(pe_doi),
        "spot_po3_phase": phase_label, "po3_phase": phase_label,
        "po3_direction": candle_direction, "po3_options_confirmation": confirmation,
        "final_ce_bias": "STRONG" if final_dir == "BULLISH" and conf >= 75 else ("YES" if final_dir == "BULLISH" else "NO"),
        "final_pe_bias": "STRONG" if final_dir == "BEARISH" and conf >= 75 else ("YES" if final_dir == "BEARISH" else "NO"),
        "final_direction": final_dir, "confidence": round(float(conf), 1),
        "reason": ", ".join(dict.fromkeys(reasons)) if reasons else "No strong options confirmation"
    }



def compute_final_signal(po3: dict[str, Any], trade_signal: Optional[TradeSignal] = None) -> dict[str, Any]:
    """One clear BUY/SELL/WAIT signal with CHOP and confidence protection."""
    po3 = po3 if isinstance(po3, dict) else {}
    direction = str(po3.get("final_direction", "NEUTRAL")).upper()
    confirmation = str(po3.get("po3_options_confirmation", "")).upper()
    confidence = float(po3.get("confidence", 0.0) or 0.0)

    trade_name = str(getattr(trade_signal, "signal", "") or "").upper() if trade_signal else ""
    trade_conf = float(getattr(trade_signal, "confidence", 0.0) or 0.0) if trade_signal else 0.0

    is_chop = (
        "CHOP" in confirmation
        or "NO CONFIRMATION" in confirmation
        or direction in ("", "NEUTRAL", "UNKNOWN")
    )

    if is_chop or confidence < 55.0:
        return {
            "signal": "WAIT",
            "next_candle": "WAIT",
            "confidence": round(confidence, 1),
            "market_status": "CHOP / WAIT",
            "reason": "CHOP or confidence below 55% — BUY/SELL blocked",
        }

    if direction == "BULLISH":
        return {
            "signal": "BUY",
            "next_candle": "UP",
            "confidence": round(min(max(confidence, trade_conf if "BUY" in trade_name else 0.0), 95.0), 1),
            "market_status": "BULLISH",
            "reason": po3.get("reason", "PO3 + options bullish confirmation"),
        }

    if direction == "BEARISH":
        return {
            "signal": "SELL",
            "next_candle": "DOWN",
            "confidence": round(min(max(confidence, trade_conf if "SELL" in trade_name else 0.0), 95.0), 1),
            "market_status": "BEARISH",
            "reason": po3.get("reason", "PO3 + options bearish confirmation"),
        }

    return {
        "signal": "WAIT",
        "next_candle": "WAIT",
        "confidence": round(confidence, 1),
        "market_status": "NEUTRAL",
        "reason": "No confirmed directional alignment",
    }


def _render_final_signal_card(final_signal: dict[str, Any]) -> None:
    sig = str(final_signal.get("signal", "WAIT")).upper()
    nxt = str(final_signal.get("next_candle", "WAIT")).upper()
    conf = float(final_signal.get("confidence", 0.0) or 0.0)

    a, b, c, d = st.columns(4)
    a.metric("FINAL SIGNAL", "🟢 BUY" if sig == "BUY" else ("🔴 SELL" if sig == "SELL" else "🟡 WAIT"))
    b.metric("NEXT CANDLE", "🟢 UP" if nxt == "UP" else ("🔴 DOWN" if nxt == "DOWN" else "🟡 WAIT"))
    c.metric("MARKET", str(final_signal.get("market_status", "NEUTRAL")))
    d.metric("CONFIDENCE", f"{conf:.0f}%")

    if sig == "BUY":
        st.success("🟢 FINAL BUY — confirmed bullish conditions")
    elif sig == "SELL":
        st.error("🔴 FINAL SELL — confirmed bearish conditions")
    else:
        st.warning("🟡 FINAL WAIT — no strong confirmation; weak signals are blocked")

    if final_signal.get("reason"):
        st.caption("Final decision: " + str(final_signal["reason"]))


def _render_po3_intelligence(po3: dict[str, Any]) -> None:
    """Render compact 11-field PO3 + Options Intelligence dashboard."""
    st.markdown('<div class="block-title">🧠 PO3 + OPTIONS INTELLIGENCE</div>', unsafe_allow_html=True)
    r1 = st.columns(4)
    r1[0].metric("Spot / Underlying", f"₹{po3['spot']:,.2f}" if po3.get("spot") else "—")
    r1[1].metric("ATM Strike", f"₹{po3['atm_strike']:,.0f}" if po3.get("atm_strike") else "—")
    r1[2].metric("PO3 Phase", po3.get("po3_phase", "UNKNOWN"))
    r1[3].metric("PO3 Direction", po3.get("po3_direction", "NEUTRAL"))

    r2 = st.columns(4)
    r2[0].metric("CE OI", f"{po3.get('ce_oi', 0):,.0f}")
    r2[1].metric("PE OI", f"{po3.get('pe_oi', 0):,.0f}")
    r2[2].metric("CE OI Change", f"{po3.get('ce_oi_change', 0):+,.0f}")
    r2[3].metric("PE OI Change", f"{po3.get('pe_oi_change', 0):+,.0f}")

    r3 = st.columns(5)
    r3[0].metric("PCR", f"{po3.get('pcr', 0):.3f}")
    r3[1].metric("Call Writing", po3.get("call_writing", "N/A"))
    r3[2].metric("Put Writing", po3.get("put_writing", "N/A"))
    r3[3].metric("Call Unwinding", po3.get("call_unwinding", "N/A"))
    r3[4].metric("Put Unwinding", po3.get("put_unwinding", "N/A"))

    direction = po3.get("final_direction", "NEUTRAL")
    if direction == "BULLISH":
        st.success(f"🟢 PO3 + OPTIONS: {po3.get('po3_options_confirmation')} | Final CE Bias: {po3.get('final_ce_bias')} | Confidence: {po3.get('confidence', 0):.0f}%")
    elif direction == "BEARISH":
        st.error(f"🔴 PO3 + OPTIONS: {po3.get('po3_options_confirmation')} | Final PE Bias: {po3.get('final_pe_bias')} | Confidence: {po3.get('confidence', 0):.0f}%")
    else:
        st.warning(f"🟡 PO3 + OPTIONS: {po3.get('po3_options_confirmation')} | Final Bias: NEUTRAL | Confidence: {po3.get('confidence', 0):.0f}%")
    if po3.get("reason"):
        st.caption("Confirmation factors: " + po3["reason"])

    if po3.get("final_signal"):
        _render_final_signal_card(po3["final_signal"])


# ══════════════════════════════════════════════════════════════════════════
# 14. AI SIGNAL ENGINE (ORIGINAL - UNMODIFIED)
# ══════════════════════════════════════════════════════════════════════════

AI_SCORE_WEIGHTS = {
    "put_writing": 0.18, "call_unwind": 0.12, "volume": 0.12, "pcr_bias": 0.12,
    "proximity": 0.12, "max_pain_proximity": 0.10, "highest_oi": 0.10,
    "delta_oi_magnitude": 0.08, "iv_stability": 0.06,
}


def compute_ai_scores(df: pd.DataFrame, spot: float, atm_strike: float,
                       max_pain: float, pcr: float) -> pd.DataFrame:
    d = df.copy()
    if d.empty:
        d["CE Score"] = pd.Series(dtype=float)
        d["PE Score"] = pd.Series(dtype=float)
        return d

    ce_oi_s = _normalize_series(d["ce_oi"])
    pe_oi_s = _normalize_series(d["pe_oi"])
    pe_chng_s = _normalize_series(d["pe_chng_oi"])
    ce_chng_s = _normalize_series(d["ce_chng_oi"])
    ce_unwind_s = _normalize_series((-d["ce_chng_oi"]).clip(lower=0))
    pe_unwind_s = _normalize_series((-d["pe_chng_oi"]).clip(lower=0))
    ce_vol_s = _normalize_series(d["ce_volume"])
    pe_vol_s = _normalize_series(d["pe_volume"])
    delta_oi_mag_s = _normalize_series(d["ce_chng_oi"].abs() + d["pe_chng_oi"].abs())

    avg_ce_iv = d.loc[d["ce_iv"] > 0, "ce_iv"].mean() if (d["ce_iv"] > 0).any() else 0.0
    avg_pe_iv = d.loc[d["pe_iv"] > 0, "pe_iv"].mean() if (d["pe_iv"] > 0).any() else 0.0
    ce_iv_stability_s = _normalize_series(-(d["ce_iv"] - avg_ce_iv).abs())
    pe_iv_stability_s = _normalize_series(-(d["pe_iv"] - avg_pe_iv).abs())

    ref = spot if spot else (atm_strike if atm_strike else float(d["strike_price"].median()))
    proximity_s = 1 - _normalize_series((d["strike_price"] - ref).abs())
    maxpain_proximity_s = 1 - _normalize_series((d["strike_price"] - max_pain).abs()) if max_pain else pd.Series(0.5, index=d.index)

    pcr_bull_bias = float(np.clip(((pcr or 1.0) - 1.0), -1, 1))
    pcr_bull_s = (pcr_bull_bias + 1) / 2
    pcr_bear_s = 1 - pcr_bull_s

    w = AI_SCORE_WEIGHTS
    ce_score = (
        pe_chng_s * w["put_writing"] + ce_unwind_s * w["call_unwind"] + ce_vol_s * w["volume"]
        + pcr_bull_s * w["pcr_bias"] + proximity_s * w["proximity"]
        + maxpain_proximity_s * w["max_pain_proximity"] + ce_oi_s * w["highest_oi"]
        + delta_oi_mag_s * w["delta_oi_magnitude"] + ce_iv_stability_s * w["iv_stability"]
    ) * 100

    pe_score = (
        ce_chng_s * w["put_writing"] + pe_unwind_s * w["call_unwind"] + pe_vol_s * w["volume"]
        + pcr_bear_s * w["pcr_bias"] + proximity_s * w["proximity"]
        + maxpain_proximity_s * w["max_pain_proximity"] + pe_oi_s * w["highest_oi"]
        + delta_oi_mag_s * w["delta_oi_magnitude"] + pe_iv_stability_s * w["iv_stability"]
    ) * 100

    d["CE Score"] = ce_score.clip(0, 100).round(1)
    d["PE Score"] = pe_score.clip(0, 100).round(1)

    def _decision(row) -> str:
        ce, pe = row["CE Score"], row["PE Score"]
        if abs(ce - pe) < 3:
            return "HOLD"
        return "BUY CE" if ce > pe else "BUY PE"

    d["AI Signal"] = d.apply(_decision, axis=1)
    d["AI Confidence %"] = d[["CE Score", "PE Score"]].max(axis=1).round(1)
    return d


def detect_institutional_smart_money(df: pd.DataFrame) -> pd.DataFrame:
    d = df.copy()
    if d.empty:
        d["Institutional Signal"] = pd.Series(dtype=object)
        d["Smart Money"] = pd.Series(dtype=bool)
        return d

    ce_oi_q75 = d["ce_oi"].quantile(0.75) if d["ce_oi"].max() > 0 else 0
    pe_oi_q75 = d["pe_oi"].quantile(0.75) if d["pe_oi"].max() > 0 else 0
    ce_vol_med = d["ce_volume"].median()
    pe_vol_med = d["pe_volume"].median()

    def _inst_signal(row) -> str:
        ce_inst = row["ce_oi"] >= ce_oi_q75 > 0 and row["ce_chng_oi"] > 0 and row["ce_volume"] >= ce_vol_med
        pe_inst = row["pe_oi"] >= pe_oi_q75 > 0 and row["pe_chng_oi"] > 0 and row["pe_volume"] >= pe_vol_med
        if ce_inst and pe_inst:
            return "Institutional Activity (Both Sides)"
        if ce_inst:
            return "Institutional Call Writing"
        if pe_inst:
            return "Institutional Put Writing"
        return "None"

    d["Institutional Signal"] = d.apply(_inst_signal, axis=1)
    d["Smart Money"] = d["Institutional Signal"] != "None"
    return d


# ══════════════════════════════════════════════════════════════════════════
# 15. BUY/SELL PRESSURE FUNCTIONS (NEW - ADDED CLEANLY)
# ══════════════════════════════════════════════════════════════════════════

def _normalize_0_100(value, min_val: float = 0.0, max_val: float = 1.0):
    """Normalize value to 0-100 range. Handles both scalars and pandas Series."""
    if max_val == min_val:
        if isinstance(value, pd.Series):
            return pd.Series(50.0, index=value.index)
        return 50.0
    
    normalized = ((value - min_val) / (max_val - min_val)) * 100
    
    if isinstance(value, pd.Series):
        return normalized.clip(0, 100)
    return max(0.0, min(100.0, normalized))


def calculate_volume_pressure(df: pd.DataFrame) -> pd.DataFrame:
    """Calculate volume-based pressure."""
    d = df.copy()
    if d.empty or "ce_volume" not in d.columns or "pe_volume" not in d.columns:
        d["call_volume_pressure"] = 0.0
        d["put_volume_pressure"] = 0.0
        d["volume_ratio"] = 1.0
        return d
    
    d["ce_volume"] = d["ce_volume"].clip(lower=0)
    d["pe_volume"] = d["pe_volume"].clip(lower=0)
    
    total_ce_vol = d["ce_volume"].sum()
    total_pe_vol = d["pe_volume"].sum()
    max_vol = max(total_ce_vol, total_pe_vol, 1)
    
    d["call_volume_pressure"] = _normalize_0_100(d["ce_volume"], 0, max_vol)
    d["put_volume_pressure"] = _normalize_0_100(d["pe_volume"], 0, max_vol)
    d["volume_ratio"] = (d["ce_volume"] / (d["pe_volume"] + 1e-6)).clip(0, 10)
    
    return d


def calculate_oi_change_pressure(df: pd.DataFrame) -> pd.DataFrame:
    """Calculate OI change momentum."""
    d = df.copy()
    if d.empty or "ce_chng_oi" not in d.columns or "pe_chng_oi" not in d.columns:
        d["call_oi_accumulation"] = 0.0
        d["put_oi_accumulation"] = 0.0
        d["oi_momentum"] = 0.0
        return d
    
    max_oi_chg = max(d["ce_chng_oi"].abs().max(), d["pe_chng_oi"].abs().max(), 1)
    
    d["call_oi_accumulation"] = (d["ce_chng_oi"] / max_oi_chg) * 50 + 50
    d["put_oi_accumulation"] = (d["pe_chng_oi"] / max_oi_chg) * 50 + 50
    d["oi_momentum"] = (d["ce_chng_oi"] - d["pe_chng_oi"]) / max_oi_chg * 100
    d["oi_momentum"] = d["oi_momentum"].clip(-100, 100)
    
    return d


def calculate_delta_pressure(df: pd.DataFrame, spot: float) -> pd.DataFrame:
    """Calculate delta-weighted pressure."""
    d = df.copy()
    if d.empty or "ce_delta" not in d.columns or "pe_delta" not in d.columns:
        d["call_delta_exposure"] = 0.0
        d["put_delta_exposure"] = 0.0
        d["delta_imbalance"] = 0.0
        return d
    
    d["call_delta_exposure"] = d["ce_delta"] * d.get("ce_oi", 0)
    d["put_delta_exposure"] = d["pe_delta"] * d.get("pe_oi", 0)
    
    total_call_delta = d["call_delta_exposure"].sum()
    total_put_delta = d["put_delta_exposure"].sum()
    
    max_delta = max(abs(total_call_delta), abs(total_put_delta), 1)
    d["call_delta_pressure"] = _normalize_0_100(total_call_delta, -max_delta, max_delta)
    d["put_delta_pressure"] = _normalize_0_100(-total_put_delta, -max_delta, max_delta)
    
    d["delta_imbalance"] = (total_call_delta - total_put_delta) / max_delta * 100
    d["delta_imbalance"] = d["delta_imbalance"].clip(-100, 100)
    
    return d


def calculate_composite_pressure(df: pd.DataFrame, spot: float, lot_size: int = 1) -> pd.DataFrame:
    """Calculate composite buy/sell pressure."""
    d = df.copy()
    
    if d.empty:
        d["buy_pressure"] = 0.0
        d["sell_pressure"] = 0.0
        d["net_pressure"] = 0.0
        d["pressure_direction"] = "NEUTRAL"
        d["aggression_level"] = 0.0
        return d
    
    for col in ["call_volume_pressure", "put_volume_pressure", "call_oi_accumulation",
                "put_oi_accumulation", "call_delta_pressure", "put_delta_pressure",
                "ce_volume", "pe_volume", "ce_chng_oi", "pe_chng_oi"]:
        if col not in d.columns:
            d[col] = 0.0
    
    if "ce_ltp" in d.columns and "ce_change" in d.columns:
        d["price_action"] = (d["ce_change"] / (d["ce_ltp"] + 1e-6)) * 100
        d["price_action"] = d["price_action"].clip(-50, 50)
    else:
        d["price_action"] = 0.0
    
    w_volume, w_oi, w_delta, w_price = 0.30, 0.35, 0.25, 0.10
    
    d["buy_pressure"] = (
        (d["call_volume_pressure"] * w_volume) +
        (d["call_oi_accumulation"] * w_oi) +
        (d["call_delta_pressure"] * w_delta) +
        (_normalize_0_100(d["price_action"], -50, 50) * w_price)
    )
    
    d["sell_pressure"] = (
        (d["put_volume_pressure"] * w_volume) +
        (d["put_oi_accumulation"] * w_oi) +
        (d["put_delta_pressure"] * w_delta) +
        (_normalize_0_100(-d["price_action"], -50, 50) * w_price)
    )
    
    d["net_pressure"] = d["buy_pressure"] - d["sell_pressure"]
    d["net_pressure"] = d["net_pressure"].clip(-100, 100)
    d["aggression_level"] = (d["buy_pressure"].abs() + d["sell_pressure"].abs()) / 2
    
    def _classify_pressure(net_p: float, aggr: float) -> str:
        if net_p > 60 and aggr > 70:
            return "STRONG BUY"
        elif net_p > 30:
            return "BUY"
        elif net_p > -30:
            return "NEUTRAL"
        elif net_p > -60:
            return "SELL"
        else:
            return "STRONG SELL" if aggr > 70 else "SELL"
    
    d["pressure_direction"] = d.apply(
        lambda row: _classify_pressure(row["net_pressure"], row["aggression_level"]),
        axis=1
    )
    
    return d


def detect_pressure_anomalies(df: pd.DataFrame) -> pd.DataFrame:
    """Detect unusual volume/OI spikes."""
    d = df.copy()
    
    if d.empty or "ce_volume" not in d.columns:
        d["volume_spike"] = False
        d["oi_surge"] = False
        d["unusual_activity"] = False
        return d
    
    vol_q75 = (df["ce_volume"] + df["pe_volume"]).quantile(0.75)
    d["total_volume"] = d["ce_volume"] + d["pe_volume"]
    d["volume_spike"] = d["total_volume"] > (vol_q75 * 1.5)
    
    if "ce_chng_oi" in d.columns and "pe_chng_oi" in d.columns:
        oi_q75 = (d["ce_chng_oi"].abs() + d["pe_chng_oi"].abs()).quantile(0.75)
        d["oi_surge"] = (d["ce_chng_oi"].abs() + d["pe_chng_oi"].abs()) > (oi_q75 * 1.5)
    else:
        d["oi_surge"] = False
    
    d["unusual_activity"] = d["volume_spike"] | d["oi_surge"]
    
    return d


def calculate_market_pressure_summary(df: pd.DataFrame, spot: float, pcr: float) -> MarketPressure:
    """Calculate market-wide pressure summary."""
    if df.empty:
        return MarketPressure()
    
    total_call_pressure = df["buy_pressure"].mean() if "buy_pressure" in df.columns else 50
    total_put_pressure = df["sell_pressure"].mean() if "sell_pressure" in df.columns else 50
    net_bias = total_call_pressure - total_put_pressure
    
    if net_bias > 40:
        sentiment = "EXTREME BULLISH" if net_bias > 60 else "BULLISH"
    elif net_bias < -40:
        sentiment = "EXTREME BEARISH" if net_bias < -60 else "BEARISH"
    else:
        sentiment = "NEUTRAL"
    
    if pcr > 1.3 and net_bias < -20:
        pcr_corr = "⚠️ Divergence"
    elif pcr < 0.7 and net_bias > 20:
        pcr_corr = "✓ Aligned"
    else:
        pcr_corr = "Neutral"
    
    volume_surge = df["volume_spike"].sum() > len(df) * 0.15 if "volume_spike" in df.columns else False
    oi_accum = (df["ce_chng_oi"].sum() > 0 if "ce_chng_oi" in df.columns else False)
    
    atm_strikes = df[df.get("ATM", False)]
    itm_ce = df[df.get("CE Moneyness", "") == "ITM"]
    otm_ce = df[df.get("CE Moneyness", "") == "OTM"]
    
    itm_pressure = itm_ce["buy_pressure"].mean() if not itm_ce.empty and "buy_pressure" in itm_ce.columns else 50
    atm_pressure = atm_strikes["buy_pressure"].mean() if not atm_strikes.empty and "buy_pressure" in atm_strikes.columns else 50
    otm_pressure = otm_ce["buy_pressure"].mean() if not otm_ce.empty and "buy_pressure" in otm_ce.columns else 50
    
    return MarketPressure(
        total_call_pressure=round(total_call_pressure, 2),
        total_put_pressure=round(total_put_pressure, 2),
        net_market_bias=round(net_bias, 2),
        market_sentiment=sentiment,
        pcr_vs_pressure=pcr_corr,
        volume_surge_detected=volume_surge,
        oi_accumulation_detected=oi_accum,
        itm_pressure=round(itm_pressure, 2),
        atm_pressure=round(atm_pressure, 2),
        otm_pressure=round(otm_pressure, 2),
    )



def add_strike_movement_score(df: pd.DataFrame) -> pd.DataFrame:
    """Estimate which option side/strike has the strongest near-term movement potential.

    This is a ranking model, not a price prediction. It combines activity, OI change,
    buy/sell pressure, aggression, anomaly flags and existing AI confidence.
    """
    d = df.copy()
    if d.empty:
        for c, default in {
            "movement_score": 0.0,
            "movement_bias": "NEUTRAL",
            "movement_strength": "LOW",
            "ce_movement_score": 0.0,
            "pe_movement_score": 0.0,
        }.items():
            d[c] = default
        return d

    def pct_rank(series: pd.Series) -> pd.Series:
        s = pd.to_numeric(series, errors="coerce").fillna(0.0)
        if s.nunique() <= 1:
            return pd.Series(50.0, index=s.index)
        return (s.rank(pct=True) * 100).clip(0, 100)

    ce_vol = pct_rank(d.get("ce_volume", pd.Series(0.0, index=d.index)))
    pe_vol = pct_rank(d.get("pe_volume", pd.Series(0.0, index=d.index)))
    ce_oi = pct_rank(d.get("ce_chng_oi", pd.Series(0.0, index=d.index)).abs())
    pe_oi = pct_rank(d.get("pe_chng_oi", pd.Series(0.0, index=d.index)).abs())

    ce_pressure = pd.to_numeric(d.get("buy_pressure", 50.0), errors="coerce").fillna(50.0).clip(0, 100)
    pe_pressure = pd.to_numeric(d.get("sell_pressure", 50.0), errors="coerce").fillna(50.0).clip(0, 100)
    aggression = pd.to_numeric(d.get("aggression_level", 50.0), errors="coerce").fillna(50.0).clip(0, 100)

    total_vol = pd.to_numeric(d.get("total_volume", 0.0), errors="coerce").fillna(0.0)
    activity = pct_rank(total_vol)
    oi_activity = pct_rank(
        pd.to_numeric(d.get("ce_chng_oi", 0.0), errors="coerce").abs().fillna(0.0)
        + pd.to_numeric(d.get("pe_chng_oi", 0.0), errors="coerce").abs().fillna(0.0)
    )
    anomaly = (
        d.get("volume_spike", pd.Series(False, index=d.index)).astype(bool).astype(float) * 60
        + d.get("oi_surge", pd.Series(False, index=d.index)).astype(bool).astype(float) * 40
    ).clip(0, 100)

    ai_conf = pd.to_numeric(d.get("AI Confidence %", 50.0), errors="coerce").fillna(50.0).clip(0, 100)
    ai_signal = d.get("AI Signal", pd.Series("", index=d.index)).astype(str).str.upper()

    ce_side = (
        ce_vol * 0.20
        + ce_oi * 0.15
        + ce_pressure * 0.35
        + aggression * 0.15
        + anomaly * 0.10
        + ai_conf.where(ai_signal.str.contains("CE|CALL|BUY"), 50.0) * 0.05
    )
    pe_side = (
        pe_vol * 0.20
        + pe_oi * 0.15
        + pe_pressure * 0.35
        + aggression * 0.15
        + anomaly * 0.10
        + ai_conf.where(ai_signal.str.contains("PE|PUT|SELL"), 50.0) * 0.05
    )

    d["ce_movement_score"] = ce_side.clip(0, 100).round(1)
    d["pe_movement_score"] = pe_side.clip(0, 100).round(1)
    d["movement_score"] = pd.concat([ce_side, pe_side], axis=1).max(axis=1).clip(0, 100).round(1)

    def bias(row: pd.Series) -> str:
        ce, pe = float(row["ce_movement_score"]), float(row["pe_movement_score"])
        if max(ce, pe) < 55:
            return "NEUTRAL"
        if abs(ce - pe) < 7:
            return "BOTH / CHOP"
        return "CE UP" if ce > pe else "PE UP"

    d["movement_bias"] = d.apply(bias, axis=1)
    d["movement_strength"] = pd.cut(
        d["movement_score"], bins=[-1, 45, 60, 75, 100],
        labels=["LOW", "MODERATE", "STRONG", "VERY STRONG"]
    ).astype(str)

    # A combined ranking that prioritizes the strongest activity around ATM.
    d["movement_rank"] = d["movement_score"].rank(method="min", ascending=False).astype(int)
    return d


def _movement_history_key(symbol: str, expiry_label: str, strike: float) -> str:
    return f"{symbol}|{expiry_label}|{float(strike):.4f}"


def compute_movement_early_warning(
    df: pd.DataFrame, symbol: str, expiry_label: str, spot: float
) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Track movement-score acceleration across live scans.

    This adds a *build-up* layer to the existing movement ranking. It does not
    predict an exact future price. It looks for repeated score improvement,
    directional pressure and option activity before the current movement score
    becomes extreme.
    """
    d = df.copy()
    default_summary = {
        "status": "WAIT", "direction": "NEUTRAL", "strike": None,
        "early_score": 0.0, "current_score": 0.0, "score_delta": 0.0,
        "score_acceleration": 0.0, "rising_scans": 0, "confidence": 0.0,
        "volume": 0.0, "oi_change": 0.0, "pressure": 50.0,
        "reason": "Waiting for repeated scans"
    }
    if d.empty or "movement_score" not in d.columns:
        for c, default in {
            "early_movement_score": 0.0,
            "movement_score_delta": 0.0,
            "movement_score_acceleration": 0.0,
            "movement_rising_scans": 0,
            "early_movement_status": "WAIT",
            "early_movement_confidence": 0.0,
        }.items():
            d[c] = default
        return d, default_summary

    history = st.session_state.setdefault(MOVEMENT_HISTORY_KEY, {})
    now = datetime.now()

    # Save the current snapshot strike-by-strike.
    for _, row in d.iterrows():
        strike = float(row.get("strike_price", 0) or 0)
        if strike <= 0:
            continue
        key = _movement_history_key(symbol, expiry_label, strike)
        series = history.get(key, [])
        series.append({
            "ts": now,
            "score": float(row.get("movement_score", 0) or 0),
            "ce_score": float(row.get("ce_movement_score", 0) or 0),
            "pe_score": float(row.get("pe_movement_score", 0) or 0),
            "buy": float(row.get("buy_pressure", 50) or 50),
            "sell": float(row.get("sell_pressure", 50) or 50),
            "volume": float(row.get("total_volume", 0) or 0),
            "oi": float(abs(row.get("ce_chng_oi", 0) or 0) + abs(row.get("pe_chng_oi", 0) or 0)),
        })
        history[key] = series[-MOVEMENT_HISTORY_MAX:]

    st.session_state[MOVEMENT_HISTORY_KEY] = history

    early_scores = []
    deltas = []
    accels = []
    rising_counts = []
    statuses = []
    confidences = []

    for _, row in d.iterrows():
        strike = float(row.get("strike_price", 0) or 0)
        key = _movement_history_key(symbol, expiry_label, strike)
        series = history.get(key, [])
        scores = [float(x.get("score", 0)) for x in series]
        current = scores[-1] if scores else float(row.get("movement_score", 0) or 0)

        delta = scores[-1] - scores[-2] if len(scores) >= 2 else 0.0
        prev_delta = scores[-2] - scores[-3] if len(scores) >= 3 else 0.0
        accel = delta - prev_delta if len(scores) >= 3 else 0.0

        rising = 0
        for i in range(len(scores) - 1, 0, -1):
            if scores[i] > scores[i - 1] + 0.5:
                rising += 1
            else:
                break

        ce = float(row.get("ce_movement_score", 0) or 0)
        pe = float(row.get("pe_movement_score", 0) or 0)
        directional_gap = abs(ce - pe)
        pressure_gap = abs(float(row.get("buy_pressure", 50) or 50) - float(row.get("sell_pressure", 50) or 50))

        # Build-up score: current activity + improving score + acceleration + direction.
        delta_component = float(np.clip(50.0 + delta * 2.5, 0, 100))
        accel_component = float(np.clip(50.0 + accel * 4.0, 0, 100))
        direction_component = float(np.clip(50.0 + directional_gap * 1.2 + pressure_gap * 0.25, 0, 100))
        early = float(np.clip(
            current * 0.50 + delta_component * 0.20 + accel_component * 0.10 + direction_component * 0.20,
            0, 100
        ))

        if rising >= MOVEMENT_MIN_RISING_SCANS and early >= MOVEMENT_STRONG_THRESHOLD:
            status = "STRONG MOVE"
        elif rising >= MOVEMENT_MIN_RISING_SCANS and early >= MOVEMENT_EARLY_THRESHOLD:
            status = "EARLY MOVE"
        elif rising >= 1 and delta > 0:
            status = "BUILDING"
        else:
            status = "WAIT"

        confidence = float(np.clip(
            35 + early * 0.45 + min(rising, 4) * 5 + min(directional_gap, 30) * 0.2,
            0, 95
        ))

        early_scores.append(round(early, 1))
        deltas.append(round(delta, 1))
        accels.append(round(accel, 1))
        rising_counts.append(rising)
        statuses.append(status)
        confidences.append(round(confidence, 1))

    d["early_movement_score"] = early_scores
    d["movement_score_delta"] = deltas
    d["movement_score_acceleration"] = accels
    d["movement_rising_scans"] = rising_counts
    d["early_movement_status"] = statuses
    d["early_movement_confidence"] = confidences

    # Pick the best candidate only after it has some evidence of building.
    candidates = d[d["early_movement_status"].isin(["BUILDING", "EARLY MOVE", "STRONG MOVE"])].copy()
    if candidates.empty:
        candidates = d.sort_values("early_movement_score", ascending=False).head(1)
    if candidates.empty:
        return d, default_summary

    best = candidates.sort_values(
        ["early_movement_score", "movement_score_delta"], ascending=False
    ).iloc[0]
    ce = float(best.get("ce_movement_score", 0) or 0)
    pe = float(best.get("pe_movement_score", 0) or 0)
    direction = "CE / UP" if ce > pe + 7 else ("PE / DOWN" if pe > ce + 7 else "BOTH / CHOP")
    reasons = []
    if float(best.get("movement_score_delta", 0) or 0) > 0:
        reasons.append(f"score +{float(best['movement_score_delta']):.1f}")
    if float(best.get("movement_score_acceleration", 0) or 0) > 0:
        reasons.append("acceleration positive")
    if int(best.get("movement_rising_scans", 0) or 0) >= 2:
        reasons.append(f"{int(best['movement_rising_scans'])} rising scans")
    if bool(best.get("volume_spike", False)):
        reasons.append("volume spike")
    if bool(best.get("oi_surge", False)):
        reasons.append("OI surge")
    if abs(float(best.get("buy_pressure", 50) or 50) - float(best.get("sell_pressure", 50) or 50)) >= 15:
        reasons.append("pressure imbalance")

    return d, {
        "status": str(best.get("early_movement_status", "WAIT")),
        "direction": direction,
        "strike": float(best.get("strike_price", 0) or 0),
        "early_score": float(best.get("early_movement_score", 0) or 0),
        "current_score": float(best.get("movement_score", 0) or 0),
        "score_delta": float(best.get("movement_score_delta", 0) or 0),
        "score_acceleration": float(best.get("movement_score_acceleration", 0) or 0),
        "rising_scans": int(best.get("movement_rising_scans", 0) or 0),
        "confidence": float(best.get("early_movement_confidence", 0) or 0),
        "volume": float(best.get("total_volume", 0) or 0),
        "oi_change": float(abs(best.get("ce_chng_oi", 0) or 0) + abs(best.get("pe_chng_oi", 0) or 0)),
        "pressure": max(float(best.get("buy_pressure", 50) or 50), float(best.get("sell_pressure", 50) or 50)),
        "reason": ", ".join(reasons) if reasons else "Activity building",
    }


def _render_movement_early_warning(early: dict[str, Any]) -> None:
    st.markdown('<div class="block-title">🚨 MOVEMENT BEFORE IT HAPPENS — EARLY WARNING</div>', unsafe_allow_html=True)
    if not early:
        st.info("Waiting for live movement history.")
        return

    status = str(early.get("status", "WAIT"))
    direction = str(early.get("direction", "NEUTRAL"))
    score = float(early.get("early_score", 0) or 0)
    confidence = float(early.get("confidence", 0) or 0)

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("STATUS", status)
    c2.metric("DIRECTION", direction)
    c3.metric("EARLY SCORE", f"{score:.0f}/100")
    c4.metric("CONFIDENCE", f"{confidence:.0f}%")
    c5.metric("RISING SCANS", str(int(early.get("rising_scans", 0) or 0)))

    c6, c7, c8, c9 = st.columns(4)
    c6.metric("STRIKE", f"{float(early.get('strike', 0) or 0):,.0f}" if early.get("strike") else "—")
    c7.metric("CURRENT MOVE", f"{float(early.get('current_score', 0) or 0):.0f}")
    c8.metric("SCORE Δ", f"{float(early.get('score_delta', 0) or 0):+.1f}")
    c9.metric("ACCELERATION", f"{float(early.get('score_acceleration', 0) or 0):+.1f}")

    if status == "STRONG MOVE":
        st.error(f"🔴 STRONG MOVE BUILDING — {direction} near {float(early.get('strike', 0) or 0):,.0f}")
    elif status == "EARLY MOVE":
        st.warning(f"🟠 EARLY MOVE — {direction} near {float(early.get('strike', 0) or 0):,.0f}")
    elif status == "BUILDING":
        st.info(f"🟡 BUILDING — {direction} near {float(early.get('strike', 0) or 0):,.0f}")
    else:
        st.success("🟢 No confirmed build-up yet. Continue monitoring repeated scans.")

    st.caption(
        f"Reasons: {early.get('reason', '—')} | "
        f"OI activity: {float(early.get('oi_change', 0) or 0):,.0f} | "
        f"Pressure: {float(early.get('pressure', 50) or 50):.0f}"
    )
    st.caption("⚠️ Early Warning is probabilistic: it detects rising activity/pressure across scans; it cannot guarantee the next move.")


def add_pressure_analysis(df: pd.DataFrame, spot: float, lot_size: int = 1) -> tuple[pd.DataFrame, MarketPressure]:
    """Full pressure analysis pipeline."""
    d = df.copy()
    
    d = calculate_volume_pressure(d)
    d = calculate_oi_change_pressure(d)
    d = calculate_delta_pressure(d, spot)
    d = calculate_composite_pressure(d, spot, lot_size)
    d = detect_pressure_anomalies(d)
    
    pcr = d["pe_oi"].sum() / d["ce_oi"].sum() if d["ce_oi"].sum() > 0 else 1.0
    market_pressure = calculate_market_pressure_summary(d, spot, pcr)
    
    return d, market_pressure


# ══════════════════════════════════════════════════════════════════════════
# 16. CHARTS (ORIGINAL - UNMODIFIED + NEW PRESSURE CHARTS)
# ══════════════════════════════════════════════════════════════════════════

def _plotly_dark_layout(fig: go.Figure, height: int = 420, title: str = "") -> go.Figure:
    fig.update_layout(
        paper_bgcolor=DARK_BG, plot_bgcolor=DARK_BG,
        font=dict(color=TEXT_MUTED, family="Courier New"),
        height=height, margin=dict(l=10, r=10, t=40 if title else 10, b=10),
        title=dict(text=title, font=dict(color=TEXT_MAIN, size=14)) if title else None,
        legend=dict(bgcolor=PANEL_BG, bordercolor=BORDER_COLOR, borderwidth=1),
    )
    return fig


def chart_oi_bars(df: pd.DataFrame, max_pain: float) -> go.Figure:
    fig = make_subplots(rows=1, cols=2, subplot_titles=("Call OI (CE)", "Put OI (PE)"),
                         shared_yaxes=True, horizontal_spacing=0.04)
    if df.empty:
        return _plotly_dark_layout(fig)
    max_oi = max(df["ce_oi"].max(), df["pe_oi"].max(), 1)
    strikes_sorted = df["strike_price"].sort_values().unique()
    gap = (strikes_sorted[1] - strikes_sorted[0]) if len(strikes_sorted) > 1 else 1

    fig.add_trace(go.Bar(
        x=-df["ce_oi"], y=df["strike_price"], orientation="h",
        marker_color=[GREEN if abs(s - max_pain) < gap / 2 else "#238636" for s in df["strike_price"]],
        name="CE OI", showlegend=False,
        hovertemplate="Strike %{y}<br>CE OI: %{customdata:,}<extra></extra>", customdata=df["ce_oi"],
    ), row=1, col=1)
    fig.add_trace(go.Bar(
        x=df["pe_oi"], y=df["strike_price"], orientation="h",
        marker_color=[RED if abs(s - max_pain) < gap / 2 else "#da3633" for s in df["strike_price"]],
        name="PE OI", showlegend=False,
        hovertemplate="Strike %{y}<br>PE OI: %{x:,}<extra></extra>",
    ), row=1, col=2)
    for col in (1, 2):
        fig.add_hline(y=max_pain, line_dash="dot", line_color=AMBER,
                      annotation_text=f"Max Pain {max_pain:,.0f}", annotation_font_color=AMBER, row=1, col=col)
    fig.update_layout(
        xaxis=dict(showticklabels=False, showgrid=False, range=[-max_oi * 1.1, 0]),
        xaxis2=dict(showticklabels=False, showgrid=False, range=[0, max_oi * 1.1]),
        yaxis=dict(showgrid=True, gridcolor=BORDER_COLOR, tickfont=dict(color=TEXT_MAIN, size=11)),
    )
    fig.update_annotations(font_color=TEXT_MUTED)
    return _plotly_dark_layout(fig, height=480)


def chart_iv_skew(df: pd.DataFrame) -> go.Figure:
    fig = go.Figure()
    if not df.empty:
        fig.add_trace(go.Scatter(x=df["strike_price"], y=df["ce_iv"], mode="lines+markers",
                                  name="CE IV", line=dict(color=GREEN, width=2)))
        fig.add_trace(go.Scatter(x=df["strike_price"], y=df["pe_iv"], mode="lines+markers",
                                  name="PE IV", line=dict(color=RED, width=2)))
    fig.update_layout(xaxis=dict(title="Strike", showgrid=True, gridcolor=BORDER_COLOR),
                       yaxis=dict(title="IV %", showgrid=True, gridcolor=BORDER_COLOR))
    return _plotly_dark_layout(fig, height=320, title="Implied Volatility Skew")


def chart_greeks(df: pd.DataFrame, greek: str) -> go.Figure:
    fig = go.Figure()
    col_ce, col_pe = f"ce_{greek}", f"pe_{greek}"
    if not df.empty and col_ce in df.columns:
        fig.add_trace(go.Scatter(x=df["strike_price"], y=df[col_ce], mode="lines+markers",
                                  name=f"CE {greek.title()}", line=dict(color=GREEN, width=2)))
        fig.add_trace(go.Scatter(x=df["strike_price"], y=df[col_pe], mode="lines+markers",
                                  name=f"PE {greek.title()}", line=dict(color=RED, width=2)))
    fig.update_layout(xaxis=dict(title="Strike", showgrid=True, gridcolor=BORDER_COLOR),
                       yaxis=dict(title=greek.title(), showgrid=True, gridcolor=BORDER_COLOR))
    return _plotly_dark_layout(fig, height=300, title=f"{greek.title()} by Strike")


def chart_movement_score(df: pd.DataFrame) -> go.Figure:
    fig = go.Figure()
    if df.empty or "movement_score" not in df.columns:
        return _plotly_dark_layout(fig, height=360, title="Strike Movement Score")
    d = df.sort_values("movement_score", ascending=False).head(20).sort_values("strike_price")
    fig.add_trace(go.Bar(
        x=d["strike_price"], y=d["movement_score"],
        text=d["movement_bias"], textposition="outside",
        marker_color=[GREEN if x >= 75 else (AMBER if x >= 60 else RED) for x in d["movement_score"]],
        name="Movement Score",
        hovertemplate="Strike %{x:,.0f}<br>Score: %{y:.1f}<br>Bias: %{text}<extra></extra>",
    ))
    fig.update_layout(
        xaxis=dict(title="Strike", showgrid=True, gridcolor=BORDER_COLOR),
        yaxis=dict(title="Movement Score (0-100)", range=[0, 105], showgrid=True, gridcolor=BORDER_COLOR),
    )
    return _plotly_dark_layout(fig, height=380, title="Top Strike Movement Potential")


def chart_gex_by_strike(gex_data: dict) -> go.Figure:
    fig = go.Figure()
    by_strike = gex_data.get("by_strike", pd.DataFrame())
    if not by_strike.empty:
        colors = [GREEN if v >= 0 else RED for v in by_strike["gex"]]
        fig.add_trace(go.Bar(x=by_strike["strike_price"], y=by_strike["gex"], marker_color=colors, name="GEX"))
    fig.update_layout(xaxis=dict(title="Strike", showgrid=True, gridcolor=BORDER_COLOR),
                       yaxis=dict(title="Gamma Exposure", showgrid=True, gridcolor=BORDER_COLOR))
    return _plotly_dark_layout(fig, height=320, title="Gamma Exposure (GEX) by Strike")


def chart_price_action(df: pd.DataFrame, title: str = "Price Action with Indicators") -> go.Figure:
    """Chart price action with VWAP, EMA, and volume."""
    if df.empty or "close" not in df.columns:
        return _plotly_dark_layout(go.Figure(), title=title)
    
    fig = make_subplots(
        rows=2, cols=1, shared_xaxes=True, vertical_spacing=0.1,
        row_heights=[0.7, 0.3],
        subplot_titles=("Price", "Volume")
    )
    
    fig.add_trace(go.Candlestick(
        x=df.index, open=df["open"], high=df["high"], low=df["low"], close=df["close"],
        name="OHLC", showlegend=True
    ), row=1, col=1)
    
    if "vwap" in df.columns:
        fig.add_trace(go.Scatter(
            x=df.index, y=df["vwap"], mode="lines", name="VWAP",
            line=dict(color=BLUE, width=2)
        ), row=1, col=1)
    
    if "ema_9" in df.columns:
        fig.add_trace(go.Scatter(
            x=df.index, y=df["ema_9"], mode="lines", name="EMA 9",
            line=dict(color=GREEN, width=1.5)
        ), row=1, col=1)
    if "ema_21" in df.columns:
        fig.add_trace(go.Scatter(
            x=df.index, y=df["ema_21"], mode="lines", name="EMA 21",
            line=dict(color=RED, width=1.5)
        ), row=1, col=1)
    
    if "volume" in df.columns:
        colors = [GREEN if df["close"].iloc[i] >= df["open"].iloc[i] else RED for i in range(len(df))]
        fig.add_trace(go.Bar(
            x=df.index, y=df["volume"], name="Volume", marker_color=colors, showlegend=True
        ), row=2, col=1)
    
    fig.update_layout(
        xaxis=dict(showgrid=True, gridcolor=BORDER_COLOR),
        yaxis=dict(showgrid=True, gridcolor=BORDER_COLOR),
        xaxis2=dict(showgrid=True, gridcolor=BORDER_COLOR),
        yaxis2=dict(showgrid=True, gridcolor=BORDER_COLOR),
    )
    
    return _plotly_dark_layout(fig, height=500, title=title)


def chart_technical_indicators(df: pd.DataFrame) -> go.Figure:
    """Chart RSI, MACD, and Momentum."""
    if df.empty:
        return _plotly_dark_layout(go.Figure())
    
    fig = make_subplots(
        rows=3, cols=1, shared_xaxes=True, vertical_spacing=0.08,
        subplot_titles=("RSI (14)", "MACD", "Momentum")
    )
    
    if "rsi" in df.columns:
        fig.add_trace(go.Scatter(
            x=df.index, y=df["rsi"], mode="lines", name="RSI",
            line=dict(color=BLUE, width=2)
        ), row=1, col=1)
        fig.add_hline(y=70, line_dash="dash", line_color=RED, annotation_text="Overbought", row=1, col=1)
        fig.add_hline(y=30, line_dash="dash", line_color=GREEN, annotation_text="Oversold", row=1, col=1)
    
    if "macd" in df.columns and "macd_signal" in df.columns:
        fig.add_trace(go.Scatter(
            x=df.index, y=df["macd"], mode="lines", name="MACD",
            line=dict(color=BLUE, width=2)
        ), row=2, col=1)
        fig.add_trace(go.Scatter(
            x=df.index, y=df["macd_signal"], mode="lines", name="MACD Signal",
            line=dict(color=RED, width=1.5)
        ), row=2, col=1)
        if "macd_hist" in df.columns:
            colors = [GREEN if v >= 0 else RED for v in df["macd_hist"]]
            fig.add_trace(go.Bar(
                x=df.index, y=df["macd_hist"], name="MACD Histogram",
                marker_color=colors
            ), row=2, col=1)
    
    if len(df) > 1:
        momentum = df["close"].pct_change() * 100
        fig.add_trace(go.Scatter(
            x=df.index, y=momentum, mode="lines", name="Momentum %",
            line=dict(color=AMBER, width=2)
        ), row=3, col=1)
        fig.add_hline(y=0, line_dash="dash", line_color=TEXT_MUTED, row=3, col=1)
    
    fig.update_yaxes(title_text="RSI", row=1, col=1)
    fig.update_yaxes(title_text="MACD", row=2, col=1)
    fig.update_yaxes(title_text="Momentum %", row=3, col=1)
    
    fig.update_layout(
        xaxis=dict(showgrid=True, gridcolor=BORDER_COLOR),
        yaxis=dict(showgrid=True, gridcolor=BORDER_COLOR),
    )
    
    return _plotly_dark_layout(fig, height=600)


def gauge_pcr(pcr: float) -> go.Figure:
    fig = go.Figure(go.Indicator(
        mode="gauge+number", value=pcr,
        number={"font": {"color": TEXT_MAIN, "size": 32, "family": "Courier New"}},
        gauge={
            "axis": {"range": [0, 3], "tickcolor": TEXT_MUTED, "tickfont": {"color": TEXT_MUTED}},
            "bar": {"color": BLUE, "thickness": 0.25}, "bgcolor": PANEL_BG, "borderwidth": 0,
            "steps": [{"range": [0, 0.7], "color": "#3b0d1a"}, {"range": [0.7, 1.3], "color": "#1c2128"},
                      {"range": [1.3, 3.0], "color": "#0d3b2e"}],
            "threshold": {"line": {"color": AMBER, "width": 3}, "value": pcr},
        },
        title={"text": "PUT / CALL RATIO", "font": {"color": TEXT_MUTED, "size": 12}},
    ))
    return _plotly_dark_layout(fig, height=220)


# ══════════════════════════════════════════════════════════════════════════
# NEW PRESSURE CHARTS
# ══════════════════════════════════════════════════════════════════════════

def chart_pressure_by_strike(df: pd.DataFrame, spot: float) -> go.Figure:
    """Chart buy/sell pressure by strike."""
    fig = go.Figure()
    
    if df.empty or "buy_pressure" not in df.columns:
        return _plotly_dark_layout(fig, title="Buy/Sell Pressure by Strike")
    
    df_sorted = df.sort_values("strike_price")
    
    fig.add_trace(go.Bar(
        x=df_sorted["buy_pressure"],
        y=df_sorted["strike_price"],
        orientation="h",
        name="Buy Pressure",
        marker_color=GREEN,
        hovertemplate="Strike %{y}<br>Buy: %{x:.0f}<extra></extra>",
    ))
    
    fig.add_trace(go.Bar(
        x=-df_sorted["sell_pressure"],
        y=df_sorted["strike_price"],
        orientation="h",
        name="Sell Pressure",
        marker_color=RED,
        hovertemplate="Strike %{y}<br>Sell: %{customdata:.0f}<extra></extra>",
        customdata=df_sorted["sell_pressure"],
    ))
    
    if spot:
        fig.add_hline(y=spot, line_dash="dash", line_color=BLUE, 
                     annotation_text=f"Spot {spot:,.0f}",
                     annotation_font_color=BLUE)
    
    fig.update_layout(
        barmode="overlay",
        xaxis=dict(title="Pressure Score", showgrid=True, gridcolor=BORDER_COLOR),
        yaxis=dict(title="Strike", showgrid=True, gridcolor=BORDER_COLOR),
    )
    
    return _plotly_dark_layout(fig, height=500, title="Buy/Sell Pressure by Strike")


def chart_net_pressure(df: pd.DataFrame, spot: float) -> go.Figure:
    """Chart net pressure (directional)."""
    fig = go.Figure()
    
    if df.empty or "net_pressure" not in df.columns:
        return _plotly_dark_layout(fig, title="Net Pressure Bias")
    
    df_sorted = df.sort_values("strike_price")
    colors = [GREEN if x > 0 else RED for x in df_sorted["net_pressure"]]
    
    fig.add_trace(go.Bar(
        x=df_sorted["strike_price"],
        y=df_sorted["net_pressure"],
        marker_color=colors,
        name="Net Pressure",
        hovertemplate="Strike %{x:,.0f}<br>Net: %{y:.0f}<extra></extra>",
    ))
    
    fig.add_hline(y=0, line_dash="dash", line_color=TEXT_MUTED)
    
    if spot:
        fig.add_vline(x=spot, line_dash="dash", line_color=BLUE,
                     annotation_text=f"Spot {spot:,.0f}",
                     annotation_font_color=BLUE)
    
    fig.update_layout(
        xaxis=dict(title="Strike", showgrid=True, gridcolor=BORDER_COLOR),
        yaxis=dict(title="Net Pressure (-100 to +100)", showgrid=True, gridcolor=BORDER_COLOR),
    )
    
    return _plotly_dark_layout(fig, height=400, title="Net Pressure Bias by Strike")


def chart_aggression_level(df: pd.DataFrame) -> go.Figure:
    """Chart aggression level."""
    fig = go.Figure()
    
    if df.empty or "aggression_level" not in df.columns:
        return _plotly_dark_layout(fig, title="Aggression Level")
    
    df_sorted = df.sort_values("strike_price")
    
    fig.add_trace(go.Scatter(
        x=df_sorted["strike_price"],
        y=df_sorted["aggression_level"],
        mode="lines+markers",
        name="Aggression",
        line=dict(color=AMBER, width=2),
        fill="tozeroy",
        fillcolor=f"rgba(210, 153, 34, 0.2)",
        hovertemplate="Strike %{x:,.0f}<br>Aggression: %{y:.0f}<extra></extra>",
    ))
    
    fig.update_layout(
        xaxis=dict(title="Strike", showgrid=True, gridcolor=BORDER_COLOR),
        yaxis=dict(title="Aggression Level (0-100)", showgrid=True, gridcolor=BORDER_COLOR, range=[0, 100]),
    )
    
    return _plotly_dark_layout(fig, height=350, title="Aggression Level by Strike")


# ══════════════════════════════════════════════════════════════════════════
# 17. HTML TABLE RENDERING (ORIGINAL - UNMODIFIED)
# ══════════════════════════════════════════════════════════════════════════

_TABLE_CSS = f"""
<style>
.oc-table-wrap {{ max-height: 620px; overflow-y: auto; border: 1px solid {BORDER_COLOR}; border-radius: 8px; }}
.oc-table {{ width: 100%; border-collapse: collapse; font-family: 'Courier New', monospace; font-size: 12.5px; }}
.oc-table th {{ background: #1F4E78; color: #ffffff; padding: 8px 10px; text-align: center;
                position: sticky; top: 0; font-size: 11px; text-transform: uppercase; letter-spacing: .04em; }}
.oc-table td {{ padding: 6px 9px; text-align: center; border-bottom: 1px solid #21262d; color: {TEXT_MAIN}; white-space: nowrap; }}
.oc-atm-row td {{ background-color: #1c2128 !important; font-weight: 700; }}
</style>
"""


def _safe_cell(val: Any) -> str:
    if val is None:
        return ""
    try:
        if isinstance(val, float) and math.isnan(val):
            return ""
    except (TypeError, ValueError):
        pass
    s = str(val)
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


def _oi_cell_style(val: float, heavy_thresh: float, max_val: float) -> str:
    if max_val <= 0:
        return f"color:{TEXT_MUTED};"
    pct = max(0.0, min(100.0, (val / max_val) * 100))
    intensity = 0.10 + pct / 250
    is_heavy = heavy_thresh > 0 and val >= heavy_thresh
    bg = f"background:linear-gradient(90deg, rgba(63,185,80,{intensity:.2f}) {pct:.0f}%, transparent {pct:.0f}%);"
    weight = "font-weight:700;" if is_heavy else ""
    return bg + weight


def _oi_change_cell_style(val: float, heavy_thresh: float) -> str:
    if val == 0:
        return f"color:{TEXT_MUTED};"
    is_large = heavy_thresh > 0 and abs(val) >= heavy_thresh
    if val > 0:
        return f"color:#0d3b2e;font-weight:700;background-color:{GREEN};" if is_large else f"color:{GREEN};"
    return f"color:#3b0d1a;font-weight:700;background-color:{RED};" if is_large else f"color:{RED};"


def _signal_cell_style(val: str) -> str:
    v = str(val).upper()
    if "BUY CE" in v or "STRONG BUY" in v or "BUY" in v:
        return f"color:{GREEN};font-weight:700;"
    if "BUY PE" in v or "SELL" in v:
        return f"color:{RED};font-weight:700;"
    if "HOLD" in v or "WAIT" in v:
        return f"color:{AMBER};font-weight:700;"
    return f"color:{TEXT_MUTED};"


def render_chain_table_html(df: pd.DataFrame, show_greeks: bool, top_n: int = 400) -> str:
    if df.empty:
        return _TABLE_CSS + "<div style='color:#8b949e;padding:12px;'>No rows to display.</div>"

    base_cols = [
        ("ce_oi", "CE OI"), ("ce_chng_oi", "CE ΔOI"), ("ce_oi_change_pct", "CE ΔOI%"),
        ("ce_volume", "CE Vol"), ("ce_iv", "CE IV"), ("ce_ltp", "CE LTP"),
        ("ce_bid", "CE Bid"), ("ce_ask", "CE Ask"),
    ]
    greek_ce_cols = [("ce_delta", "CE Δ"), ("ce_gamma", "CE Γ"), ("ce_theta", "CE Θ"), ("ce_vega", "CE V")]
    mid_cols = [("strike_price", "STRIKE"), ("CE Buildup", "CE Build"), ("PE Buildup", "PE Build"),
                ("AI Signal", "AI Signal"), ("movement_score", "MOVE %"),
                ("movement_bias", "MOVE BIAS"), ("movement_strength", "MOVE STR")]
    greek_pe_cols = [("pe_delta", "PE Δ"), ("pe_gamma", "PE Γ"), ("pe_theta", "PE Θ"), ("pe_vega", "PE V")]
    pe_cols = [
        ("pe_bid", "PE Bid"), ("pe_ask", "PE Ask"), ("pe_ltp", "PE LTP"), ("pe_iv", "PE IV"),
        ("pe_volume", "PE Vol"), ("pe_oi_change_pct", "PE ΔOI%"), ("pe_chng_oi", "PE ΔOI"), ("pe_oi", "PE OI"),
    ]

    cols = base_cols + (greek_ce_cols if show_greeks else []) + mid_cols + \
        (greek_pe_cols if show_greeks else []) + pe_cols
    cols = [(k, label) for k, label in cols if k in df.columns]

    fmt = {
        "ce_oi": "{:,.0f}", "ce_chng_oi": "{:+,.0f}", "ce_oi_change_pct": "{:+.1f}%",
        "ce_volume": "{:,.0f}", "ce_iv": "{:.1f}", "ce_ltp": "{:.2f}", "ce_bid": "{:.2f}", "ce_ask": "{:.2f}",
        "ce_delta": "{:.3f}", "ce_gamma": "{:.5f}", "ce_theta": "{:.3f}", "ce_vega": "{:.3f}",
        "strike_price": "{:,.0f}", "movement_score": "{:.1f}",
        "pe_delta": "{:.3f}", "pe_gamma": "{:.5f}", "pe_theta": "{:.3f}", "pe_vega": "{:.3f}",
        "pe_bid": "{:.2f}", "pe_ask": "{:.2f}", "pe_ltp": "{:.2f}", "pe_iv": "{:.1f}",
        "pe_volume": "{:,.0f}", "pe_oi_change_pct": "{:+.1f}%", "pe_chng_oi": "{:+,.0f}", "pe_oi": "{:,.0f}",
    }

    heavy_ce_oi = df["ce_oi"].quantile(0.80) if df["ce_oi"].max() > 0 else 0
    heavy_pe_oi = df["pe_oi"].quantile(0.80) if df["pe_oi"].max() > 0 else 0
    heavy_ce_chng = df["ce_chng_oi"].abs().quantile(0.80) if (df["ce_chng_oi"] != 0).any() else 0
    heavy_pe_chng = df["pe_chng_oi"].abs().quantile(0.80) if (df["pe_chng_oi"] != 0).any() else 0
    max_ce_oi, max_pe_oi = df["ce_oi"].max(), df["pe_oi"].max()

    view = df.head(top_n)
    header_html = "".join(f"<th>{label}</th>" for _, label in cols)
    rows_html = []
    for _, row in view.iterrows():
        is_atm = bool(row.get("ATM", False))
        cells = []
        for key, _ in cols:
            val = row.get(key, "")
            spec = fmt.get(key)
            display_val = spec.format(val) if spec and pd.notna(val) else ("" if pd.isna(val) else val)
            style = ""
            if key == "ce_oi":
                style = _oi_cell_style(val, heavy_ce_oi, max_ce_oi)
            elif key == "pe_oi":
                style = _oi_cell_style(val, heavy_pe_oi, max_pe_oi)
            elif key == "ce_chng_oi":
                style = _oi_change_cell_style(val, heavy_ce_chng)
            elif key == "pe_chng_oi":
                style = _oi_change_cell_style(val, heavy_pe_chng)
            elif key == "AI Signal":
                style = _signal_cell_style(val)
            elif key == "movement_bias":
                style = _signal_cell_style(val)
            elif key == "movement_score":
                try:
                    score = float(val)
                    style = f"font-weight:700;color:{GREEN if score >= 75 else (AMBER if score >= 60 else TEXT_MUTED)};"
                except Exception:
                    style = ""
            cells.append(f'<td style="{style}">{_safe_cell(display_val)}</td>')
        row_class = "oc-atm-row" if is_atm else ""
        rows_html.append(f'<tr class="{row_class}">{"".join(cells)}</tr>')

    return (
        _TABLE_CSS
        + f'<div class="oc-table-wrap"><table class="oc-table"><thead><tr>{header_html}</tr></thead>'
        + f'<tbody>{"".join(rows_html)}</tbody></table></div>'
    )


# ══════════════════════════════════════════════════════════════════════════
# 18. EXCEL EXPORT (ORIGINAL - UNMODIFIED)
# ══════════════════════════════════════════════════════════════════════════

FILL_HEADER = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_AMBER = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FONT_HEADER = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(*(Side(style="thin", color="30363D"),) * 4)


def _style_header_row(ws, row_idx: int = 1) -> None:
    for cell in ws[row_idx]:
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _autosize_columns(ws) -> None:
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length + 3, 10), 40)


def _apply_borders(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.border = THIN_BORDER


def _write_dataframe(ws, df: pd.DataFrame, start_row: int = 1) -> None:
    for j, col_name in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=j, value=str(col_name))
    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for j, val in enumerate(row, start=1):
            if isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = float(val) if not math.isnan(val) else None
            elif isinstance(val, (np.bool_,)):
                val = bool(val)
            ws.cell(row=i, column=j, value=val)
    _style_header_row(ws, start_row)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate
    ws.auto_filter.ref = ws.dimensions
    _conditional_color_signal_columns(ws, list(df.columns), start_row=start_row + 1)
    _apply_borders(ws)
    _autosize_columns(ws)


def _conditional_color_signal_columns(ws, header_values: list, start_row: int) -> None:
    target_cols = [
        idx + 1 for idx, h in enumerate(header_values)
        if h and any(k in str(h) for k in ("Signal", "Buildup", "Institutional", "Smart Money"))
    ]
    for row in ws.iter_rows(min_row=start_row):
        for col_idx in target_cols:
            cell = row[col_idx - 1]
            val = str(cell.value or "").upper()
            fill = None
            if "BUY CE" in val or "LONG BUILDUP" in val or "INSTITUTIONAL" in val or "TRUE" in val:
                fill = FILL_GREEN
            elif "BUY PE" in val or "SHORT BUILDUP" in val:
                fill = FILL_RED
            elif "HOLD" in val or "WAIT" in val or "FLAT" in val:
                fill = FILL_AMBER
            if fill:
                cell.fill = fill


def export_excel_report(df: pd.DataFrame, meta: dict, pcr: float, max_pain: float,
                         support: Optional[float], resistance: Optional[float],
                         symbol: str, expiry_label: str, iv_rank: float,
                         iv_percentile: float, gex_dex: dict, market_pressure: Optional[MarketPressure] = None,
                         trade_signal: Optional[TradeSignal] = None,
                         po3_intelligence: Optional[dict] = None) -> io.BytesIO:
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    summary_rows = [
        ("Symbol", symbol), ("Expiry", expiry_label),
        ("Generated At", datetime.now().strftime("%d-%b-%Y %H:%M:%S")),
        ("Spot Price", round(meta.get("spot_price", 0.0), 2)),
        ("PCR", pcr), ("Max Pain", max_pain),
        ("Support (Max PE OI)", support), ("Resistance (Max CE OI)", resistance),
        ("IV Rank", iv_rank), ("IV Percentile", iv_percentile),
        ("Total GEX", round(gex_dex.get("total_gex", 0.0), 2)),
        ("Total DEX", round(gex_dex.get("total_dex", 0.0), 2)),
        ("Gamma Flip Strike", gex_dex.get("gamma_flip")),
        ("Total CE OI", int(df["ce_oi"].sum()) if not df.empty else 0),
        ("Total PE OI", int(df["pe_oi"].sum()) if not df.empty else 0),
    ]
    
    if market_pressure:
        summary_rows.extend([
            ("", ""),
            ("MARKET PRESSURE", ""),
            ("Market Sentiment", market_pressure.market_sentiment),
            ("Net Market Bias", market_pressure.net_market_bias),
            ("Buy Pressure", market_pressure.total_call_pressure),
            ("Sell Pressure", market_pressure.total_put_pressure),
            ("Volume Surge Detected", market_pressure.volume_surge_detected),
            ("OI Accumulation", market_pressure.oi_accumulation_detected),
        ])
    
    if trade_signal:
        summary_rows.extend([
            ("", ""),
            ("TRADE SIGNAL (Price Action)", ""),
            ("Signal", trade_signal.signal),
            ("Entry", round(trade_signal.entry, 2)),
            ("Stop Loss", round(trade_signal.stop_loss, 2)),
            ("Target 1", round(trade_signal.target_1, 2)),
            ("Target 2", round(trade_signal.target_2, 2)),
            ("Target 3", round(trade_signal.target_3, 2)),
            ("Risk:Reward", round(trade_signal.risk_reward_ratio, 2)),
            ("Probability", f"{trade_signal.probability:.1f}%"),
            ("Confidence", f"{trade_signal.confidence:.1f}%"),
        ])

    if po3_intelligence:
        summary_rows.extend([
            ("", ""), ("PO3 + OPTIONS INTELLIGENCE", ""),
            ("PO3 Phase", po3_intelligence.get("po3_phase")),
            ("PO3 Direction", po3_intelligence.get("po3_direction")),
            ("ATM Strike", po3_intelligence.get("atm_strike")),
            ("CE OI", po3_intelligence.get("ce_oi")),
            ("PE OI", po3_intelligence.get("pe_oi")),
            ("CE OI Change", po3_intelligence.get("ce_oi_change")),
            ("PE OI Change", po3_intelligence.get("pe_oi_change")),
            ("PCR", po3_intelligence.get("pcr")),
            ("Call Writing", po3_intelligence.get("call_writing")),
            ("Put Writing", po3_intelligence.get("put_writing")),
            ("Call Unwinding", po3_intelligence.get("call_unwinding")),
            ("Put Unwinding", po3_intelligence.get("put_unwinding")),
            ("PO3 + Options Confirmation", po3_intelligence.get("po3_options_confirmation")),
            ("Final CE Bias", po3_intelligence.get("final_ce_bias")),
            ("Final PE Bias", po3_intelligence.get("final_pe_bias")),
            ("Confidence", po3_intelligence.get("confidence")),
        ])
    
    ws_summary.cell(row=1, column=1, value="Metric")
    ws_summary.cell(row=1, column=2, value="Value")
    _style_header_row(ws_summary, 1)
    for i, (label, value) in enumerate(summary_rows, start=2):
        ws_summary.cell(row=i, column=1, value=label)
        ws_summary.cell(row=i, column=2, value=value)
    ws_summary.freeze_panes = "A2"
    _apply_borders(ws_summary)
    _autosize_columns(ws_summary)

    ws_chain = wb.create_sheet("Option Chain")
    chain_export_cols = [c for c in [
        "strike_price", "ce_oi", "ce_chng_oi", "ce_oi_change_pct", "ce_volume", "ce_iv", "ce_ltp",
        "ce_bid", "ce_ask", "CE Buildup", "CE Moneyness", "AI Signal", "AI Confidence %",
        "Institutional Signal", "Smart Money", "PE Moneyness", "PE Buildup",
        "pe_bid", "pe_ask", "pe_ltp", "pe_iv", "pe_volume", "pe_oi_change_pct", "pe_chng_oi", "pe_oi",
    ] if c in df.columns]
    _write_dataframe(ws_chain, df[chain_export_cols])

    ws_greeks = wb.create_sheet("Greeks")
    greek_cols = [c for c in [
        "strike_price", "ce_delta", "ce_gamma", "ce_theta", "ce_vega",
        "pe_delta", "pe_gamma", "pe_theta", "pe_vega",
    ] if c in df.columns]
    if greek_cols:
        _write_dataframe(ws_greeks, df[greek_cols])

    ws_signals = wb.create_sheet("AI Signals")
    signal_cols = [c for c in [
        "strike_price", "AI Signal", "AI Confidence %", "CE Score", "PE Score",
        "Institutional Signal", "Smart Money",
    ] if c in df.columns]
    if signal_cols:
        sig_df = df[signal_cols].sort_values("AI Confidence %", ascending=False) if "AI Confidence %" in df.columns else df[signal_cols]
        _write_dataframe(ws_signals, sig_df)
    
    if "buy_pressure" in df.columns:
        ws_pressure = wb.create_sheet("Buy-Sell Pressure")
        pressure_cols = [c for c in [
            "strike_price", "buy_pressure", "sell_pressure", "net_pressure",
            "pressure_direction", "aggression_level", "volume_spike", "oi_surge",
        ] if c in df.columns]
        if pressure_cols:
            _write_dataframe(ws_pressure, df[pressure_cols])

    if "movement_score" in df.columns:
        ws_move = wb.create_sheet("Strike Movement")
        movement_cols = [c for c in [
            "movement_rank", "strike_price", "movement_score", "movement_bias", "movement_strength",
            "ce_movement_score", "pe_movement_score", "buy_pressure", "sell_pressure",
            "aggression_level", "total_volume", "ce_chng_oi", "pe_chng_oi",
            "volume_spike", "oi_surge", "AI Signal", "AI Confidence %",
        ] if c in df.columns]
        if movement_cols:
            move_df = df[movement_cols].sort_values("movement_score", ascending=False)
            _write_dataframe(ws_move, move_df)

    # PO3 Intelligence sheet (safe optional export)
    if po3_intelligence and isinstance(po3_intelligence, dict):
        ws_po3 = wb.create_sheet("PO3 Intelligence")
        ws_po3.cell(row=1, column=1, value="Metric")
        ws_po3.cell(row=1, column=2, value="Value")
        _style_header_row(ws_po3, 1)

        for row_idx, (key, value) in enumerate(po3_intelligence.items(), start=2):
            ws_po3.cell(row=row_idx, column=1, value=str(key))
            if isinstance(value, (dict, list, tuple, set)):
                value = str(value)
            elif isinstance(value, np.generic):
                value = value.item()
            elif isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
                value = None
            ws_po3.cell(row=row_idx, column=2, value=value)

        ws_po3.freeze_panes = "A2"
        _apply_borders(ws_po3)
        _autosize_columns(ws_po3)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_csv_bytes(df: pd.DataFrame) -> bytes:
    if df is None or df.empty:
        return b""
    out = df.copy()
    out = out.replace([np.inf, -np.inf], np.nan)
    return out.to_csv(index=False).encode("utf-8")


# ══════════════════════════════════════════════════════════════════════════
# 19. STREAMLIT UI (ORIGINAL WITH PRESSURE ENHANCEMENTS)
# ══════════════════════════════════════════════════════════════════════════

def _configure_page() -> None:
    try:
        st.set_page_config(
            page_title="NSE Options Chain Dashboard + Price Action + Buy/Sell Pressure",
            page_icon="📊", layout="wide", initial_sidebar_state="expanded",
        )
    except Exception as e:
        logger.warning("st.set_page_config() skipped: %s", e)


def _inject_css() -> None:
    st.markdown(f"""
    <style>
    .stApp {{ background-color: {DARK_BG}; }}
    section[data-testid="stSidebar"] {{ background-color: {PANEL_BG}; border-right: 1px solid {BORDER_COLOR}; }}
    div[data-testid="metric-container"] {{
        background: {PANEL_BG}; border: 1px solid {BORDER_COLOR}; border-radius: 8px; padding: 14px 18px;
    }}
    div[data-testid="metric-container"] label {{ color: {TEXT_MUTED} !important; font-size: 12px;
        text-transform: uppercase; letter-spacing: 0.08em; }}
    div[data-testid="metric-container"] div[data-testid="stMetricValue"] {{
        color: {TEXT_MAIN} !important; font-size: 21px; font-weight: 700; font-family: 'Courier New', monospace; }}
    h1, h2, h3 {{ color: {TEXT_MAIN} !important; }}
    .block-title {{ color: {BLUE}; font-size: 13px; font-weight: 600; text-transform: uppercase;
        letter-spacing: 0.1em; margin-bottom: 8px; }}
    button[data-baseweb="tab"] {{ color: {TEXT_MUTED} !important; }}
    button[data-baseweb="tab"][aria-selected="true"] {{ color: {BLUE} !important; border-bottom: 2px solid {BLUE}; }}
    hr {{ border-color: {BORDER_COLOR}; }}
    .intel-card {{ background: {PANEL_BG}; border: 1px solid {BORDER_COLOR}; border-radius: 8px;
        padding: 14px 16px; margin-bottom: 8px; }}
    .intel-label {{ color: {TEXT_MUTED}; font-size: 11px; text-transform: uppercase; letter-spacing: .08em; }}
    .intel-value {{ color: {TEXT_MAIN}; font-size: 20px; font-weight: 700; font-family: 'Courier New', monospace; }}
    .trade-signal-card {{ background: linear-gradient(135deg, {PANEL_BG} 0%, #1c2128 100%); 
        border: 2px solid {BLUE}; border-radius: 12px; padding: 16px 18px; margin: 12px 0; }}
    .trade-signal-buy {{ border-left: 4px solid {GREEN}; }}
    .trade-signal-sell {{ border-left: 4px solid {RED}; }}
    </style>
    """, unsafe_allow_html=True)


def _pcr_sentiment_badge(pcr: float) -> str:
    if pcr > 1.3:
        return f'<span style="color:{GREEN};font-weight:700;">🟢 Bullish (High PCR)</span>'
    if pcr < 0.7:
        return f'<span style="color:{RED};font-weight:700;">🔴 Bearish (Low PCR)</span>'
    return f'<span style="color:{AMBER};font-weight:700;">🟡 Neutral</span>'


def _sidebar_config() -> dict:
    with st.sidebar:
        st.markdown("### ⚙️ Configuration")
        instrument_type = st.radio("Instrument Type", ["Index", "F&O Stock"], key="oc_instr_type")
        is_index = instrument_type == "Index"

        if is_index:
            symbol = st.selectbox("Index", list(INDEX_SYMBOLS.keys()), key="oc_index_select")
            if symbol in NSE_UNSUPPORTED_INDICES:
                st.caption(f"ℹ️ {symbol} is BSE-listed — requires a connected FYERS client.")
        else:
            raw_symbol = st.text_input(
                "Stock Symbol (e.g. RELIANCE, TCS, INFY)", "RELIANCE", key="oc_stock_input"
            )
            symbol = normalize_stock_symbol(raw_symbol)

        strike_count = st.slider("Strikes Around ATM", 5, 40, 15, step=5, key="oc_strike_count")
        show_greeks = st.checkbox("Show Greeks in chain table", value=True, key="oc_show_greeks")
        min_ai_conf = st.slider("Min AI Confidence %", 0, 100, 55, step=5, key="oc_min_ai_conf")
        strike_search_raw = st.text_input("Search Strike Price", value="", key="oc_strike_search")
        strike_search = 0.0
        if strike_search_raw.strip():
            try:
                strike_search = float(strike_search_raw.strip())
            except ValueError:
                st.caption("⚠️ Enter a valid numeric strike price.")

        default_lot = DEFAULT_LOT_SIZES.get(symbol, DEFAULT_LOT_SIZES["_STOCK_DEFAULT"])
        lot_size = st.number_input(
            "Lot Size", min_value=1, value=default_lot, step=1, key="oc_lot_size",
        )

        st.divider()
        st.markdown("### 📊 Price Action Analysis")
        analyze_price_action = st.checkbox("Fetch & Analyze Price Action (requires FYERS)", value=False, key="oc_price_action")

        # ADDITIVE: optional scalping layer; old price-action logic remains untouched.
        scalping_mode = st.checkbox(
            "⚡ Enable Scalping Mode (FYERS)",
            value=False,
            key="oc_scalping_mode",
            help="Adds 1M/3M/5M early-warning analysis without changing the existing MTF signal.",
        )

        st.divider()
        st.markdown("### 🔄 Auto Refresh")
        auto_refresh = st.checkbox("Enable auto-refresh", value=False, key="oc_auto_refresh")
        refresh_secs = st.slider("Refresh interval (seconds)", 10, 120, 20, step=5, key="oc_refresh_secs",
                                  disabled=not auto_refresh)

        st.divider()
        debug_mode = st.checkbox("Debug info", value=False, key="oc_debug_mode")
        col_free, col_live = st.columns(2)
        with col_free:
            free_run = st.button("🆓 FREE RUN", use_container_width=True, type="secondary",
                                 help="Runs the NSE option-chain scanner without requiring a FYERS client.")
        with col_live:
            fetch_clicked = st.button("📡 FETCH LIVE", use_container_width=True)

    return {
        "is_index": is_index, "symbol": symbol, "strike_count": strike_count,
        "show_greeks": show_greeks, "min_ai_conf": min_ai_conf, "strike_search": strike_search,
        "lot_size": lot_size, "auto_refresh": auto_refresh, "refresh_secs": refresh_secs,
        "debug_mode": debug_mode, "fetch_clicked": (fetch_clicked or free_run),
        "free_run": free_run,
        "analyze_price_action": analyze_price_action,
        "scalping_mode": scalping_mode,
    }


def _do_fetch_and_process(cfg: dict, fyers: Any = None) -> Optional[dict]:
    """Full fetch -> parse -> validate -> analytics pipeline."""
    preferred_expiry = st.session_state.get("oc_selected_expiry", "")
    stock_name = cfg["symbol"] if not cfg["is_index"] else ""
    fetch_result = fetch_chain_unified(
        fyers, cfg["symbol"], cfg["is_index"], stock_name, preferred_expiry, cfg["strike_count"],
    )
    if cfg["debug_mode"]:
        st.write("**Fetch result:**", fetch_result.get("ok"), fetch_result.get("source"), fetch_result.get("error"))

    if not fetch_result.get("ok"):
        st.error(
            f"⚠️ Could not fetch option chain for **{cfg['symbol']}**: "
            f"{fetch_result.get('error', 'Unknown error.')} "
        )
        return None

    df_all: pd.DataFrame = fetch_result["df"]
    meta: dict = fetch_result["meta"]
    data_source: str = fetch_result.get("source", "UNKNOWN")

    if not validate_chain_df(df_all):
        st.error(
            f"⚠️ Received a response for **{cfg['symbol']}**, but it did not contain a usable "
            "option chain."
        )
        return None

    spot = meta["spot_price"]
    df = filter_strikes_around_atm(df_all, spot, cfg["strike_count"])
    if df.empty:
        df = df_all

    expiry_label = meta["selected_expiry"]
    if spot:
        atm_pos = int((df["strike_price"] - float(spot)).abs().to_numpy().argmin())
        atm_strike = float(df.iloc[atm_pos]["strike_price"])
    else:
        atm_strike = float(df["strike_price"].median())

    df = add_greeks_columns(df, spot, expiry_label)
    df = classify_buildup(df)
    df = classify_moneyness(df, spot)
    df = compute_ai_scores(df, spot, atm_strike, calc_max_pain(df), calc_pcr(df))
    df = detect_institutional_smart_money(df)
    
    # ✅ ADD PRESSURE ANALYSIS
    df, market_pressure = add_pressure_analysis(df, spot, cfg["lot_size"])
    df = add_strike_movement_score(df)
    df, movement_early_warning = compute_movement_early_warning(
        df, cfg["symbol"], meta["selected_expiry"], spot
    )

    pcr = calc_pcr(df)
    max_pain = calc_max_pain(df)
    support, resistance = calc_support_resistance(df)
    max_oi = calc_max_oi(df)

    atm_iv = _atm_iv(df, spot)
    update_iv_history(cfg["symbol"], expiry_label, atm_iv)
    iv_rank, iv_percentile = compute_iv_rank_percentile(cfg["symbol"], expiry_label, atm_iv)

    gex_dex = compute_gex_dex(df, spot, cfg["lot_size"])

    oi_shift_notes = detect_oi_shift(cfg["symbol"], expiry_label, support, resistance)

    price_action_data = None
    trade_signal = None
    if cfg["analyze_price_action"] and fyers is not None:
        fyers_symbol_candidates = (
            _fyers_index_candidates(cfg["symbol"]) if cfg["is_index"] else fyers_stock_symbol_candidates(stock_name)
        )
        fyers_symbol = fyers_symbol_candidates[0] if fyers_symbol_candidates else None
        
        if fyers_symbol:
            df_dict = {}
            for tf_name, tf_mins in TIMEFRAMES.items():
                df_tf = fetch_fyers_candles(fyers, fyers_symbol, tf_mins, count=100)
                df_dict[tf_name] = df_tf
            
            if any(df_dict.values()):
                for tf_name in df_dict:
                    if df_dict[tf_name] is not None and not df_dict[tf_name].empty:
                        df_dict[tf_name] = add_technical_indicators(df_dict[tf_name])
                
                mss = detect_mss(df_dict)
                trade_signal = generate_trade_signal(df_dict, spot, mss, fyers is not None)
                
                price_action_data = {
                    "df_dict": df_dict,
                    "mss": mss,
                    "trade_signal": trade_signal,
                }

    scalping_data = None
    if cfg.get("scalping_mode") and fyers is not None:
        fyers_symbol_candidates = (
            _fyers_index_candidates(cfg["symbol"]) if cfg["is_index"] else fyers_stock_symbol_candidates(stock_name)
        )

        # Try every resolved FYERS symbol variant instead of only the first one.
        scalp_dict = {}
        scalping_symbol = None
        for candidate in fyers_symbol_candidates:
            test_dict = {}
            for tf_name, tf_mins in SCALPING_TIMEFRAMES.items():
                test_dict[tf_name] = fetch_fyers_candles(
                    fyers, candidate, tf_mins, count=120
                )

            usable = sum(
                1 for df_tf in test_dict.values()
                if isinstance(df_tf, pd.DataFrame) and not df_tf.empty
            )
            if usable:
                scalping_symbol = candidate
                scalp_dict = test_dict
                break

        if scalp_dict:
            scalping_data = compute_scalping_early_warning(scalp_dict, spot)
            scalping_data["df_dict"] = scalp_dict
            scalping_data["fyers_symbol"] = scalping_symbol
        else:
            scalping_data = {
                "enabled": False,
                "trigger": "FYERS candle fetch failed",
                "score": 0.0,
                "confidence": 0.0,
                "direction": "WATCH",
                "entry": None,
                "sl": None,
                "targets": [],
                "reasons": [
                    "FYERS returned no usable 1M/3M/5M/15M candles.",
                    "Check the FYERS symbol/permissions and market-data availability.",
                ],
                "df_dict": {},
                "fyers_symbol": None,
            }

    po3_price_df = None
    if price_action_data and price_action_data.get("df_dict"):
        po3_price_df = price_action_data["df_dict"].get("5M")
    po3_intelligence = compute_po3_options_intelligence(df, spot, cfg["symbol"], po3_price_df)
    po3_intelligence["final_signal"] = compute_final_signal(po3_intelligence, trade_signal)

    return {
        "df": df, "meta": meta, "spot": spot, "atm_strike": atm_strike, "expiry_label": expiry_label,
        "pcr": pcr, "max_pain": max_pain, "support": support, "resistance": resistance, "max_oi": max_oi,
        "atm_iv": atm_iv, "iv_rank": iv_rank, "iv_percentile": iv_percentile, "gex_dex": gex_dex,
        "oi_shift_notes": oi_shift_notes, "data_source": data_source,
        "price_action_data": price_action_data, "trade_signal": trade_signal,
        "market_pressure": market_pressure,
        "po3_intelligence": po3_intelligence,
        "final_signal": po3_intelligence.get("final_signal", {}),
        "scalping_data": scalping_data,
        "movement_early_warning": movement_early_warning,
    }


def _render_summary_cards(state: dict) -> None:
    """Enhanced with pressure metrics."""
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Spot Price", f"₹{state['spot']:,.2f}" if state["spot"] else "—")
    c2.metric("ATM Strike", f"₹{state['atm_strike']:,.0f}")
    c3.metric("PCR", f"{state['pcr']:.3f}")
    c4.metric("Max Pain", f"₹{state['max_pain']:,.0f}")
    c5.metric("IV Rank / %ile", f"{state['iv_rank']:.0f} / {state['iv_percentile']:.0f}")

    c6, c7, c8, c9, c10 = st.columns(5)
    c6.metric("Support", f"₹{state['support']:,.0f}" if state["support"] else "—")
    c7.metric("Resistance", f"₹{state['resistance']:,.0f}" if state["resistance"] else "—")
    c8.metric("Total GEX", f"{state['gex_dex'].get('total_gex', 0):,.0f}")
    c9.metric("Total DEX", f"{state['gex_dex'].get('total_dex', 0):,.0f}")
    
    fs = state.get("final_signal", {})
    if fs:
        c10.metric("FINAL SIGNAL", fs.get("signal", "WAIT"),
                   delta=f"{float(fs.get('confidence', 0) or 0):.0f}%")
    elif state.get("trade_signal"):
        c10.metric("FINAL SIGNAL", state["trade_signal"].signal,
                   delta=f"{state['trade_signal'].confidence:.0f}%")
    else:
        c10.metric("FINAL SIGNAL", "WAIT")

    # NEW: Pressure metrics
    mp = state.get("market_pressure")
    if mp:
        p1, p2, p3, p4, p5 = st.columns(5)
        p1.metric("Buy Pressure", f"{mp.total_call_pressure:.0f}")
        p2.metric("Sell Pressure", f"{mp.total_put_pressure:.0f}")
        p3.metric("Market Bias", f"{mp.net_market_bias:+.0f}", delta=mp.market_sentiment.split()[0])
        p4.metric("Volume Spike", "🔴 YES" if mp.volume_surge_detected else "🟢 No")
        p5.metric("OI Surge", "🔴 YES" if mp.oi_accumulation_detected else "🟢 No")


def _render_ai_signal_cards(state: dict, min_conf: float) -> None:
    df = state["df"]
    qualifying = df[df["AI Confidence %"] >= min_conf].sort_values("AI Confidence %", ascending=False)
    if qualifying.empty:
        st.info(f"No strikes meet the {min_conf:.0f}% AI confidence threshold.")
        return
    for _, row in qualifying.head(15).iterrows():
        signal = row["AI Signal"]
        color = GREEN if "CE" in signal else (RED if "PE" in signal else AMBER)
        st.markdown(f"""
        <div class="intel-card">
          <div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;">
            <div><b style="color:{TEXT_MAIN};">{row['strike_price']:,.0f}</b>
              &nbsp; <span style="color:{color};font-weight:700;">{_safe_cell(signal)}</span></div>
            <div class="intel-label">Confidence
              <span style="color:{TEXT_MAIN};font-weight:700;font-size:15px;">{row['AI Confidence %']:.0f}%</span></div>
          </div>
          <div style="margin-top:8px;color:{TEXT_MUTED};font-size:12px;">
            CE Score {row['CE Score']:.1f} &nbsp;|&nbsp; PE Score {row['PE Score']:.1f}
          </div>
        </div>
        """, unsafe_allow_html=True)


def run_dashboard(fyers: Any = None) -> None:
    _configure_page()
    _inject_css()
    st.markdown("## 📊 Options Chain + Price Action + Buy/Sell Pressure")

    # Prominent Run button: keeps the dashboard usable even when the sidebar is collapsed.
    run_col1, run_col2, run_col3 = st.columns([1.2, 1.2, 4.6])
    with run_col1:
        run_clicked = st.button("▶️ RUN SCANNER", use_container_width=True, type="primary",
                                help="Fetch live option-chain data and run the scanner.")
    with run_col2:
        refresh_clicked = st.button("🔄 REFRESH NOW", use_container_width=True,
                                    help="Fetch the latest available data now.")

    cfg = _sidebar_config()
    cfg["fetch_clicked"] = bool(cfg.get("fetch_clicked") or run_clicked or refresh_clicked)
    if run_clicked or refresh_clicked:
        cfg["free_run"] = True
        # Do not require FYERS just to run the NSE option-chain scanner.
        if fyers is None:
            cfg["analyze_price_action"] = False

    if cfg["symbol"] != st.session_state.get("oc_last_symbol"):
        st.session_state["oc_last_symbol"] = cfg["symbol"]
        st.session_state.pop("oc_state", None)
        st.session_state.pop("oc_selected_expiry", None)

    if cfg["fetch_clicked"] or cfg["auto_refresh"]:
        # RUN SCANNER / FREE RUN intentionally works with NSE-only option-chain data.
        # Price-action/MSS signals require a connected FYERS client.
        if cfg.get("free_run") and fyers is None:
            cfg["analyze_price_action"] = False
        with st.spinner(f"Fetching live data for {cfg['symbol']}…"):
            result = _do_fetch_and_process(cfg, fyers)
        if result is not None:
            st.session_state["oc_state"] = result
            st.session_state["oc_selected_expiry"] = result["expiry_label"]

    state = st.session_state.get("oc_state")
    if state is None:
        st.info("👈 Choose an instrument and click **Fetch Live Data**.")
        return

    df: pd.DataFrame = state["df"]
    meta = state["meta"]

    expiry_options = meta.get("expiry_dates", [])
    if expiry_options:
        current = state["expiry_label"] if state["expiry_label"] in expiry_options else expiry_options[0]
        selected = st.selectbox(
            "Expiry", expiry_options, index=expiry_options.index(current), key="oc_expiry_selectbox"
        )
        if selected != st.session_state.get("oc_selected_expiry"):
            st.session_state["oc_selected_expiry"] = selected
            with st.spinner("Reloading for selected expiry…"):
                refreshed = _do_fetch_and_process(cfg, fyers)
            if refreshed is not None:
                st.session_state["oc_state"] = refreshed
                state = refreshed
                df = state["df"]

    if cfg["debug_mode"]:
        with st.expander("🔍 Debug", expanded=False):
            st.write(f"Rows: {meta.get('total_rows_seen')} seen, {meta.get('rows_parsed')} parsed")
            st.write(f"Source: {state.get('data_source', 'UNKNOWN')}")
            st.dataframe(df.head(3), use_container_width=True)

    _render_summary_cards(state)

    # Big-movement alert: ranking/pressure based, not a guaranteed prediction.
    if not df.empty and "movement_score" in df.columns:
        top_move = df.loc[df["movement_score"].idxmax()]
        if float(top_move.get("movement_score", 0)) >= 75:
            bias = str(top_move.get("movement_bias", "NEUTRAL"))
            alert_text = (
                f"🚨 BIG MOVEMENT ALERT: Strike {top_move['strike_price']:,.0f} | "
                f"{bias} | Score {float(top_move['movement_score']):.0f}/100 | "
                f"Buy {float(top_move.get('buy_pressure', 0)):.0f} / "
                f"Sell {float(top_move.get('sell_pressure', 0)):.0f}"
            )
            st.warning(alert_text)

    _render_movement_early_warning(state.get("movement_early_warning") or {})

    if cfg.get("scalping_mode"):
        _render_scalping_panel(state.get("scalping_data") or {})

    source = state.get("data_source", "UNKNOWN")
    source_badge = "🟢 FYERS" if source == "FYERS" else ("🟡 NSE" if source == "NSE" else "⚪ Unknown")
    mode_badge = "⚡ SCALPING ON" if cfg.get("scalping_mode") else "NORMAL MODE"
    st.caption(
        f"📡 Source: **{source_badge}** | Mode: **{mode_badge}** | Sentiment: {_pcr_sentiment_badge(state['pcr'])} | "
        f"Last update: **{meta.get('fetched_at', datetime.now()).strftime('%H:%M:%S')}**"
    , unsafe_allow_html=True)

    for note in state.get("oi_shift_notes", []):
        st.info(f"🔀 {note}")

    if cfg["strike_search"]:
        match = df[(df["strike_price"] - cfg["strike_search"]).abs() < 0.5]
        if not match.empty:
            r = match.iloc[0]
            st.success(
                f"🔎 {cfg['strike_search']:,.0f} → CE {r['ce_ltp']:.2f} (OI {r['ce_oi']:,.0f}) | "
                f"PE {r['pe_ltp']:.2f} (OI {r['pe_oi']:,.0f})"
            )

    st.divider()

    if state.get("price_action_data") and state["price_action_data"].get("df_dict"):
        tab_chain, tab_charts, tab_pressure, tab_movement, tab_greeks, tab_ai, tab_gex, tab_po3, tab_price_action, tab_export = st.tabs([
            "📋 Chain", "📈 OI", "💪 Pressure", "🎯 Strike Movement", "🧮 Greeks", "🤖 AI", "⚡ GEX", "🧠 PO3 Intelligence", "💹 Price Action", "📥 Export",
        ])
    else:
        tab_chain, tab_charts, tab_pressure, tab_movement, tab_greeks, tab_ai, tab_gex, tab_po3, tab_export = st.tabs([
            "📋 Chain", "📈 OI", "💪 Pressure", "🎯 Strike Movement", "🧮 Greeks", "🤖 AI", "⚡ GEX", "🧠 PO3 Intelligence", "📥 Export",
        ])

    with tab_chain:
        st.markdown(render_chain_table_html(df, cfg["show_greeks"]), unsafe_allow_html=True)

    with tab_charts:
        st.plotly_chart(chart_oi_bars(df, state["max_pain"]), use_container_width=True, config={"displayModeBar": False})
        col_a, col_b = st.columns(2)
        with col_a:
            st.plotly_chart(gauge_pcr(state["pcr"]), use_container_width=True, config={"displayModeBar": False})
        with col_b:
            st.plotly_chart(chart_iv_skew(df), use_container_width=True, config={"displayModeBar": False})

    # NEW: Pressure Tab
    with tab_pressure:
        st.markdown('<div class="block-title">💪 Buy/Sell Pressure Analysis</div>', unsafe_allow_html=True)
        
        mp = state.get("market_pressure")
        if mp:
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                color = GREEN if mp.net_market_bias > 0 else RED
                st.metric("Market Bias", f"{mp.net_market_bias:+.0f}", delta=mp.market_sentiment)
            with col2:
                st.metric("Buy Pressure", f"{mp.total_call_pressure:.0f}")
            with col3:
                st.metric("Sell Pressure", f"{mp.total_put_pressure:.0f}")
            with col4:
                st.metric("Aggression", f"{(df['aggression_level'].mean() if 'aggression_level' in df.columns else 50):.0f}")
            
            st.divider()
            
            col_itm, col_atm, col_otm = st.columns(3)
            with col_itm:
                st.metric("ITM Pressure", f"{mp.itm_pressure:.0f}")
            with col_atm:
                st.metric("ATM Pressure", f"{mp.atm_pressure:.0f}")
            with col_otm:
                st.metric("OTM Pressure", f"{mp.otm_pressure:.0f}")
            
            st.divider()
            
            anom_col1, anom_col2 = st.columns(2)
            with anom_col1:
                if mp.volume_surge_detected:
                    st.warning("🚨 **VOLUME SPIKE** — Unusual activity detected")
                else:
                    st.info("✓ Volume levels normal")
            with anom_col2:
                if mp.oi_accumulation_detected:
                    st.warning("🚨 **OI SURGE** — Strong accumulation detected")
                else:
                    st.info("✓ OI changes normal")
            
            st.divider()
            
            st.plotly_chart(chart_pressure_by_strike(df, state["spot"]), use_container_width=True, config={"displayModeBar": False})
            
            col_net, col_agg = st.columns(2)
            with col_net:
                st.plotly_chart(chart_net_pressure(df, state["spot"]), use_container_width=True, config={"displayModeBar": False})
            with col_agg:
                st.plotly_chart(chart_aggression_level(df), use_container_width=True, config={"displayModeBar": False})

    with tab_movement:
        st.markdown('<div class="block-title">🎯 Strike Movement Scanner</div>', unsafe_allow_html=True)
        st.caption("Movement Score is a ranking model based on current option-chain activity; it is not a guaranteed price prediction.")
        move_cols = [c for c in [
            "movement_rank", "strike_price", "movement_score", "movement_bias", "movement_strength",
            "ce_movement_score", "pe_movement_score", "buy_pressure", "sell_pressure",
            "aggression_level", "total_volume", "ce_chng_oi", "pe_chng_oi",
            "volume_spike", "oi_surge", "AI Signal", "AI Confidence %",
        ] if c in df.columns]
        move_view = df[move_cols].sort_values("movement_score", ascending=False).copy()
        top_n = st.slider("Top strikes", 5, min(30, max(5, len(move_view))), min(15, max(5, len(move_view))), key="movement_top_n") if len(move_view) >= 5 else len(move_view)
        if len(move_view):
            best = move_view.iloc[0]
            st.success(
                f"🎯 TOP MOVEMENT: {float(best['strike_price']):,.0f} | "
                f"{best.get('movement_bias', 'NEUTRAL')} | "
                f"Score {float(best.get('movement_score', 0)):.0f}/100 | "
                f"Strength {best.get('movement_strength', 'LOW')}"
            )
            early_cols = [c for c in [
                "strike_price", "early_movement_score", "early_movement_status",
                "movement_score_delta", "movement_score_acceleration",
                "movement_rising_scans", "early_movement_confidence"
            ] if c in move_view.columns]
            if early_cols:
                st.dataframe(
                    move_view.sort_values("early_movement_score", ascending=False).head(top_n)[early_cols],
                    use_container_width=True, hide_index=True
                )
            st.dataframe(move_view.head(top_n), use_container_width=True, hide_index=True)
            st.plotly_chart(chart_movement_score(df), use_container_width=True, config={"displayModeBar": False})
        else:
            st.info("No strike movement data available.")

    with tab_greeks:
        g1, g2 = st.columns(2)
        with g1:
            st.plotly_chart(chart_greeks(df, "delta"), use_container_width=True, config={"displayModeBar": False})
            st.plotly_chart(chart_greeks(df, "theta"), use_container_width=True, config={"displayModeBar": False})
        with g2:
            st.plotly_chart(chart_greeks(df, "gamma"), use_container_width=True, config={"displayModeBar": False})
            st.plotly_chart(chart_greeks(df, "vega"), use_container_width=True, config={"displayModeBar": False})

    with tab_ai:
        st.markdown('<div class="block-title">🤖 AI Trade Signals</div>', unsafe_allow_html=True)
        _render_ai_signal_cards(state, cfg["min_ai_conf"])

    with tab_gex:
        e1, e2, e3 = st.columns(3)
        e1.metric("Total GEX", f"{state['gex_dex'].get('total_gex', 0):,.0f}")
        e2.metric("Total DEX", f"{state['gex_dex'].get('total_dex', 0):,.0f}")
        gf = state["gex_dex"].get("gamma_flip")
        e3.metric("Gamma Flip", f"{gf:,.0f}" if gf else "—")
        st.plotly_chart(chart_gex_by_strike(state["gex_dex"]), use_container_width=True, config={"displayModeBar": False})

    with tab_po3:
        _render_po3_intelligence(state.get("po3_intelligence", {}))

    if state.get("price_action_data") and state["price_action_data"].get("df_dict"):
        with tab_price_action:
            st.markdown('<div class="block-title">💹 Price Action</div>', unsafe_allow_html=True)
            df_dict = state["price_action_data"]["df_dict"]
            
            for tf_name, df_tf in df_dict.items():
                if df_tf is not None and not df_tf.empty:
                    with st.expander(f"📊 {tf_name}", expanded=(tf_name == "5M")):
                        st.plotly_chart(chart_price_action(df_tf, title=f"{tf_name}"),
                                       use_container_width=True, config={"displayModeBar": False})

    with tab_export:
        st.markdown('<div class="block-title">📥 Export</div>', unsafe_allow_html=True)
        col_x, col_y = st.columns(2)
        with col_x:
            try:
                excel_buf = export_excel_report(
                    df=df,
                    meta=meta,
                    pcr=state.get("pcr", 0.0),
                    max_pain=state.get("max_pain", 0.0),
                    support=state.get("support"),
                    resistance=state.get("resistance"),
                    symbol=cfg.get("symbol", "OPTION_CHAIN"),
                    expiry_label=state.get("expiry_label", ""),
                    iv_rank=state.get("iv_rank", 0.0),
                    iv_percentile=state.get("iv_percentile", 0.0),
                    gex_dex=state.get("gex_dex", {}),
                    market_pressure=state.get("market_pressure"),
                    trade_signal=state.get("trade_signal"),
                    po3_intelligence=state.get("po3_intelligence", {}),
                )
                st.download_button(
                    "📥 Excel", data=excel_buf,
                    file_name=f"oc_{cfg['symbol']}_{datetime.now().strftime('%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"Excel export failed: {e}")
        with col_y:
            try:
                csv_bytes = export_csv_bytes(df)
                st.download_button(
                    "📥 CSV", data=csv_bytes,
                    file_name=f"oc_{cfg['symbol']}_{datetime.now().strftime('%H%M%S')}.csv",
                    mime="text/csv", use_container_width=True,
                )
            except Exception as e:
                st.error(f"CSV export failed: {e}")

    st.caption(
        f"**NSE Options + FYERS Price Action + Buy/Sell Pressure** | Last: {meta.get('fetched_at', datetime.now()).strftime('%H:%M:%S')} | "
        "Educational tool — not financial advice."
    )

    if cfg["auto_refresh"]:
        time.sleep(cfg["refresh_secs"])
        st.rerun()


def show_option_chain(fyers: Any = None) -> None:
    """Entry point for hosting apps."""
    if fyers is not None:
        logger.info("show_option_chain() received FYERS client — using as PRIMARY data source.")
    else:
        logger.info("show_option_chain() — no FYERS client (NSE fallback only).")
    run_dashboard(fyers)


if __name__ == "__main__":
    run_dashboard()
