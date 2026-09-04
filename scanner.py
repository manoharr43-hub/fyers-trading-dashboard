import streamlit as st
import pandas as pd
import numpy as np
import requests
import time
import io
import os
import re
import json
import csv
import gc
import logging
import sys
from datetime import datetime, timedelta
from typing import List, Optional, Tuple, Dict, Any
from concurrent.futures import ThreadPoolExecutor, as_completed


# ============================================================
# INDEPENDENT STRONG SIGNALS / MARKET DASHBOARD HELPERS
# ============================================================

def _show_strong_signals_button():
    """Independent Strong Signals runner."""
    if st.button("🔥 RUN STRONG SIGNALS", key="run_strong_signals", use_container_width=True):
        source = st.session_state.get("strong_source", "NSE Stocks")
        df = st.session_state.get("nse_df" if source == "NSE Stocks" else "fo_df")

        if df is None or len(df) == 0:
            st.warning(f"⚠️ Run the {source} scanner first.")
            return

        confidence_col = next(
            (c for c in ["AI CONFIDENCE %", "CONFIDENCE %", "Confidence %"]
             if c in df.columns),
            None
        )
        signal_col = next(
            (c for c in ["AI SIGNAL", "SIGNAL", "Signal"]
             if c in df.columns),
            None
        )

        if confidence_col is None:
            st.error("❌ Confidence column not found in scanner output.")
            return

        out = df.copy()
        conf = pd.to_numeric(out[confidence_col], errors="coerce").fillna(0)
        out = out[conf >= 75].copy()

        if signal_col:
            out = out[
                ~out[signal_col].astype(str).str.upper().isin(
                    ["NEUTRAL", "NO SIGNAL", "NONE", "N/A"]
                )
            ]

        if len(out) == 0:
            st.warning("No strong signals ≥75% found.")
        else:
            out = out.sort_values(confidence_col, ascending=False)
            st.success(f"🔥 {len(out)} strong signals found.")
            st.dataframe(out, use_container_width=True, height=450)


def _show_market_dashboard_button():
    """Independent Market Dashboard runner."""
    if st.button("📊 RUN MARKET DASHBOARD", key="run_market_dashboard", use_container_width=True):
        source = st.session_state.get("dash_source", "NSE Stocks")
        df = st.session_state.get("nse_df" if source == "NSE Stocks" else "fo_df")

        if df is None or len(df) == 0:
            st.warning(f"⚠️ Run the {source} scanner first.")
            return

        signal_col = next(
            (c for c in ["AI SIGNAL", "SIGNAL", "Signal"]
             if c in df.columns),
            None
        )
        confidence_col = next(
            (c for c in ["AI CONFIDENCE %", "CONFIDENCE %", "Confidence %"]
             if c in df.columns),
            None
        )

        if signal_col is None:
            st.error("❌ Signal column not found.")
            return

        s = df[signal_col].astype(str).str.upper()

        buy = int(s.str.contains("BUY", na=False).sum())
        sell = int(s.str.contains("SELL", na=False).sum())
        neutral = int(s.isin(["NEUTRAL", "NONE", "NO SIGNAL", "N/A"]).sum())

        strong_buy = 0
        strong_sell = 0
        avg_conf = 0.0

        if confidence_col:
            conf = pd.to_numeric(
                df[confidence_col], errors="coerce"
            ).fillna(0)

            strong_buy = int(
                ((s.str.contains("BUY", na=False)) & (conf >= 75)).sum()
            )
            strong_sell = int(
                ((s.str.contains("SELL", na=False)) & (conf >= 75)).sum()
            )
            avg_conf = float(conf.mean()) if len(conf) else 0.0

        total = len(df)

        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("TOTAL", total)
        c2.metric("🟢 BUY", buy)
        c3.metric("🔴 SELL", sell)
        c4.metric("🔥 STRONG BUY", strong_buy)
        c5.metric("🔥 STRONG SELL", strong_sell)

        st.metric("Average Confidence", f"{avg_conf:.1f}%")

        if buy > sell * 1.15:
            sentiment = "🟢 BULLISH"
        elif sell > buy * 1.15:
            sentiment = "🔴 BEARISH"
        else:
            sentiment = "🟡 NEUTRAL"

        st.subheader(f"Market Sentiment: {sentiment}")

        dash = pd.DataFrame({
            "Metric": [
                "Total Stocks",
                "BUY",
                "SELL",
                "NEUTRAL",
                "Strong BUY",
                "Strong SELL",
                "Average Confidence"
            ],
            "Value": [
                total,
                buy,
                sell,
                neutral,
                strong_buy,
                strong_sell,
                f"{avg_conf:.1f}%"
            ]
        })
        st.dataframe(dash, use_container_width=True, hide_index=True)




def _excel_download_button(df: pd.DataFrame, filename_prefix: str, key: str, label: str = "📥 DOWNLOAD EXCEL"):
    """Render an Excel download button for a DataFrame."""
    if df is None or df.empty:
        return
    try:
        excel_data = _format_excel_output(df, filename_prefix)
        filename = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        st.download_button(
            label,
            data=excel_data,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=key,
            use_container_width=True,
        )
    except Exception as e:
        st.error(f"❌ Excel export failed: {str(e)[:150]}")

# ════════════════════════════════════════════════════════════════════════════════
# IMPORTS & CONFIG (FROM ORIGINAL)
# ════════════════════════════════════════════════════════════════════════════════
try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

try:
    from zoneinfo import ZoneInfo
    IST = ZoneInfo("Asia/Kolkata")
except Exception:
    from datetime import timezone
    IST = timezone(timedelta(hours=5, minutes=30))

try:
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ════════════════════════════════════════════════════════════════════════════════
# CONSTANTS (ORIGINAL + NEW)
# ════════════════════════════════════════════════════════════════════════════════
DATE_FROM = (datetime.today() - timedelta(days=365)).strftime("%Y-%m-%d")
DATE_TO = datetime.today().strftime("%Y-%m-%d")
FYERS_NSE_CM_SYMBOL_MASTER = "https://public.fyers.in/sym_details/NSE_CM.csv"
NIFTY_BENCHMARK_SYMBOL = "NSE:NIFTY50-INDEX"
MAX_WORKERS = 8
BATCH_SIZE = 50
BATCH_PAUSE_SECONDS = 1.0
DEFAULT_SCAN_STOCKS = 2300
FYERS_APP_ID = os.environ.get("FYERS_APP_ID", "")
OPTIONS_STRIKE_COUNT = 10
OPTIONS_HTTP_TIMEOUT = 15

# ════════════════════════════════════════════════════════════════════════════════
# V17 NEW CONSTANTS
# ════════════════════════════════════════════════════════════════════════════════
DEFAULT_CONFIDENCE_THRESHOLD = 70
DEFAULT_RVOL_THRESHOLD = 1.2
DEFAULT_STRONG_RVOL = 1.5
GOLDEN_CROSS_MIN_SIGNALS = 1
DEATH_CROSS_MIN_SIGNALS = 1
ADDITIONAL_ANALYSIS_LOOKBACK_DAYS = 30
REFRESH_INTERVALS = [30, 60, 120]

# ════════════════════════════════════════════════════════════════════════════════
# MOMENTUM MOVERS CONSTANTS (NEW)
# ════════════════════════════════════════════════════════════════════════════════
MOMENTUM_MIN_SCORE = 65
MOMENTUM_STRONG_SCORE = 85
MOMENTUM_DEVELOPING_SCORE = 55
# LIVE movement scanner: recent 5M price action only. No consolidation requirement.
LIVE_MOVE_MIN_PCT = 0.35
LIVE_MOVE_BIG_PCT = 0.70
LIVE_MOVE_MIN_RVOL = 1.30
LIVE_MOVE_STRONG_RVOL = 1.80
LIVE_MOVE_MIN_BODY_PCT = 50.0
LIVE_MOVE_MIN_SCORE = 65
LIVE_MOVE_STRONG_SCORE = 85
LIVE_MOVE_LOOKBACK_DAYS = 2
MOMENTUM_MIN_MOVE_PCT = 0.10
MOMENTUM_MIN_RVOL = 1.20
MOMENTUM_MIN_BODY_PCT = 20
MOMENTUM_DISPLAY_COUNT = 10

# ════════════════════════════════════════════════════════════════════════════════
# 15-MIN REVERSAL SCANNER CONSTANTS (ORIGINAL)
# ════════════════════════════════════════════════════════════════════════════════
REVERSAL_RESOLUTION = "15"
REVERSAL_LOOKBACK_DAYS = 5
REVERSAL_CONFIRMATION_BARS = 0
REVERSAL_ATR_LENGTH = 5
REVERSAL_ATR_MULTIPLIER = 2.8
REVERSAL_MIN_MOVE_PCT = 0.015
REVERSAL_CUSTOM_ABS = 0.05

# ════════════════════════════════════════════════════════════════════════════════
# VOLUME BIG MOVEMENT CONSTANTS (ORIGINAL)
# ════════════════════════════════════════════════════════════════════════════════
VOL_BIGMOVE_MIN_RVOL = 2.5
VOL_BIGMOVE_MIN_BODY_PCT = 0.8
VOL_BIGMOVE_LOOKBACK = 20

# ════════════════════════════════════════════════════════════════════════════════
# MULTI-TIMEFRAME CONSTANTS (NEW)
# ════════════════════════════════════════════════════════════════════════════════
SWING_LOOKBACK_PERIODS = 20
EMA_PERIODS = [9, 21, 50, 200]
VWAP_LOOKBACK = 20
MSS_MIN_CONFIRMATION = 1
STRUCTURE_TIMEFRAMES = ["5", "15", "60"]
MASTER_SIGNAL_LOOKBACK_DAYS = 10

# ════════════════════════════════════════════════════════════════════════════════
# LOGGING & UTILITIES (ORIGINAL)
# ════════════════════════════════════════════════════════════════════════════════
def _now_ist() -> datetime:
    return datetime.now(IST)


# ============================================================
# UNIVERSAL SCANNER SIGNAL TIME
# Works with AI SIGNAL, AMD SIGNAL and PIN SIGNAL.
# Existing scanner rules are not changed.
# ============================================================
def _add_signal_time_columns(df: pd.DataFrame, signal_col: str,
                             symbol_col: str = "Symbol") -> pd.DataFrame:
    if df is None or df.empty or signal_col not in df.columns:
        return df.copy() if isinstance(df, pd.DataFrame) else df

    if symbol_col not in df.columns:
        for candidate in ("Symbol", "SYMBOL", "symbol"):
            if candidate in df.columns:
                symbol_col = candidate
                break
    if symbol_col not in df.columns:
        return df.copy()

    if "scanner_signal_time_history" not in st.session_state:
        st.session_state["scanner_signal_time_history"] = {}

    history = st.session_state["scanner_signal_time_history"]
    now = _now_ist()
    out = df.copy()
    first_list, last_list, age_list = [], [], []

    for _, row in out.iterrows():
        symbol = str(row.get(symbol_col, "")).strip()
        signal = str(row.get(signal_col, "")).strip().upper()

        # Any actionable BUY/SELL signal gets a time.
        active = bool(symbol) and ("BUY" in signal or "SELL" in signal)

        if not active:
            first_list.append("-")
            last_list.append("-")
            age_list.append("-")
            continue

        key = f"{signal_col}::{symbol}::{signal}"

        # If direction changes, the old BUY/SELL timer is removed.
        prefix = f"{signal_col}::{symbol}::"
        for old_key in list(history.keys()):
            if old_key.startswith(prefix) and old_key != key:
                history.pop(old_key, None)

        if key not in history:
            history[key] = {"first_seen": now, "last_seen": now}
        else:
            history[key]["last_seen"] = now

        first_seen = history[key]["first_seen"]
        last_seen = history[key]["last_seen"]
        seconds = max(0, int((now - first_seen).total_seconds()))

        if seconds < 60:
            age = "JUST NOW"
        elif seconds < 3600:
            age = f"{seconds // 60} min ago"
        else:
            age = f"{seconds // 3600}h {(seconds % 3600) // 60}m ago"

        first_list.append(first_seen.strftime("%d-%b-%Y %I:%M:%S %p"))
        last_list.append(last_seen.strftime("%d-%b-%Y %I:%M:%S %p"))
        age_list.append(age)

    out["SIGNAL TIME"] = first_list
    out["LAST SEEN"] = last_list
    out["SIGNAL AGE"] = age_list
    return out

def _ensure_app_folders() -> None:
    for folder in ("logs", "charts", "exports"):
        os.makedirs(folder, exist_ok=True)

_ensure_app_folders()
logger = logging.getLogger("nse_scanner_v17")
logger.setLevel(logging.DEBUG)

if not logger.handlers:
    fh = logging.FileHandler("logs/scanner.log", encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    fh.setFormatter(formatter)
    logger.addHandler(fh)

def _candle_signal_timestamp(df, is_daily: bool = False, resolution: str = "15") -> Tuple[str, str]:
    ts = df["Time"].iloc[-1]
    if ts.tzinfo is None:
        ts = ts.tz_localize("UTC")
    ts_ist = ts.tz_convert(IST)
    if is_daily:
        close_ts = ts_ist.replace(hour=15, minute=30, second=0, microsecond=0)
    else:
        try:
            minutes = int(resolution)
        except Exception:
            minutes = 15
        close_ts = ts_ist + timedelta(minutes=minutes)
    return close_ts.strftime("%d-%b-%Y"), close_ts.strftime("%H:%M:%S") + " IST"

def _generated_timestamp() -> str:
    return _now_ist().strftime("%d-%b-%Y %H:%M:%S IST")

# ════════════════════════════════════════════════════════════════════════════════
# CORE INDICATORS (RETAINED FROM ORIGINAL)
# ════════════════════════════════════════════════════════════════════════════════
def calculate_rsi(close, period: int = 14):
    delta = close.diff()
    gain = delta.clip(lower=0)
    loss = -delta.clip(upper=0)
    avg_gain = gain.ewm(alpha=1 / period, min_periods=period, adjust=False).mean()
    avg_loss = loss.ewm(alpha=1 / period, min_periods=period, adjust=False).mean()
    rs = avg_gain / avg_loss.replace(0, np.nan)
    return (100 - (100 / (1 + rs))).fillna(50)

def calculate_macd(close):
    ema12 = close.ewm(span=12, adjust=False).mean()
    ema26 = close.ewm(span=26, adjust=False).mean()
    macd_line = ema12 - ema26
    signal_line = macd_line.ewm(span=9, adjust=False).mean()
    return macd_line, signal_line, macd_line - signal_line

def calculate_atr(df, period: int = 14):
    h, l, c = df["High"], df["Low"], df["Close"]
    pc = c.shift(1)
    tr = pd.concat([h - l, (h - pc).abs(), (l - pc).abs()], axis=1).max(axis=1)
    return tr.ewm(alpha=1 / period, min_periods=period, adjust=False).mean()

def _last_valid_atr(df, period: int = 14) -> float:
    atr_series = calculate_atr(df, period)
    val = atr_series.iloc[-1] if len(atr_series) else np.nan
    if pd.isna(val) or val <= 0:
        last_close = float(df["Close"].iloc[-1]) if len(df) else 0.0
        val = max(last_close * 0.005, 0.01)
    return float(val)

# ════════════════════════════════════════════════════════════════════════════════
# BUYING/SELLING PRESSURE INDICATORS
# ════════════════════════════════════════════════════════════════════════════════
def calculate_buying_selling_pressure(df) -> Dict[str, Any]:
    if len(df) < 5:
        return {
            "buying_pressure": 50, "selling_pressure": 50,
            "buying_volume": 0, "selling_volume": 0,
            "pressure_ratio": 1.0, "trend": "NEUTRAL"
        }
    
    recent = df.tail(20).copy()
    buying_vol = 0.0
    selling_vol = 0.0
    
    for idx in range(len(recent)):
        candle = recent.iloc[idx]
        close = float(candle["Close"])
        open_ = float(candle["Open"])
        volume = float(candle["Volume"])
        
        if close > open_:
            buying_vol += volume
        elif close < open_:
            selling_vol += volume
        else:
            buying_vol += volume * 0.5
            selling_vol += volume * 0.5
    
    total_vol = buying_vol + selling_vol
    if total_vol == 0:
        return {
            "buying_pressure": 50, "selling_pressure": 50,
            "buying_volume": 0, "selling_volume": 0,
            "pressure_ratio": 1.0, "trend": "NEUTRAL"
        }
    
    bp_pct = (buying_vol / total_vol) * 100
    sp_pct = (selling_vol / total_vol) * 100
    pressure_ratio = buying_vol / selling_vol if selling_vol > 0 else float('inf')
    
    if bp_pct > 65:
        trend = "STRONG_BUYING"
    elif bp_pct > 55:
        trend = "BUYING"
    elif sp_pct > 65:
        trend = "STRONG_SELLING"
    elif sp_pct > 55:
        trend = "SELLING"
    else:
        trend = "NEUTRAL"
    
    return {
        "buying_pressure": round(bp_pct, 1),
        "selling_pressure": round(sp_pct, 1),
        "buying_volume": round(buying_vol, 0),
        "selling_volume": round(selling_vol, 0),
        "pressure_ratio": round(pressure_ratio, 2) if pressure_ratio != float('inf') else 0,
        "trend": trend
    }

# ════════════════════════════════════════════════════════════════════════════════
# MOMENTUM SCORING ENGINE (NEW - V17)
# ════════════════════════════════════════════════════════════════════════════════
def calculate_movement_metrics(df) -> Dict[str, float]:
    """Measure the latest CLOSED candle and short-term movement."""
    empty = {
        "move_5m_pct": 0.0,
        "move_15m_pct": 0.0,
        "candle_body_pct": 0.0,
        "atr_normalized_move": 0.0,
        "price_acceleration": 0.0,
    }
    if df is None or len(df) < 10:
        return empty
    try:
        d = df.reset_index(drop=True).copy()
        last = d.iloc[-1]
        prev = d.iloc[-2]
        close = float(last["Close"])
        open_ = float(last["Open"])
        high = float(last["High"])
        low = float(last["Low"])
        prev_close = float(prev["Close"])
        base_15 = float(d["Close"].iloc[-4]) if len(d) >= 4 else prev_close
        rng = max(high - low, 0.0)
        body = abs(close - open_)
        atr_series = calculate_atr(d, 14)
        atr = float(atr_series.iloc[-1]) if len(atr_series) and pd.notna(atr_series.iloc[-1]) else max(close * 0.005, 0.01)
        prev_body = abs(float(prev["Close"]) - float(prev["Open"]))
        return {
            "move_5m_pct": round(abs(close-prev_close)/prev_close*100, 2) if prev_close else 0.0,
            "move_15m_pct": round(abs(close-base_15)/base_15*100, 2) if base_15 else 0.0,
            "candle_body_pct": round(body/rng*100, 1) if rng else 0.0,
            "atr_normalized_move": round(body/atr, 2) if atr else 0.0,
            "price_acceleration": round(body/prev_body, 2) if prev_body else 1.0,
        }
    except Exception:
        return empty


def _bigmove_pivots(df: pd.DataFrame, left: int = 2, right: int = 2):
    """Confirmed pivots, excluding the current unconfirmed candle."""
    if df is None or len(df) < left + right + 5:
        return [], []
    d = df.reset_index(drop=True)
    highs = d["High"].astype(float).to_numpy()
    lows = d["Low"].astype(float).to_numpy()
    ph, pl = [], []
    for i in range(left, len(d) - right):
        if highs[i] >= max(highs[i-left:i]) and highs[i] > max(highs[i+1:i+right+1]):
            ph.append((i, float(highs[i])))
        if lows[i] <= min(lows[i-left:i]) and lows[i] < min(lows[i+1:i+right+1]):
            pl.append((i, float(lows[i])))
    return ph, pl


def detect_last_consolidation(df: pd.DataFrame) -> Dict[str, Any]:
    """Find the most recent tight range immediately before the current move."""
    empty = {
        "found": False, "start_idx": None, "end_idx": None, "bars": 0,
        "high": None, "low": None, "range_pct": 0.0, "range_atr": 0.0,
    }
    if df is None or len(df) < 20:
        return empty
    try:
        d = df.reset_index(drop=True).copy()
        last_idx = len(d) - 1
        atr_s = calculate_atr(d, 14)
        atr = float(atr_s.iloc[-1]) if pd.notna(atr_s.iloc[-1]) else max(float(d["Close"].iloc[-1])*0.005, 0.01)
        best = None
        # The latest consolidation must end before the breakout candle.
        for end_idx in range(last_idx - 1, max(2, last_idx - BIGMOVE_LOOKBACK_BARS), -1):
            for bars in range(BIGMOVE_CONSOLIDATION_MIN_BARS, BIGMOVE_CONSOLIDATION_MAX_BARS + 1):
                start_idx = end_idx - bars + 1
                if start_idx < 2:
                    continue
                w = d.iloc[start_idx:end_idx+1]
                hi = float(w["High"].max()); lo = float(w["Low"].min())
                mid = (hi + lo) / 2.0
                if mid <= 0:
                    continue
                range_pct = (hi-lo)/mid*100.0
                bar_ranges = (w["High"] - w["Low"]).astype(float)
                median_bar_range = float(bar_ranges.median()) if len(bar_ranges) else 0.0
                # Tight range + no single abnormal expansion inside consolidation.
                if range_pct > BIGMOVE_MAX_RANGE_PCT:
                    continue
                if median_bar_range > atr * BIGMOVE_MAX_BAR_ATR_MULT:
                    continue
                # Prefer the latest qualifying range; allow only a small gap to breakout.
                gap = last_idx - end_idx
                if gap > 2:
                    continue
                score = (gap * 1000) + (BIGMOVE_CONSOLIDATION_MAX_BARS - bars) + range_pct
                if best is None or score < best[0]:
                    best = (score, start_idx, end_idx, bars, hi, lo, range_pct, (hi-lo)/atr if atr else 0.0)
        if best is None:
            return empty
        _, start_idx, end_idx, bars, hi, lo, range_pct, range_atr = best
        return {
            "found": True, "start_idx": int(start_idx), "end_idx": int(end_idx),
            "bars": int(bars), "high": hi, "low": lo,
            "range_pct": round(range_pct, 2), "range_atr": round(range_atr, 2),
        }
    except Exception:
        return empty


def detect_big_move_setup(df: pd.DataFrame) -> Dict[str, Any]:
    """Detect BIG BUY/BIG SELL from the latest consolidation breakout and HH/HL or LH/LL."""
    out = {
        "signal": "NO BIG MOVE", "direction": "NONE", "score": 0.0,
        "reason": "", "consolidation": detect_last_consolidation(df),
        "breakout_level": None, "breakout_move_pct": 0.0,
        "body_pct": 0.0, "body_atr": 0.0, "rvol": 0.0,
        "structure": "NONE", "hh_hl": False, "lh_ll": False,
        "current_move_pct": 0.0,
    }
    if df is None or len(df) < 25:
        out["reason"] = "Insufficient 5M candles"
        return out
    try:
        d = df.reset_index(drop=True).copy()
        c = detect_last_consolidation(d)
        if not c["found"]:
            out["reason"] = "No recent tight consolidation"
            return out

        last = d.iloc[-1]
        close = float(last["Close"]); open_ = float(last["Open"])
        high = float(last["High"]); low = float(last["Low"])
        body = abs(close-open_); rng = max(high-low, 0.0)
        atr_s = calculate_atr(d, 14)
        atr = float(atr_s.iloc[-1]) if pd.notna(atr_s.iloc[-1]) else max(close*0.005, 0.01)
        vol_avg = float(d["Volume"].iloc[-21:-1].mean()) if len(d) >= 22 else float(d["Volume"].iloc[:-1].mean())
        rvol = float(last["Volume"])/vol_avg if vol_avg > 0 else 0.0
        body_pct = body/rng*100 if rng else 0.0
        body_atr = body/atr if atr else 0.0

        buy_break = close > float(c["high"])
        sell_break = close < float(c["low"])
        breakout_level = c["high"] if buy_break else c["low"] if sell_break else None
        breakout_move_pct = abs(close-breakout_level)/breakout_level*100 if breakout_level else 0.0
        prev_close = float(d["Close"].iloc[-2])
        current_move_pct = abs(close-prev_close)/prev_close*100 if prev_close else 0.0

        # Latest confirmed HH/HL or LH/LL structure. The current breakout candle
        # is not treated as a confirmed pivot, so we never use an unfinished pivot.
        ph, pl = _bigmove_pivots(d)
        hh_hl = False; lh_ll = False
        if len(ph) >= 2 and len(pl) >= 2:
            hh_hl = ph[-1][1] > ph[-2][1] and pl[-1][1] > pl[-2][1]
            lh_ll = ph[-1][1] < ph[-2][1] and pl[-1][1] < pl[-2][1]
        structure = "HH/HL" if hh_hl else "LH/LL" if lh_ll else "NONE"

        # Score is evidence, not an EMA/RSI/AI score.
        score = 0.0
        reasons = []
        if buy_break or sell_break:
            score += 30; reasons.append("consolidation breakout")
        if body_atr >= BIGMOVE_MIN_BODY_ATR:
            score += 20; reasons.append(f"big candle {body_atr:.2f} ATR")
        if body_pct >= BIGMOVE_MIN_BODY_PCT:
            score += 10; reasons.append(f"body {body_pct:.0f}%")
        if rvol >= BIGMOVE_MIN_RVOL:
            score += 20; reasons.append(f"RVOL {rvol:.2f}x")
        if breakout_move_pct >= BIGMOVE_MIN_BREAK_PCT:
            score += 10; reasons.append(f"break {breakout_move_pct:.2f}%")
        if hh_hl or lh_ll:
            score += 10; reasons.append(structure)

        # BIG MOVE requires the directional breakout + strong candle + volume.
        core_ok = (body_atr >= BIGMOVE_MIN_BODY_ATR and body_pct >= BIGMOVE_MIN_BODY_PCT
                   and rvol >= BIGMOVE_MIN_RVOL and breakout_move_pct >= BIGMOVE_MIN_BREAK_PCT)
        if buy_break and core_ok:
            out["direction"] = "UP"
            out["structure"] = structure
            out["hh_hl"] = hh_hl
            out["signal"] = "🟢 BIG BUY" if score >= BIGMOVE_STRONG_SCORE else "🟢 BUY MOVE"
        elif sell_break and core_ok:
            out["direction"] = "DOWN"
            out["structure"] = structure
            out["lh_ll"] = lh_ll
            out["signal"] = "🔴 BIG SELL" if score >= BIGMOVE_STRONG_SCORE else "🔴 SELL MOVE"
        else:
            if not buy_break and not sell_break:
                reasons.append("no consolidation break")
            if body_atr < BIGMOVE_MIN_BODY_ATR:
                reasons.append(f"candle < {BIGMOVE_MIN_BODY_ATR:.2f} ATR")
            if body_pct < BIGMOVE_MIN_BODY_PCT:
                reasons.append(f"body < {BIGMOVE_MIN_BODY_PCT:.0f}%")
            if rvol < BIGMOVE_MIN_RVOL:
                reasons.append(f"RVOL < {BIGMOVE_MIN_RVOL:.2f}x")
            if breakout_move_pct < BIGMOVE_MIN_BREAK_PCT:
                reasons.append(f"break < {BIGMOVE_MIN_BREAK_PCT:.2f}%")

        out.update({
            "score": round(min(100.0, score), 1),
            "reason": " | ".join(reasons),
            "breakout_level": breakout_level,
            "breakout_move_pct": round(breakout_move_pct, 2),
            "body_pct": round(body_pct, 1),
            "body_atr": round(body_atr, 2),
            "rvol": round(rvol, 2),
            "current_move_pct": round(current_move_pct, 2),
            "consolidation": c,
        })
        return out
    except Exception as e:
        out["reason"] = f"BIG MOVE error: {type(e).__name__}"
        return out

def calculate_momentum_score(analysis_5m: Dict, analysis_15m: Dict, analysis_1h: Dict, movement_metrics: Dict, is_bullish: bool) -> Dict[str, Any]:
    """Calculate momentum score (0-100) based on multiple factors."""
    data_5m = analysis_5m.get("data", {})
    data_15m = analysis_15m.get("data", {})
    data_1h = analysis_1h.get("data", {})
    
    if not data_5m:
        return {"score": 0, "status": "INSUFFICIENT_DATA"}
    
    score = 50
    score_breakdown = {}
    
    try:
        # 1. PRICE MOVEMENT (20 points)
        move_5m = movement_metrics.get("move_5m_pct", 0)
        move_15m = movement_metrics.get("move_15m_pct", 0)
        
        if move_5m >= 0.5 and move_15m >= 1.0:
            score += 20
        elif move_5m >= 0.3 and move_15m >= 0.5:
            score += 15
        elif move_5m >= 0.1:
            score += 10
        else:
            score -= 5
        
        score_breakdown["price_movement"] = 20
        
        # 2. 5M MOMENTUM (15 points)
        tf_5m = data_5m.get("structure_trend", "NEUTRAL")
        if (is_bullish and tf_5m == "BULLISH") or (not is_bullish and tf_5m == "BEARISH"):
            score += 15
            score_breakdown["5m_momentum"] = 15
        else:
            score -= 10
            score_breakdown["5m_momentum"] = -10
        
        # 3. 15M TREND (15 points)
        tf_15m = data_15m.get("structure_trend", "NEUTRAL")
        if (is_bullish and tf_15m in ["BULLISH", "NEUTRAL"]) or (not is_bullish and tf_15m in ["BEARISH", "NEUTRAL"]):
            score += 15
            score_breakdown["15m_trend"] = 15
        elif (is_bullish and tf_15m == "BEARISH") or (not is_bullish and tf_15m == "BULLISH"):
            score -= 8
            score_breakdown["15m_trend"] = -8
        else:
            score += 7
            score_breakdown["15m_trend"] = 7
        
        # 4. 1H TREND (10 points)
        tf_1h = data_1h.get("structure_trend", "NEUTRAL")
        if (is_bullish and tf_1h in ["BULLISH", "NEUTRAL"]) or (not is_bullish and tf_1h in ["BEARISH", "NEUTRAL"]):
            score += 10
            score_breakdown["1h_trend"] = 10
        elif (is_bullish and tf_1h == "BEARISH") or (not is_bullish and tf_1h == "BULLISH"):
            score -= 5
            score_breakdown["1h_trend"] = -5
        
        # 5. VWAP (10 points)
        vwap = data_5m.get("vwap")
        price = data_5m.get("last_close", 0)
        if vwap is not None:
            if (is_bullish and price > vwap) or (not is_bullish and price < vwap):
                score += 10
                score_breakdown["vwap"] = 10
            else:
                score -= 5
                score_breakdown["vwap"] = -5
        
        # 6. EMA ALIGNMENT (10 points)
        ema_trend = data_5m.get("ema_trend", "NEUTRAL")
        if (is_bullish and ema_trend == "BULLISH") or (not is_bullish and ema_trend == "BEARISH"):
            score += 10
            score_breakdown["ema"] = 10
        elif (is_bullish and ema_trend == "BEARISH") or (not is_bullish and ema_trend == "BULLISH"):
            score -= 8
            score_breakdown["ema"] = -8
        else:
            score += 3
            score_breakdown["ema"] = 3
        
        # 7. BUYING/SELLING PRESSURE (10 points)
        pressure_trend = data_5m.get("pressure_trend", "NEUTRAL")
        
        if is_bullish:
            if pressure_trend == "STRONG_BUYING":
                score += 10
                score_breakdown["pressure"] = 10
            elif pressure_trend == "BUYING":
                score += 7
                score_breakdown["pressure"] = 7
            elif pressure_trend == "STRONG_SELLING":
                score -= 8
                score_breakdown["pressure"] = -8
            elif pressure_trend == "SELLING":
                score -= 3
                score_breakdown["pressure"] = -3
        else:
            if pressure_trend == "STRONG_SELLING":
                score += 10
                score_breakdown["pressure"] = 10
            elif pressure_trend == "SELLING":
                score += 7
                score_breakdown["pressure"] = 7
            elif pressure_trend == "STRONG_BUYING":
                score -= 8
                score_breakdown["pressure"] = -8
            elif pressure_trend == "BUYING":
                score -= 3
                score_breakdown["pressure"] = -3
        
        # 8. RVOL (10 points)
        rvol = data_5m.get("rvol", 1.0)
        if rvol >= 2.0:
            score += 10
            score_breakdown["rvol"] = 10
        elif rvol >= 1.5:
            score += 8
            score_breakdown["rvol"] = 8
        elif rvol >= 1.2:
            score += 5
            score_breakdown["rvol"] = 5
        elif rvol < 0.8:
            score -= 5
            score_breakdown["rvol"] = -5
        else:
            score_breakdown["rvol"] = 0
        
        score = max(0, min(100, score))
        
        if score >= 85:
            if is_bullish:
                status = "🟢 STRONG MOMENTUM BUY"
            else:
                status = "🔴 STRONG MOMENTUM SELL"
        elif score >= 70:
            if is_bullish:
                status = "🟢 MOMENTUM BUY"
            else:
                status = "🔴 MOMENTUM SELL"
        elif score >= 55:
            if is_bullish:
                status = "🟡 DEVELOPING BUY"
            else:
                status = "🟠 DEVELOPING SELL"
        else:
            status = "NONE"
        
        return {
            "score": round(score, 1),
            "status": status,
            "breakdown": score_breakdown,
        }
    
    except Exception as e:
        return {"score": 0, "status": "ERROR"}

# ════════════════════════════════════════════════════════════════════════════════
# NEXT CANDLE BIAS CALCULATION (NEW)
# ════════════════════════════════════════════════════════════════════════════════
def calculate_next_candle_bias(df, timeframe_analysis: Dict[str, Any]) -> Dict[str, Any]:
    """Calculate Next Candle Bias using weighted scoring from confirmed data."""
    if df is None or len(df) < 5:
        return {"bias": "NEUTRAL", "confidence": 0.0}
    
    try:
        data = timeframe_analysis.get("data", {})
        if not data:
            return {"bias": "NEUTRAL", "confidence": 0.0}
        
        buy_score = 50.0
        scores_count = 0
        
        last_close = float(df["Close"].iloc[-1])
        last_open = float(df["Open"].iloc[-1])
        
        if last_close > last_open:
            buy_score += 8
            scores_count += 1
        elif last_close < last_open:
            buy_score -= 8
            scores_count += 1
        
        pressure_trend = data.get("pressure_trend", "NEUTRAL")
        if pressure_trend == "STRONG_BUYING":
            buy_score += 12
        elif pressure_trend == "BUYING":
            buy_score += 8
        elif pressure_trend == "STRONG_SELLING":
            buy_score -= 12
        elif pressure_trend == "SELLING":
            buy_score -= 8
        scores_count += 1
        
        vwap = data.get("vwap")
        if vwap is not None:
            if last_close > vwap:
                buy_score += 6
            elif last_close < vwap:
                buy_score -= 6
            scores_count += 1
        
        ema_trend = data.get("ema_trend", "NEUTRAL")
        if ema_trend == "BULLISH":
            buy_score += 7
        elif ema_trend == "BEARISH":
            buy_score -= 7
        scores_count += 1
        
        rsi = data.get("rsi", 50)
        if rsi > 70:
            buy_score -= 4
        elif rsi < 30:
            buy_score += 4
        elif rsi > 60:
            buy_score += 3
        elif rsi < 40:
            buy_score -= 3
        scores_count += 1
        
        if data.get("macd_bullish"):
            buy_score += 5
        else:
            buy_score -= 5
        scores_count += 1
        
        structure_trend = data.get("structure_trend", "NEUTRAL")
        if structure_trend == "BULLISH":
            buy_score += 6
        elif structure_trend == "BEARISH":
            buy_score -= 6
        scores_count += 1
        
        if data.get("bullish_choch") or data.get("bullish_mss") or data.get("bullish_cisd"):
            buy_score += 10
        elif data.get("bearish_choch") or data.get("bearish_mss") or data.get("bearish_cisd"):
            buy_score -= 10
        scores_count += 1
        
        rvol = data.get("rvol", 1.0)
        if rvol > 1.5:
            if buy_score > 50:
                buy_score += 4
            else:
                buy_score -= 4
        elif rvol < 0.8:
            buy_score = buy_score * 0.95
        scores_count += 1
        
        buy_score = max(0, min(100, buy_score))
        confidence = round(abs(buy_score - 50) * 0.8, 1)
        confidence = max(0, min(100, confidence))
        
        if buy_score >= 65:
            bias = "🟢 BUY"
        elif buy_score <= 35:
            bias = "🔴 SELL"
        else:
            bias = "🟡 NEUTRAL"
        
        return {
            "bias": bias,
            "confidence": confidence,
            "score": buy_score,
        }
    
    except Exception as e:
        return {"bias": "NEUTRAL", "confidence": 0.0}

# ════════════════════════════════════════════════════════════════════════════════
# NEW INDICATORS (MULTI-TIMEFRAME)
# ════════════════════════════════════════════════════════════════════════════════
def calculate_vwap(df) -> pd.Series:
    """Calculate VWAP"""
    if "Volume" not in df.columns or len(df) == 0:
        return pd.Series([np.nan] * len(df), index=df.index)
    
    high = df["High"].astype(float)
    low = df["Low"].astype(float)
    close = df["Close"].astype(float)
    volume = df["Volume"].astype(float)
    
    typical_price = (high + low + close) / 3
    vwap = (typical_price * volume).cumsum() / volume.cumsum()
    return vwap.fillna(method="ffill").fillna(close)

def calculate_ema(close, period: int = 9) -> pd.Series:
    """Calculate EMA"""
    return close.ewm(span=period, adjust=False).mean()

def find_swing_highs_lows(df, lookback: int = SWING_LOOKBACK_PERIODS) -> Dict[str, Any]:
    """Return the latest confirmed swing high/low using only CLOSED candles."""
    empty = {"swing_high": None, "swing_high_idx": None,
             "swing_high_bars_ago": None, "swing_low": None,
             "swing_low_idx": None, "swing_low_bars_ago": None}
    if df is None or len(df) < 7:
        return empty
    d = df.tail(max(lookback, 7)).reset_index(drop=True).copy()
    highs, lows = d["High"].astype(float).to_numpy(), d["Low"].astype(float).to_numpy()
    pivot_highs, pivot_lows = [], []
    for i in range(2, len(d) - 2):
        if highs[i] > highs[i-1] and highs[i] >= highs[i-2] and highs[i] > highs[i+1] and highs[i] >= highs[i+2]:
            pivot_highs.append(i)
        if lows[i] < lows[i-1] and lows[i] <= lows[i-2] and lows[i] < lows[i+1] and lows[i] <= lows[i+2]:
            pivot_lows.append(i)
    if pivot_highs:
        i = pivot_highs[-1]
        empty.update(swing_high=float(highs[i]), swing_high_idx=int(i), swing_high_bars_ago=int(len(d)-1-i))
    if pivot_lows:
        i = pivot_lows[-1]
        empty.update(swing_low=float(lows[i]), swing_low_idx=int(i), swing_low_bars_ago=int(len(d)-1-i))
    return empty

def _confirmed_pivots(df, left: int = 2, right: int = 2):
    """Return confirmed pivot highs/lows."""
    if df is None or len(df) < left + right + 3:
        return [], []
    d = df.reset_index(drop=True)
    highs, lows = d["High"].astype(float).to_numpy(), d["Low"].astype(float).to_numpy()
    ph, pl = [], []
    for i in range(left, len(d) - right):
        if highs[i] >= max(highs[i-left:i]) and highs[i] > max(highs[i+1:i+right+1]):
            ph.append((i, float(highs[i])))
        if lows[i] <= min(lows[i-left:i]) and lows[i] < min(lows[i+1:i+right+1]):
            pl.append((i, float(lows[i])))
    return ph, pl

def detect_structure(df) -> Dict[str, Any]:
    """Detect HH/HL/LH/LL from confirmed pivots only."""
    ph, pl = _confirmed_pivots(df)
    result = {"type":"UNKNOWN", "trend":"NEUTRAL", "current_high":None,
              "current_low":None, "prev_high":None, "prev_low":None, "strength": 0}
    if len(ph) >= 2:
        result["prev_high"], result["current_high"] = ph[-2][1], ph[-1][1]
    if len(pl) >= 2:
        result["prev_low"], result["current_low"] = pl[-2][1], pl[-1][1]
    if len(ph) >= 2 and len(pl) >= 2:
        hh = ph[-1][1] > ph[-2][1]
        hl = pl[-1][1] > pl[-2][1]
        lh = ph[-1][1] < ph[-2][1]
        ll = pl[-1][1] < pl[-2][1]
        if hh and hl:
            result["type"], result["trend"] = "HH/HL", "BULLISH"
            result["strength"] = min(100, abs((ph[-1][1] - ph[-2][1]) / ph[-2][1] * 100) * 10)
        elif lh and ll:
            result["type"], result["trend"] = "LH/LL", "BEARISH"
            result["strength"] = min(100, abs((pl[-1][1] - pl[-2][1]) / pl[-2][1] * 100) * 10)
        elif hh:
            result["type"], result["trend"] = "HH", "BULLISH"
            result["strength"] = 60
        elif ll:
            result["type"], result["trend"] = "LL", "BEARISH"
            result["strength"] = 60
        elif hl:
            result["type"], result["trend"] = "HL", "BULLISH"
            result["strength"] = 40
        elif lh:
            result["type"], result["trend"] = "LH", "BEARISH"
            result["strength"] = 40
    return result

def detect_choch(df) -> Dict[str, Any]:
    """Detect a NEW confirmed CHoCH on the latest CLOSED candle only."""
    ph, pl = _confirmed_pivots(df)
    out = {"bullish_choch":False, "bearish_choch":False, "choch_price":None,
           "choch_type":"NONE", "confirmation":"NONE"}
    if len(ph) < 2 or len(pl) < 2 or len(df) < 10:
        out["confirmation"] = "PENDING"
        return out
    prev_close = float(df["Close"].iloc[-2])
    close = float(df["Close"].iloc[-1])
    bearish_structure = ph[-1][1] < ph[-2][1] and pl[-1][1] < pl[-2][1]
    bullish_structure = ph[-1][1] > ph[-2][1] and pl[-1][1] > pl[-2][1]
    
    min_move = max(ph[-1][1] * 0.001, 0.5)
    
    if bearish_structure and prev_close <= ph[-1][1] and close > (ph[-1][1] + min_move):
        out.update(bullish_choch=True, choch_price=ph[-1][1], choch_type="BULLISH_CHoCH", confirmation="CONFIRMED")
    elif bullish_structure and prev_close >= pl[-1][1] and close < (pl[-1][1] - min_move):
        out.update(bearish_choch=True, choch_price=pl[-1][1], choch_type="BEARISH_CHoCH", confirmation="CONFIRMED")
    return out

def detect_cisd(df: pd.DataFrame) -> Dict[str, Any]:
    """Detect a NEW confirmed CISD event on the latest CLOSED candle only."""
    result = {"bullish_cisd": False, "bearish_cisd": False, "cisd_type": "NONE", "cisd_price": None}
    if df is None or len(df) < 5:
        return result
    d = df.reset_index(drop=True)
    last = d.iloc[-1]
    prev_close = float(d["Close"].iloc[-2])
    prior = d.iloc[:-1]
    
    bearish = prior[prior["Close"] < prior["Open"]].tail(3)
    bullish = prior[prior["Close"] > prior["Open"]].tail(3)
    
    min_move = float(last["Close"]) * 0.001
    
    if not bearish.empty:
        level = float(bearish.iloc[-1]["High"])
        if prev_close <= level and float(last["Close"]) > (level + min_move):
            result.update(bullish_cisd=True, cisd_type="BULLISH_CISD", cisd_price=float(last["Close"]))
    
    if not bullish.empty and not result["bullish_cisd"]:
        level = float(bullish.iloc[-1]["Low"])
        if prev_close >= level and float(last["Close"]) < (level - min_move):
            result.update(bearish_cisd=True, cisd_type="BEARISH_CISD", cisd_price=float(last["Close"]))
    return result

def detect_mss(df) -> Dict[str, Any]:
    """Detect a NEW confirmed MSS event on the latest CLOSED candle only."""
    ph, pl = _confirmed_pivots(df)
    out = {"bullish_mss":False, "bearish_mss":False, "mss_type":"NONE", "confirmation":"NONE"}
    if len(ph) < 2 or len(pl) < 2 or len(df) < 10:
        out["confirmation"] = "PENDING"
        return out
    prev_close = float(df["Close"].iloc[-2])
    close = float(df["Close"].iloc[-1])
    bearish_structure = ph[-1][1] < ph[-2][1] and pl[-1][1] < pl[-2][1]
    bullish_structure = ph[-1][1] > ph[-2][1] and pl[-1][1] > pl[-2][1]
    
    min_move = max(ph[-1][1] * 0.001, 0.5)
    
    if bearish_structure and prev_close <= ph[-1][1] and close > (ph[-1][1] + min_move):
        out.update(bullish_mss=True, mss_type="BULLISH_MSS", confirmation="CONFIRMED")
    elif bullish_structure and prev_close >= pl[-1][1] and close < (pl[-1][1] - min_move):
        out.update(bearish_mss=True, mss_type="BEARISH_MSS", confirmation="CONFIRMED")
    return out

# ════════════════════════════════════════════════════════════════════════════════
# SYMBOL LOADING (RETAINED WITH ENHANCEMENT FOR F&O)
# ════════════════════════════════════════════════════════════════════════════════
_VALID_EQ_SYMBOL_RE = re.compile(r"^NSE:[A-Z0-9&\-]+-EQ$")
_FO_EQUITY_PATTERN = re.compile(r"^NSE:[A-Z0-9&\-]+-EQ$")

def _validate_symbols(symbols) -> List[str]:
    seen = set()
    valid = []
    for s in symbols:
        if not isinstance(s, str):
            continue
        s = s.strip().upper()
        if not s or s in seen:
            continue
        if not _VALID_EQ_SYMBOL_RE.match(s):
            continue
        seen.add(s)
        valid.append(s)
    return valid

@st.cache_data(ttl=60 * 60 * 12)
def load_nse_equity_symbols() -> List[str]:
    try:
        resp = requests.get(FYERS_NSE_CM_SYMBOL_MASTER, timeout=20)
        resp.raise_for_status()
    except Exception as e:
        st.error(f"Could not download Fyers symbol master: {e}")
        return []
    lines = [ln for ln in resp.text.strip().split("\n") if ln.strip()]
    if not lines:
        return []
    sample = lines[:min(500, len(lines))]
    split_sample = [ln.split(",") for ln in sample]
    max_cols = max((len(p) for p in split_sample), default=0)
    best_col, best_hits = None, 0
    for col_idx in range(max_cols):
        hits = sum(1 for parts in split_sample if len(parts) > col_idx and parts[col_idx].strip().startswith("NSE:") and parts[col_idx].strip().endswith("-EQ"))
        if hits > best_hits:
            best_col, best_hits = col_idx, hits
    if best_col is None or best_hits == 0:
        st.error("Could not locate trading-symbol column.")
        return []
    symbols = []
    for line in lines:
        parts = line.split(",")
        if len(parts) <= best_col:
            continue
        sym = parts[best_col].strip()
        if sym.startswith("NSE:") and sym.endswith("-EQ"):
            symbols.append(sym)
    return sorted(set(_validate_symbols(symbols)))

@st.cache_data(ttl=60 * 60 * 12)
def load_fo_stocks() -> List[str]:
    """Load NSE equity symbols that have active equity-derivative contracts."""
    try:
        cm_symbols = set(load_nse_equity_symbols())
        if not cm_symbols:
            return []

        urls = [
            "https://public.fyers.in/sym_details/NSE_FO.csv",
            "http://public.fyers.in/sym_details/NSE_FO.csv",
        ]
        text = None
        last_error = None
        for url in urls:
            try:
                r = requests.get(
                    url,
                    timeout=30,
                    headers={"User-Agent": "Mozilla/5.0", "Accept": "text/csv,*/*"},
                )
                r.raise_for_status()
                if r.text and len(r.text) > 100:
                    text = r.text
                    break
            except Exception as exc:
                last_error = exc

        if not text:
            logging.warning("F&O symbol master unavailable: %s", last_error)
            return []

        fo_underlyings = set()
        reader = csv.reader(io.StringIO(text))
        for row in reader:
            if len(row) < 14:
                continue

            short_sym = str(row[13]).strip().upper()
            contract_symbol = str(row[9]).strip().upper() if len(row) > 9 else ""
            exchange = str(row[10]).strip() if len(row) > 10 else ""

            if exchange not in {"10", "NSE"}:
                continue
            if not short_sym or short_sym in {"NONE", "NAN"}:
                continue

            if contract_symbol.startswith("NSE:") and short_sym:
                candidate = f"NSE:{short_sym}-EQ"
                if candidate in cm_symbols:
                    fo_underlyings.add(candidate)

        result = sorted(fo_underlyings)
        logging.info("Loaded %d NSE F&O equity underlyings", len(result))
        return result

    except Exception as e:
        logging.exception("F&O symbol loading failed: %s", e)
        return []

# ════════════════════════════════════════════════════════════════════════════════
# SAFE HISTORY FETCH (RETAINED)
# ════════════════════════════════════════════════════════════════════════════════
_HISTORY_MAX_RETRIES = 3
_HISTORY_BASE_DELAY_SECONDS = 1.0

def _safe_history(fyers, params: dict, max_retries: int = _HISTORY_MAX_RETRIES, base_delay: float = _HISTORY_BASE_DELAY_SECONDS):
    symbol = params.get("symbol", "UNKNOWN")
    last_err = "unknown error"
    for attempt in range(1, max_retries + 1):
        try:
            resp = fyers.history(params)
        except requests.exceptions.Timeout:
            last_err = "timeout"
        except requests.exceptions.ConnectionError:
            last_err = "network error"
        except requests.exceptions.RequestException as e:
            last_err = f"request error: {e}"
        except (ValueError, TypeError) as e:
            last_err = f"invalid response: {e}"
        except Exception as e:
            last_err = f"unexpected error: {e}"
        else:
            if not isinstance(resp, dict):
                last_err = "empty/invalid response"
            else:
                status = resp.get("s")
                if status == "ok":
                    candles = resp.get("candles")
                    if not isinstance(candles, list):
                        last_err = "malformed candle data"
                    else:
                        return resp, None
                else:
                    message = str(resp.get("message", status or "unknown"))
                    if "rate" in message.lower() or "limit" in message.lower():
                        last_err = f"rate limited: {message}"
                        time.sleep(base_delay * attempt * 2)
                        continue
                    return None, message
        if attempt < max_retries:
            time.sleep(base_delay * attempt)
    return None, f"{symbol}: {last_err} (after {max_retries} attempts)"

# ════════════════════════════════════════════════════════════════════════════════
# SCAN STATS (ORIGINAL)
# ════════════════════════════════════════════════════════════════════════════════
class ScanStats:
    def __init__(self, total: int):
        self.total = total
        self.scanned = 0
        self.successful = 0
        self.skipped = 0
        self.failed = 0
        self._start = time.time()

    def record(self, has_result: bool, has_error: bool) -> None:
        self.scanned += 1
        if has_result:
            self.successful += 1
        elif has_error:
            self.failed += 1
        else:
            self.skipped += 1

    @property
    def elapsed_seconds(self) -> float:
        return time.time() - self._start

def _display_scan_summary(stats: "ScanStats") -> None:
    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Total Stocks", stats.total)
    c2.metric("Scanned", stats.scanned)
    c3.metric("Successful", stats.successful)
    c4.metric("Skipped", stats.skipped)
    c5.metric("Failed", stats.failed)
    c6.metric("Scan Time", f"{stats.elapsed_seconds:.1f}s")

# ════════════════════════════════════════════════════════════════════════════════
# TIMEFRAME DATA FETCHER (NEW)
# ════════════════════════════════════════════════════════════════════════════════
def _fetch_timeframe_data(fyers, symbol, resolution: str, lookback_days: int = 30) -> Optional[pd.DataFrame]:
    """Fetch OHLCV data for a specific timeframe."""
    date_from = (datetime.today() - timedelta(days=lookback_days)).strftime("%Y-%m-%d")
    date_to = datetime.today().strftime("%Y-%m-%d")
    
    resp, err = _safe_history(fyers, {
        "symbol": symbol,
        "resolution": resolution,
        "date_format": "1",
        "range_from": date_from,
        "range_to": date_to,
        "cont_flag": "1",
    })
    
    if err or not resp:
        return None
    
    candles = resp.get("candles")
    if not candles or len(candles) < 10:
        return None
    
    try:
        df = pd.DataFrame(candles, columns=["Time", "Open", "High", "Low", "Close", "Volume"])
        df["Time"] = pd.to_datetime(df["Time"], unit="s", utc=True).dt.tz_convert("Asia/Kolkata")
        df[["Open", "High", "Low", "Close", "Volume"]] = df[["Open", "High", "Low", "Close", "Volume"]].apply(pd.to_numeric, errors="coerce")
        df = df.dropna(subset=["Open", "High", "Low", "Close"]).sort_values("Time").reset_index(drop=True)
        
        if len(df) < 10:
            return None
        
        if len(df) > 1:
            last_time = df["Time"].iloc[-1]
            candle_age = (_now_ist() - last_time).total_seconds() / 60
            res_minutes = int(resolution)
            if candle_age < res_minutes + 1:
                df = df.iloc[:-1].reset_index(drop=True)
        
        if len(df) < 10:
            return None
        
        return df
    
    except Exception as e:
        return None

# ════════════════════════════════════════════════════════════════════════════════
# TIMEFRAME ANALYSIS ENGINE (NEW - WITH PRESSURE)
# ════════════════════════════════════════════════════════════════════════════════
def analyze_timeframe(fyers, symbol: str, resolution: str) -> Dict[str, Any]:
    """Analyze a specific timeframe."""
    df = _fetch_timeframe_data(fyers, symbol, resolution, lookback_days=30)
    
    if df is None or len(df) < 10:
        return {
            "timeframe": resolution,
            "status": "DATA_UNAVAILABLE",
            "data": None,
        }
    
    try:
        rsi = calculate_rsi(df["Close"])
        macd_line, macd_sig, macd_hist = calculate_macd(df["Close"])
        atr = calculate_atr(df)
        vwap = calculate_vwap(df)
        ema9 = calculate_ema(df["Close"], 9)
        ema21 = calculate_ema(df["Close"], 21)
        ema50 = calculate_ema(df["Close"], 50)
        ema200 = calculate_ema(df["Close"], 200)
        
        pressure = calculate_buying_selling_pressure(df)
        
        structure = detect_structure(df)
        choch = detect_choch(df)
        mss = detect_mss(df)
        cisd = detect_cisd(df)
        swings = find_swing_highs_lows(df)
        
        vol_avg20 = float(df["Volume"].tail(20).mean()) if "Volume" in df.columns else 0
        last_vol = float(df["Volume"].iloc[-1]) if "Volume" in df.columns else 0
        rvol = round(last_vol / vol_avg20, 2) if vol_avg20 > 0 else 0.0
        
        last_close = float(df["Close"].iloc[-1])
        last_high = float(df["High"].iloc[-1])
        last_low = float(df["Low"].iloc[-1])
        last_open = float(df["Open"].iloc[-1])
        
        ema_trend = "BULLISH" if ema9.iloc[-1] > ema21.iloc[-1] > ema50.iloc[-1] else "BEARISH" if ema9.iloc[-1] < ema21.iloc[-1] < ema50.iloc[-1] else "NEUTRAL"
        
        rsi_val = float(rsi.iloc[-1])
        rsi_overbought = rsi_val > 70
        rsi_oversold = rsi_val < 30
        
        macd_bullish = macd_line.iloc[-1] > macd_sig.iloc[-1]
        candle_date, candle_close_time = _candle_signal_timestamp(df, is_daily=False, resolution=resolution)
        
        return {
            "timeframe": resolution,
            "status": "OK",
            "data": {
                "last_close": last_close,
                "last_high": last_high,
                "last_low": last_low,
                "last_open": last_open,
                "signal_candle_date": candle_date,
                "signal_candle_time": candle_close_time,
                "signal_generated_at": _generated_timestamp(),
                "structure_type": structure["type"],
                "structure_trend": structure["trend"],
                "structure_strength": round(structure.get("strength", 0), 1),
                "current_high": structure["current_high"],
                "current_low": structure["current_low"],
                "prev_high": structure["prev_high"],
                "prev_low": structure["prev_low"],
                "bullish_choch": choch["bullish_choch"],
                "bearish_choch": choch["bearish_choch"],
                "choch_price": choch["choch_price"],
                "choch_type": choch["choch_type"],
                "bullish_mss": mss["bullish_mss"],
                "bearish_mss": mss["bearish_mss"],
                "mss_type": mss["mss_type"],
                "bullish_cisd": cisd["bullish_cisd"],
                "bearish_cisd": cisd["bearish_cisd"],
                "cisd_type": cisd["cisd_type"],
                "cisd_price": cisd["cisd_price"],
                "swing_high": swings["swing_high"],
                "swing_low": swings["swing_low"],
                "swing_high_bars_ago": swings["swing_high_bars_ago"],
                "swing_low_bars_ago": swings["swing_low_bars_ago"],
                "vwap": float(vwap.iloc[-1]) if len(vwap) > 0 else None,
                "ema9": float(ema9.iloc[-1]) if len(ema9) > 0 else None,
                "ema21": float(ema21.iloc[-1]) if len(ema21) > 0 else None,
                "ema50": float(ema50.iloc[-1]) if len(ema50) > 0 else None,
                "ema200": float(ema200.iloc[-1]) if len(ema200) > 0 else None,
                "ema_trend": ema_trend,
                "rsi": round(rsi_val, 1),
                "rsi_overbought": rsi_overbought,
                "rsi_oversold": rsi_oversold,
                "macd_bullish": macd_bullish,
                "macd_value": round(float(macd_line.iloc[-1]), 4),
                "rvol": rvol,
                "atr": round(float(atr.iloc[-1]), 2),
                "buying_pressure": pressure["buying_pressure"],
                "selling_pressure": pressure["selling_pressure"],
                "pressure_trend": pressure["trend"],
                "buying_volume": pressure["buying_volume"],
                "selling_volume": pressure["selling_volume"],
                "pressure_ratio": pressure["pressure_ratio"],
            },
            "df": df,
        }
    
    except Exception as e:
        return {
            "timeframe": resolution,
            "status": "ERROR",
            "error": str(e),
            "data": None,
        }

# ════════════════════════════════════════════════════════════════════════════════
# OPTIONS CHAIN ANALYSIS (NEW)
# ════════════════════════════════════════════════════════════════════════════════
def _fyers_optionchain_request(fyers, symbol: str, strikecount: int = OPTIONS_STRIKE_COUNT, timestamp: str = "", greeks: bool = True):
    """Call FYERS v3 option-chain endpoint directly, with SDK fallback."""
    app_id = getattr(fyers, "client_id", None) or FYERS_APP_ID
    token = getattr(fyers, "token", None) or os.environ.get("FYERS_ACCESS_TOKEN", "")
    if app_id and token:
        headers = {"Authorization": f"{app_id}:{token}"}
        params = {"symbol": symbol, "strikecount": min(int(strikecount), 50), "greeks": "1" if greeks else "0"}
        if timestamp:
            params["timestamp"] = str(timestamp)
        r = requests.get("https://api-t1.fyers.in/data/options-chain-v3", headers=headers, params=params, timeout=OPTIONS_HTTP_TIMEOUT)
        r.raise_for_status()
        return r.json()
    fn = getattr(fyers, "option_chain", None) or getattr(fyers, "optionchain", None)
    if fn:
        try:
            return fn(symbol=symbol, strikecount=min(int(strikecount), 50), timestamp=timestamp, greeks=greeks)
        except TypeError:
            return fn(data={"symbol":symbol, "strikecount":min(int(strikecount),50), "timestamp":timestamp, "greeks":greeks})
    raise RuntimeError("FYERS option-chain API is not available")

def _calculate_max_pain(chain_rows):
    strikes = sorted({float(x.get("strike_price")) for x in chain_rows if x.get("option_type") in ("CE", "PE") and x.get("strike_price") not in (None, -1)})
    if not strikes:
        return None
    ce = {(float(x.get("strike_price")), float(x.get("oi",0) or 0)) for x in chain_rows if x.get("option_type") == "CE"}
    pe = {(float(x.get("strike_price")), float(x.get("oi",0) or 0)) for x in chain_rows if x.get("option_type") == "PE"}
    ce_map, pe_map = dict(ce), dict(pe)
    pains = {}
    for settle in strikes:
        pain = sum(max(settle-k,0)*oi for k,oi in ce_map.items()) + sum(max(k-settle,0)*oi for k,oi in pe_map.items())
        pains[settle] = pain
    return min(pains, key=pains.get) if pains else None

def fetch_options_chain_data(fyers, symbol: str, expiry_timestamp: str = "") -> Dict[str, Any]:
    """Fetch live FYERS v3 options chain; never fabricate missing data."""
    empty = {"status":"DATA_UNAVAILABLE", "message":"No live option-chain data", "expiry":None, "spot":None,
             "atm_strike":None, "ce_oi":None, "pe_oi":None, "ce_oi_change":None, "pe_oi_change":None,
             "pcr":None, "max_pain":None, "ce_volume":None, "pe_volume":None, "call_writing":False,
             "put_writing":False, "call_unwinding":False, "put_unwinding":False, "options_bias":"NEUTRAL", "chain":[], "expiry_data":[]}
    try:
        resp = _fyers_optionchain_request(fyers, symbol, timestamp=expiry_timestamp)
        if not isinstance(resp, dict) or resp.get("s") not in (None, "ok"):
            empty["message"] = str(resp.get("message", "Option-chain request failed")) if isinstance(resp, dict) else "Invalid response"
            return empty
        data = resp.get("data", resp)
        chain = data.get("optionsChain") or []
        expiry_data = data.get("expiryData") or []
        if not expiry_timestamp and expiry_data:
            nearest = expiry_data[0]
            expiry_timestamp = str(nearest.get("expiry", ""))
            if expiry_timestamp:
                resp = _fyers_optionchain_request(fyers, symbol, timestamp=expiry_timestamp)
                data = resp.get("data", resp) if isinstance(resp, dict) else {}
                chain = data.get("optionsChain") or chain
        spot_row = next((x for x in chain if x.get("option_type", "") == ""), None)
        spot = float(spot_row.get("ltp")) if spot_row and spot_row.get("ltp") is not None else None
        legs = [x for x in chain if x.get("option_type") in ("CE","PE") and x.get("strike_price") not in (None,-1)]
        strikes = sorted({float(x["strike_price"]) for x in legs})
        atm = min(strikes, key=lambda k: abs(k-spot)) if strikes and spot is not None else None
        call_oi = float(data.get("callOi", 0) or 0); put_oi = float(data.get("putOi", 0) or 0)
        call_chg = sum(float(x.get("oich",0) or 0) for x in legs if x.get("option_type")=="CE")
        put_chg = sum(float(x.get("oich",0) or 0) for x in legs if x.get("option_type")=="PE")
        ce_vol = sum(float(x.get("volume",0) or 0) for x in legs if x.get("option_type")=="CE")
        pe_vol = sum(float(x.get("volume",0) or 0) for x in legs if x.get("option_type")=="PE")
        near = [x for x in legs if atm is not None and abs(float(x["strike_price"])-atm) <= max((strikes[1]-strikes[0]) if len(strikes)>1 else 1, 1)*3]
        ce_writing = sum(1 for x in near if x.get("option_type")=="CE" and float(x.get("oich",0) or 0)>0 and float(x.get("ltpch",0) or 0)<0)
        pe_writing = sum(1 for x in near if x.get("option_type")=="PE" and float(x.get("oich",0) or 0)>0 and float(x.get("ltpch",0) or 0)<0)
        ce_unwind = sum(1 for x in near if x.get("option_type")=="CE" and float(x.get("oich",0) or 0)<0 and float(x.get("ltpch",0) or 0)>0)
        pe_unwind = sum(1 for x in near if x.get("option_type")=="PE" and float(x.get("oich",0) or 0)<0 and float(x.get("ltpch",0) or 0)>0)
        pcr = put_oi/call_oi if call_oi > 0 else None
        bull = (pcr is not None and pcr >= 1.05) or pe_writing > ce_writing
        bear = (pcr is not None and pcr <= 0.80) or ce_writing > pe_writing
        if bull and not bear: bias = "🟢 BULLISH"
        elif bear and not bull: bias = "🔴 BEARISH"
        else: bias = "🟡 NEUTRAL"
        return {"status":"OK", "message":"Live FYERS option chain", "expiry": expiry_timestamp or (expiry_data[0].get("date") if expiry_data else None),
                "spot":spot, "atm_strike":atm, "ce_oi":call_oi, "pe_oi":put_oi, "ce_oi_change":call_chg,
                "pe_oi_change":put_chg, "pcr":round(pcr,2) if pcr is not None else None,
                "max_pain":_calculate_max_pain(legs), "ce_volume":ce_vol, "pe_volume":pe_vol,
                "call_writing":ce_writing>0, "put_writing":pe_writing>0, "call_unwinding":ce_unwind>0,
                "put_unwinding":pe_unwind>0, "options_bias":bias, "chain":legs, "expiry_data":expiry_data}
    except Exception as e:
        empty["status"] = "ERROR"; empty["message"] = str(e)[:180]
        return empty

# ════════════════════════════════════════════════════════════════════════════════
# STRICT MULTI-TIMEFRAME SIGNAL VALIDATION ENGINE (FIXED)
# ════════════════════════════════════════════════════════════════════════════════
def calculate_master_signal(symbol: str, analysis_5m: Dict, analysis_15m: Dict, analysis_1h: Dict, options_data: Dict = None) -> Dict[str, Any]:
    """STRICT master signal validation engine."""
    if options_data is None:
        options_data = {"status":"DATA_UNAVAILABLE", "options_bias":"NEUTRAL"}
    
    data_5m = analysis_5m.get("data") if analysis_5m.get("status") == "OK" else None
    data_15m = analysis_15m.get("data") if analysis_15m.get("status") == "OK" else None
    data_1h = analysis_1h.get("data") if analysis_1h.get("status") == "OK" else None
    
    if not data_5m or not data_15m or not data_1h:
        return {
            "final_signal": "🟡 NEUTRAL",
            "confidence": 0.0,
            "entry": None,
            "stop_loss": None,
            "target1": None,
            "target2": None,
            "rr_ratio": None,
            "scores": {"5m": 0, "15m": 0, "1h": 0, "pressure": 0, "options": 0},
            "signal_reason": "Insufficient timeframe data",
        }
    
    def classify_tf_direction(data: Dict) -> str:
        trend = data.get("structure_trend", "NEUTRAL")
        if trend == "BULLISH":
            return "BULLISH"
        elif trend == "BEARISH":
            return "BEARISH"
        else:
            return "NEUTRAL"
    
    tf_5m = classify_tf_direction(data_5m)
    tf_15m = classify_tf_direction(data_15m)
    tf_1h = classify_tf_direction(data_1h)
    
    is_bullish_aligned = (
        (tf_5m == "BULLISH") and
        (tf_15m in ["BULLISH", "NEUTRAL"]) and
        (tf_1h != "BEARISH")
    )
    
    is_bearish_aligned = (
        (tf_5m == "BEARISH") and
        (tf_15m in ["BEARISH", "NEUTRAL"]) and
        (tf_1h != "BULLISH")
    )
    
    hard_conflict = (tf_5m == "BULLISH" and tf_1h == "BEARISH") or (tf_5m == "BEARISH" and tf_1h == "BULLISH")
    
    if hard_conflict or (not is_bullish_aligned and not is_bearish_aligned):
        return {
            "final_signal": "🟡 NEUTRAL",
            "confidence": 20.0 if hard_conflict else 35.0,
            "entry": None,
            "stop_loss": None,
            "target1": None,
            "target2": None,
            "rr_ratio": None,
            "scores": {
                "5m": 50,
                "15m": 50,
                "1h": 50,
                "pressure": 50,
                "options": 50,
            },
            "signal_reason": f"Timeframe conflict: 5M {tf_5m} vs 15M {tf_15m} vs 1H {tf_1h}",
        }
    
    pressure_trend = data_5m.get("pressure_trend", "NEUTRAL")
    pressure_buy = pressure_trend in ["BUYING", "STRONG_BUYING"]
    pressure_sell = pressure_trend in ["SELLING", "STRONG_SELLING"]
    pressure_strong_buy = pressure_trend == "STRONG_BUYING"
    pressure_strong_sell = pressure_trend == "STRONG_SELLING"
    
    if is_bullish_aligned and not pressure_buy:
        is_bullish_aligned = False
    
    if is_bearish_aligned and not pressure_sell:
        is_bearish_aligned = False
    
    if not is_bullish_aligned and not is_bearish_aligned:
        return {
            "final_signal": "🟡 NEUTRAL",
            "confidence": 30.0,
            "entry": None,
            "stop_loss": None,
            "target1": None,
            "target2": None,
            "rr_ratio": None,
            "scores": {
                "5m": 50,
                "15m": 50,
                "1h": 50,
                "pressure": 50,
                "options": 50,
            },
            "signal_reason": f"Pressure conflict: {pressure_trend}",
        }
    
    last_close = data_5m.get("last_close", 0)
    vwap = data_5m.get("vwap")
    
    if vwap is not None:
        price_above_vwap = last_close > vwap
        
        if is_bullish_aligned and not price_above_vwap:
            if not (data_5m.get("bullish_choch") or data_5m.get("bullish_mss")):
                is_bullish_aligned = False
        
        if is_bearish_aligned and price_above_vwap:
            if not (data_5m.get("bearish_choch") or data_5m.get("bearish_mss")):
                is_bearish_aligned = False
    
    if not is_bullish_aligned and not is_bearish_aligned:
        return {
            "final_signal": "🟡 NEUTRAL",
            "confidence": 30.0,
            "entry": None,
            "stop_loss": None,
            "target1": None,
            "target2": None,
            "rr_ratio": None,
            "scores": {
                "5m": 50,
                "15m": 50,
                "1h": 50,
                "pressure": 50,
                "options": 50,
            },
            "signal_reason": f"VWAP conflict: Price {last_close:.2f} vs VWAP {vwap:.2f if vwap else 'N/A'}",
        }
    
    ema_trend = data_5m.get("ema_trend", "NEUTRAL")
    
    if is_bullish_aligned and ema_trend == "BEARISH":
        is_bullish_aligned = False
    
    if is_bearish_aligned and ema_trend == "BULLISH":
        is_bearish_aligned = False
    
    if not is_bullish_aligned and not is_bearish_aligned:
        return {
            "final_signal": "🟡 NEUTRAL",
            "confidence": 30.0,
            "entry": None,
            "stop_loss": None,
            "target1": None,
            "target2": None,
            "rr_ratio": None,
            "scores": {
                "5m": 50,
                "15m": 50,
                "1h": 50,
                "pressure": 50,
                "options": 50,
            },
            "signal_reason": f"EMA conflict: Trend {ema_trend} vs Structure {tf_5m}",
        }
    
    confirmation_count = 0
    total_factors = 0
    
    if tf_5m == "BULLISH":
        confirmation_count += 1
    elif tf_5m == "BEARISH":
        confirmation_count -= 1
    total_factors += 1
    
    if tf_15m == "BULLISH":
        confirmation_count += 1
    elif tf_15m == "BEARISH":
        confirmation_count -= 1
    total_factors += 1
    
    if tf_1h == "BULLISH":
        confirmation_count += 1
    elif tf_1h == "BEARISH":
        confirmation_count -= 1
    total_factors += 1
    
    if pressure_buy:
        confirmation_count += 1
    elif pressure_sell:
        confirmation_count -= 1
    total_factors += 1
    
    if vwap is not None:
        if is_bullish_aligned and price_above_vwap:
            confirmation_count += 1
        elif is_bearish_aligned and not price_above_vwap:
            confirmation_count += 1
        total_factors += 1
    
    if ema_trend == "BULLISH":
        confirmation_count += 1
    elif ema_trend == "BEARISH":
        confirmation_count -= 1
    total_factors += 1
    
    has_bullish_structure = (
        data_5m.get("bullish_choch") or 
        data_5m.get("bullish_mss") or 
        data_5m.get("bullish_cisd") or
        (data_5m.get("structure_type") in ["HH/HL", "HL"])
    )
    has_bearish_structure = (
        data_5m.get("bearish_choch") or 
        data_5m.get("bearish_mss") or 
        data_5m.get("bearish_cisd") or
        (data_5m.get("structure_type") in ["LH/LL", "LH"])
    )
    
    if is_bullish_aligned and has_bullish_structure:
        confirmation_count += 1
    elif is_bearish_aligned and has_bearish_structure:
        confirmation_count += 1
    total_factors += 1
    
    rvol = data_5m.get("rvol", 1.0)
    if rvol >= 1.2:
        confirmation_count += 1
    elif rvol < 0.8:
        confirmation_count -= 1
    total_factors += 1
    
    raw_confidence = (confirmation_count / total_factors) * 100 if total_factors > 0 else 0
    confidence = max(0, min(100, abs(raw_confidence)))
    
    df_5m = analysis_5m.get("df")
    if df_5m is not None:
        next_bias = calculate_next_candle_bias(df_5m, analysis_5m)
    else:
        next_bias = {"bias": "NEUTRAL", "confidence": 0.0}
    
    next_bias_str = next_bias.get("bias", "NEUTRAL")
    next_bias_conf = next_bias.get("confidence", 0.0)
    
    if is_bullish_aligned and "SELL" in next_bias_str and next_bias_conf >= 70:
        return {
            "final_signal": "🟡 NEUTRAL",
            "confidence": 40.0,
            "entry": None,
            "stop_loss": None,
            "target1": None,
            "target2": None,
            "rr_ratio": None,
            "scores": {
                "5m": 70,
                "15m": 60,
                "1h": 55,
                "pressure": 70 if pressure_strong_buy else 60,
                "options": 50,
            },
            "signal_reason": f"Next Candle Bias conflict: Bullish setup but Next Bias {next_bias_str}",
        }
    
    if is_bearish_aligned and "BUY" in next_bias_str and next_bias_conf >= 70:
        return {
            "final_signal": "🟡 NEUTRAL",
            "confidence": 40.0,
            "entry": None,
            "stop_loss": None,
            "target1": None,
            "target2": None,
            "rr_ratio": None,
            "scores": {
                "5m": 30,
                "15m": 40,
                "1h": 45,
                "pressure": 30 if pressure_strong_sell else 40,
                "options": 50,
            },
            "signal_reason": f"Next Candle Bias conflict: Bearish setup but Next Bias {next_bias_str}",
        }
    
    options_bias = options_data.get("options_bias", "NEUTRAL")
    options_conflict = False
    
    if is_bullish_aligned and "BEARISH" in str(options_bias):
        options_conflict = True
    elif is_bearish_aligned and "BULLISH" in str(options_bias):
        options_conflict = True
    
    if options_conflict:
        if not (has_bullish_structure or has_bearish_structure):
            confidence = min(confidence, 50.0)
    
    composite_score = 50
    
    if is_bullish_aligned:
        if tf_5m == "BULLISH" and tf_15m == "BULLISH":
            composite_score += 20
        elif tf_5m == "BULLISH":
            composite_score += 15
        
        if tf_1h == "BULLISH":
            composite_score += 5
        
        if pressure_strong_buy:
            composite_score += 15
        elif pressure_buy:
            composite_score += 10
        
        if vwap is not None and price_above_vwap:
            composite_score += 8
        
        if has_bullish_structure:
            composite_score += 10
        
        if rvol >= 1.5:
            composite_score += 5
        elif rvol < 1.2:
            composite_score -= 3
        
        if composite_score >= 80 and pressure_strong_buy and has_bullish_structure:
            final_signal = "🟢 STRONG BUY"
            confidence = min(100, confidence + 10)
        elif composite_score >= 70 and pressure_buy:
            final_signal = "🟢 BUY"
            confidence = min(95, confidence + 5)
        else:
            final_signal = "🟡 NEUTRAL"
            confidence = min(confidence, 60)
    
    elif is_bearish_aligned:
        if tf_5m == "BEARISH" and tf_15m == "BEARISH":
            composite_score -= 20
        elif tf_5m == "BEARISH":
            composite_score -= 15
        
        if tf_1h == "BEARISH":
            composite_score -= 5
        
        if pressure_strong_sell:
            composite_score -= 15
        elif pressure_sell:
            composite_score -= 10
        
        if vwap is not None and not price_above_vwap:
            composite_score -= 8
        
        if has_bearish_structure:
            composite_score -= 10
        
        if rvol >= 1.5:
            composite_score -= 5
        elif rvol < 1.2:
            composite_score += 3
        
        if composite_score <= 20 and pressure_strong_sell and has_bearish_structure:
            final_signal = "🔴 STRONG SELL"
            confidence = min(100, confidence + 10)
        elif composite_score <= 30 and pressure_sell:
            final_signal = "🔴 SELL"
            confidence = min(95, confidence + 5)
        else:
            final_signal = "🟡 NEUTRAL"
            confidence = min(confidence, 60)
    
    else:
        final_signal = "🟡 NEUTRAL"
        confidence = 35.0
    
    reason_parts = []
    
    if "BUY" in final_signal:
        reason_parts.append(f"5M {tf_5m}")
        reason_parts.append(f"15M {tf_15m}")
        reason_parts.append(f"1H {tf_1h}")
        if pressure_buy:
            reason_parts.append(f"Buy Pressure {data_5m.get('buying_pressure', 'N/A')}%")
        if vwap is not None and price_above_vwap:
            reason_parts.append("Above VWAP")
        if ema_trend == "BULLISH":
            reason_parts.append("EMA Bullish")
        if has_bullish_structure:
            reason_parts.append(f"{data_5m.get('structure_type', 'Structure')} Bullish")
        if rvol >= 1.5:
            reason_parts.append(f"RVOL {rvol}x")
    
    elif "SELL" in final_signal:
        reason_parts.append(f"5M {tf_5m}")
        reason_parts.append(f"15M {tf_15m}")
        reason_parts.append(f"1H {tf_1h}")
        if pressure_sell:
            reason_parts.append(f"Sell Pressure {data_5m.get('selling_pressure', 'N/A')}%")
        if vwap is not None and not price_above_vwap:
            reason_parts.append("Below VWAP")
        if ema_trend == "BEARISH":
            reason_parts.append("EMA Bearish")
        if has_bearish_structure:
            reason_parts.append(f"{data_5m.get('structure_type', 'Structure')} Bearish")
        if rvol >= 1.5:
            reason_parts.append(f"RVOL {rvol}x")
    
    else:
        reason_parts = ["Insufficient confirmation", f"TF: 5M {tf_5m} 15M {tf_15m} 1H {tf_1h}", f"Pressure: {pressure_trend}"]
    
    signal_reason = " + ".join(reason_parts)
    
    entry = sl = t1 = t2 = rr_ratio = None
    
    if "BUY" in final_signal or "SELL" in final_signal:
        entry = round(last_close, 2)
        atr_5m = data_5m.get("atr", 0)
        swing_high = data_5m.get("swing_high")
        swing_low = data_5m.get("swing_low")
        
        if "BUY" in final_signal:
            sl = round(swing_low - atr_5m * 0.5, 2) if swing_low else round(entry - atr_5m * 2, 2)
            t1 = round(entry + atr_5m * 1.5, 2)
            t2 = round(entry + atr_5m * 2.5, 2)
        else:
            sl = round(swing_high + atr_5m * 0.5, 2) if swing_high else round(entry + atr_5m * 2, 2)
            t1 = round(entry - atr_5m * 1.5, 2)
            t2 = round(entry - atr_5m * 2.5, 2)
        
        risk = abs(entry - sl)
        reward = abs(t1 - entry)
        rr_ratio = round(reward / risk, 2) if risk > 0 else 0.0
    
    return {
        "final_signal": final_signal,
        "confidence": confidence,
        "entry": entry,
        "stop_loss": sl,
        "target1": t1,
        "target2": t2,
        "rr_ratio": rr_ratio,
        "scores": {
            "5m": 70 if tf_5m == "BULLISH" else 30 if tf_5m == "BEARISH" else 50,
            "15m": 70 if tf_15m == "BULLISH" else 30 if tf_15m == "BEARISH" else 50,
            "1h": 70 if tf_1h == "BULLISH" else 30 if tf_1h == "BEARISH" else 50,
            "pressure": 70 if pressure_strong_buy else 60 if pressure_buy else 30 if pressure_strong_sell else 40 if pressure_sell else 50,
            "options": 75 if "BULLISH" in str(options_bias) else 25 if "BEARISH" in str(options_bias) else 50,
        },
        "signal_reason": signal_reason,
    }

# ════════════════════════════════════════════════════════════════════════════════
# NORMALIZED SIGNAL CLASSIFICATION (for filters)
# ════════════════════════════════════════════════════════════════════════════════
def normalize_signal(signal_str: str) -> str:
    """Return normalized signal class: BUY, SELL, or NEUTRAL."""
    if pd.isna(signal_str) or signal_str is None:
        return "NEUTRAL"
    sig = str(signal_str).upper()
    if "BUY" in sig and "SELL" not in sig:
        return "BUY"
    elif "SELL" in sig:
        return "SELL"
    else:
        return "NEUTRAL"

# ════════════════════════════════════════════════════════════════════════════════
# V17 NEW FEATURES: GOLDEN CROSS / DEATH CROSS DETECTION (DAILY)
# ════════════════════════════════════════════════════════════════════════════════
def detect_golden_death_cross(fyers, symbol: str) -> Dict[str, Any]:
    """Detect Golden Cross and Death Cross using DAILY closed candles only."""
    empty = {
        "golden_cross": False,
        "death_cross": False,
        "ema50": None,
        "ema200": None,
        "ema_trend": "NEUTRAL",
        "signal": "NONE",
        "signal_date": "N/A",
        "ltp": None,
        "reason": "DATA_UNAVAILABLE"
    }
    
    try:
        date_from = (datetime.today() - timedelta(days=100)).strftime("%Y-%m-%d")
        date_to = datetime.today().strftime("%Y-%m-%d")
        
        resp, err = _safe_history(fyers, {
            "symbol": symbol,
            "resolution": "1D",
            "date_format": "1",
            "range_from": date_from,
            "range_to": date_to,
            "cont_flag": "1",
        })
        
        if err or not resp:
            empty["reason"] = "FETCH_ERROR"
            return empty
        
        candles = resp.get("candles")
        if not candles or len(candles) < 50:
            empty["reason"] = "INSUFFICIENT_DATA"
            return empty
        
        df = pd.DataFrame(candles, columns=["Time", "Open", "High", "Low", "Close", "Volume"])
        df["Time"] = pd.to_datetime(df["Time"], unit="s", utc=True).dt.tz_convert("Asia/Kolkata")
        df[["Open", "High", "Low", "Close", "Volume"]] = df[["Open", "High", "Low", "Close", "Volume"]].apply(pd.to_numeric, errors="coerce")
        df = df.dropna(subset=["Close"]).sort_values("Time").reset_index(drop=True)
        
        if len(df) < 50:
            empty["reason"] = "INSUFFICIENT_DATA"
            return empty
        
        ema50 = calculate_ema(df["Close"], 50)
        ema200 = calculate_ema(df["Close"], 200)
        
        if len(ema50) < 2 or len(ema200) < 2:
            empty["reason"] = "EMA_CALCULATION_ERROR"
            return empty
        
        prev_ema50 = float(ema50.iloc[-2])
        prev_ema200 = float(ema200.iloc[-2])
        
        curr_ema50 = float(ema50.iloc[-1])
        curr_ema200 = float(ema200.iloc[-1])
        
        ltp = float(df["Close"].iloc[-1])
        signal_date = df["Time"].iloc[-1].strftime("%d-%b-%Y")
        
        if curr_ema50 > curr_ema200:
            ema_trend = "BULLISH"
        elif curr_ema50 < curr_ema200:
            ema_trend = "BEARISH"
        else:
            ema_trend = "NEUTRAL"
        
        golden_cross = (prev_ema50 <= prev_ema200) and (curr_ema50 > curr_ema200)
        death_cross = (prev_ema50 >= prev_ema200) and (curr_ema50 < curr_ema200)
        
        if golden_cross:
            signal = "🟢 GOLDEN CROSS"
        elif death_cross:
            signal = "🔴 DEATH CROSS"
        else:
            signal = "NONE"
        
        return {
            "golden_cross": golden_cross,
            "death_cross": death_cross,
            "ema50": round(curr_ema50, 2),
            "ema200": round(curr_ema200, 2),
            "ema_trend": ema_trend,
            "signal": signal,
            "signal_date": signal_date,
            "ltp": round(ltp, 2),
            "reason": "OK"
        }
    
    except Exception as e:
        empty["reason"] = f"ERROR: {str(e)}"
        return empty

# ════════════════════════════════════════════════════════════════════════════════
# MARKET STATISTICS
# ════════════════════════════════════════════════════════════════════════════════
def calculate_market_stats(results_df: pd.DataFrame) -> Dict[str, Any]:
    """Calculate market statistics from scan results."""
    if results_df is None or len(results_df) == 0:
        return {
            "total": 0,
            "buy": 0,
            "sell": 0,
            "neutral": 0,
            "strong_buy": 0,
            "strong_sell": 0,
            "avg_confidence": 0,
            "buy_pct": 0,
            "sell_pct": 0,
            "neutral_pct": 0,
        }
    
    try:
        total = len(results_df)
        
        normalized = results_df.get("AI SIGNAL", pd.Series([])).apply(normalize_signal)
        buy_count = len(normalized[normalized == "BUY"])
        sell_count = len(normalized[normalized == "SELL"])
        neutral_count = len(normalized[normalized == "NEUTRAL"])
        
        strong_buy = len(results_df[results_df.get("AI SIGNAL", pd.Series([])).astype(str).str.contains("STRONG BUY", na=False)])
        strong_sell = len(results_df[results_df.get("AI SIGNAL", pd.Series([])).astype(str).str.contains("STRONG SELL", na=False)])
        
        try:
            avg_conf = pd.to_numeric(results_df.get("AI CONFIDENCE %", pd.Series([])), errors='coerce').mean()
            avg_conf = 0 if pd.isna(avg_conf) else avg_conf
        except:
            avg_conf = 0
        
        buy_pct = (buy_count / total * 100) if total > 0 else 0
        sell_pct = (sell_count / total * 100) if total > 0 else 0
        neutral_pct = (neutral_count / total * 100) if total > 0 else 0
        
        return {
            "total": total,
            "buy": buy_count,
            "sell": sell_count,
            "neutral": neutral_count,
            "strong_buy": strong_buy,
            "strong_sell": strong_sell,
            "avg_confidence": round(avg_conf, 1),
            "buy_pct": round(buy_pct, 1),
            "sell_pct": round(sell_pct, 1),
            "neutral_pct": round(neutral_pct, 1),
        }
    
    except Exception as e:
        logger.error(f"Error calculating market stats: {e}")
        return {
            "total": 0,
            "buy": 0,
            "sell": 0,
            "neutral": 0,
            "strong_buy": 0,
            "strong_sell": 0,
            "avg_confidence": 0,
            "buy_pct": 0,
            "sell_pct": 0,
            "neutral_pct": 0,
        }

# ════════════════════════════════════════════════════════════════════════════════
# SAFE EXCEL EXPORT WITH FORMATTING
# ════════════════════════════════════════════════════════════════════════════════
def _safe_convert_df(df: pd.DataFrame) -> pd.DataFrame:
    """Convert all values to safe types for Excel export."""
    df_safe = df.copy()
    for col in df_safe.columns:
        try:
            df_safe[col] = df_safe[col].fillna("N/A")
            if df_safe[col].dtype in ['float64', 'float32']:
                df_safe[col] = df_safe[col].replace([np.inf, -np.inf], "N/A")
            df_safe[col] = df_safe[col].astype(str)
        except Exception:
            pass
    return df_safe

def _format_excel_output(df, scanner_type: str = "NSE") -> bytes:
    """Create professionally formatted Excel with colors and styling."""
    buf = io.BytesIO()
    
    try:
        df_export = _safe_convert_df(df)
        
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df_export.to_excel(writer, index=False, sheet_name="Signals")
            
            workbook = writer.book
            worksheet = writer.sheets["Signals"]
            
            if OPENPYXL_AVAILABLE:
                try:
                    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF", size=11)
                    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    
                    for col_num in range(1, len(df_export.columns) + 1):
                        cell = worksheet.cell(row=1, column=col_num)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_align
                    
                    for row_num in range(2, len(df_export) + 2):
                        for col_num in range(1, len(df_export.columns) + 1):
                            cell = worksheet.cell(row=row_num, column=col_num)
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            col_title = df_export.columns[col_num - 1]
                            
                            try:
                                cell_value = str(cell.value)
                                
                                if col_title in ["AI SIGNAL", "MOMENTUM SIGNAL"]:
                                    if "BUY" in cell_value and "SELL" not in cell_value:
                                        cell.fill = green_fill
                                    elif "SELL" in cell_value:
                                        cell.fill = red_fill
                                    else:
                                        cell.fill = yellow_fill
                                
                                elif col_title in ["AI CONFIDENCE %", "MOMENTUM SCORE"]:
                                    try:
                                        conf_val = float(cell_value)
                                        if conf_val >= 70:
                                            cell.fill = green_fill
                                        elif conf_val >= 50:
                                            cell.fill = yellow_fill
                                        else:
                                            cell.fill = red_fill
                                    except:
                                        pass
                            except Exception:
                                pass
                    
                    worksheet.freeze_panes = "A2"
                    
                    for col_num, col_title in enumerate(df_export.columns, 1):
                        max_length = len(str(col_title)) + 2
                        adjusted_width = min(max_length + 2, 60)
                        col_letter = get_column_letter(col_num)
                        worksheet.column_dimensions[col_letter].width = adjusted_width
                
                except Exception as format_err:
                    logging.warning(f"Excel formatting error: {format_err}")
        
        try:
            summary_data = {
                "Metric": ["Total Stocks", "BUY Signals", "SELL Signals", "NEUTRAL Signals", "Avg Confidence %"],
                "Value": [
                    str(len(df)),
                    str(len(df[df["AI SIGNAL"].apply(lambda x: normalize_signal(x)) == "BUY"])) if "AI SIGNAL" in df.columns else "N/A",
                    str(len(df[df["AI SIGNAL"].apply(lambda x: normalize_signal(x)) == "SELL"])) if "AI SIGNAL" in df.columns else "N/A",
                    str(len(df[df["AI SIGNAL"].apply(lambda x: normalize_signal(x)) == "NEUTRAL"])) if "AI SIGNAL" in df.columns else "N/A",
                    f"{pd.to_numeric(df.get('AI CONFIDENCE %', pd.Series([])), errors='coerce').mean():.1f}%" if "AI CONFIDENCE %" in df.columns else "N/A"
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, index=False, sheet_name="Summary")
        except Exception:
            pass
    
    except Exception as e:
        logging.error(f"Excel export error: {e}")
        try:
            buf = io.BytesIO()
            df_export = _safe_convert_df(df)
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                df_export.to_excel(writer, index=False, sheet_name="Signals")
        except Exception as final_err:
            logging.error(f"Final Excel fallback failed: {final_err}")
            buf = io.BytesIO()
            buf.write(to_csv_bytes(df))
    
    buf.seek(0)
    return buf.getvalue()

def to_csv_bytes(df) -> bytes:
    """Convert DataFrame to CSV bytes safely."""
    try:
        df_safe = _safe_convert_df(df)
        return df_safe.to_csv(index=False).encode("utf-8")
    except Exception as e:
        logging.error(f"CSV export error: {e}")
        return b"Error exporting to CSV"

def to_json_bytes(df) -> bytes:
    """Convert DataFrame to JSON bytes safely."""
    try:
        df_safe = _safe_convert_df(df)
        return df_safe.to_json(orient="records", indent=2, force_ascii=False).encode("utf-8")
    except Exception as e:
        logging.error(f"JSON export error: {e}")
        return b'{"error": "Could not export to JSON"}'

# ════════════════════════════════════════════════════════════════════════════════
# NSE SIGNAL SCANNER WORKER
# ════════════════════════════════════════════════════════════════════════════════
def _fetch_nse_signal(fyers, symbol: str):
    """Worker for NSE stocks scanning - NO options data."""
    stock_ticker = symbol.replace("NSE:", "").replace("-EQ", "") if isinstance(symbol, str) else str(symbol)
    
    if not isinstance(symbol, str) or not _VALID_EQ_SYMBOL_RE.match(symbol):
        return None, f"{symbol}: invalid format"
    
    try:
        analysis_5m = analyze_timeframe(fyers, symbol, "5")
        analysis_15m = analyze_timeframe(fyers, symbol, "15")
        analysis_1h = analyze_timeframe(fyers, symbol, "60")
        
        if all(a.get("status") != "OK" for a in [analysis_5m, analysis_15m, analysis_1h]):
            return None, None
        
        master = calculate_master_signal(symbol, analysis_5m, analysis_15m, analysis_1h)
        
        if analysis_5m.get("status") == "OK" and analysis_5m.get("data"):
            next_bias = calculate_next_candle_bias(analysis_5m.get("df"), analysis_5m)
        else:
            next_bias = {"bias": "NEUTRAL", "confidence": 0.0}
        
        ltp = None
        for analysis in [analysis_5m, analysis_15m, analysis_1h]:
            if analysis.get("status") == "OK" and analysis.get("data"):
                ltp = analysis["data"]["last_close"]
                break
        
        if ltp is None:
            return None, None
        
        data_5m = analysis_5m.get("data") if analysis_5m.get("status") == "OK" else {}
        data_15m = analysis_15m.get("data") if analysis_15m.get("status") == "OK" else {}
        data_1h = analysis_1h.get("data") if analysis_1h.get("status") == "OK" else {}
        
        return {
            "Symbol": stock_ticker,
            "LTP": round(float(ltp), 2),
            "Trend": data_5m.get("structure_trend", "N/A"),
            "5M Trend": data_5m.get("structure_trend", "N/A"),
            "15M Trend": data_15m.get("structure_trend", "N/A"),
            "1H Trend": data_1h.get("structure_trend", "N/A"),
            "5M Structure": data_5m.get("structure_type", "N/A"),
            "15M Structure": data_15m.get("structure_type", "N/A"),
            "1H Structure": data_1h.get("structure_type", "N/A"),
            "5M BOS": "✅" if data_5m.get("bullish_choch") else "❌" if data_5m.get("bearish_choch") else "−",
            "15M BOS": "✅" if data_15m.get("bullish_choch") else "❌" if data_15m.get("bearish_choch") else "−",
            "1H BOS": "✅" if data_1h.get("bullish_choch") else "❌" if data_1h.get("bearish_choch") else "−",
            "5M CHoCH": "✅" if data_5m.get("bullish_choch") else "❌" if data_5m.get("bearish_choch") else "−",
            "15M CHoCH": "✅" if data_15m.get("bullish_choch") else "❌" if data_15m.get("bearish_choch") else "−",
            "1H CHoCH": "✅" if data_1h.get("bullish_choch") else "❌" if data_1h.get("bearish_choch") else "−",
            "5M MSS": "✅" if data_5m.get("bullish_mss") else "❌" if data_5m.get("bearish_mss") else "−",
            "15M MSS": "✅" if data_15m.get("bullish_mss") else "❌" if data_15m.get("bearish_mss") else "−",
            "1H MSS": "✅" if data_1h.get("bullish_mss") else "❌" if data_1h.get("bearish_mss") else "−",
            "5M CISD": "✅" if data_5m.get("bullish_cisd") else "❌" if data_5m.get("bearish_cisd") else "−",
            "VWAP": round(data_5m.get("vwap", 0), 2) if data_5m.get("vwap") else "N/A",
            "RSI": round(data_5m.get("rsi", 50), 1),
            "MACD": "🟢" if data_5m.get("macd_bullish") else "🔴",
            "ATR": data_5m.get("atr", "N/A"),
            "Volume": data_5m.get("rvol", "N/A"),
            "RVOL": data_5m.get("rvol", 0),
            "🟢 BUY PRESSURE %": data_5m.get("buying_pressure", "N/A"),
            "🔴 SELL PRESSURE %": data_5m.get("selling_pressure", "N/A"),
            "PRESSURE SIGNAL": data_5m.get("pressure_trend", "N/A"),
            "NEXT CANDLE BIAS": next_bias.get("bias", "NEUTRAL"),
            "NEXT CANDLE CONFIDENCE %": next_bias.get("confidence", 0.0),
            "AI SIGNAL": master["final_signal"],
            "AI CONFIDENCE %": master["confidence"],
            "SIGNAL REASON": master["signal_reason"],
            "ENTRY": master["entry"],
            "STOP LOSS": master["stop_loss"],
            "TARGET 1": master["target1"],
            "TARGET 2": master["target2"],
            "RISK:REWARD": master["rr_ratio"],
        }, None
    
    except Exception as e:
        return None, f"{symbol}: error ({type(e).__name__})"

# ════════════════════════════════════════════════════════════════════════════════
# F&O SIGNAL SCANNER WORKER
# ════════════════════════════════════════════════════════════════════════════════
def _fetch_fo_signal(fyers, symbol: str):
    """Worker for F&O stocks scanning - WITH options data."""
    stock_ticker = symbol.replace("NSE:", "").replace("-EQ", "") if isinstance(symbol, str) else str(symbol)
    
    if not isinstance(symbol, str) or not _VALID_EQ_SYMBOL_RE.match(symbol):
        return None, f"{symbol}: invalid format"
    
    try:
        analysis_5m = analyze_timeframe(fyers, symbol, "5")
        analysis_15m = analyze_timeframe(fyers, symbol, "15")
        analysis_1h = analyze_timeframe(fyers, symbol, "60")
        
        if all(a.get("status") != "OK" for a in [analysis_5m, analysis_15m, analysis_1h]):
            return None, None
        
        options_data = fetch_options_chain_data(fyers, symbol)
        
        master = calculate_master_signal(symbol, analysis_5m, analysis_15m, analysis_1h, options_data)
        
        if analysis_5m.get("status") == "OK" and analysis_5m.get("data"):
            next_bias = calculate_next_candle_bias(analysis_5m.get("df"), analysis_5m)
        else:
            next_bias = {"bias": "NEUTRAL", "confidence": 0.0}
        
        ltp = None
        for analysis in [analysis_5m, analysis_15m, analysis_1h]:
            if analysis.get("status") == "OK" and analysis.get("data"):
                ltp = analysis["data"]["last_close"]
                break
        
        if ltp is None:
            return None, None
        
        data_5m = analysis_5m.get("data") if analysis_5m.get("status") == "OK" else {}
        data_15m = analysis_15m.get("data") if analysis_15m.get("status") == "OK" else {}
        data_1h = analysis_1h.get("data") if analysis_1h.get("status") == "OK" else {}
        
        return {
            "Symbol": stock_ticker,
            "LTP": round(float(ltp), 2),
            "Trend": data_5m.get("structure_trend", "N/A"),
            "5M Trend": data_5m.get("structure_trend", "N/A"),
            "15M Trend": data_15m.get("structure_trend", "N/A"),
            "1H Trend": data_1h.get("structure_trend", "N/A"),
            "5M Structure": data_5m.get("structure_type", "N/A"),
            "15M Structure": data_15m.get("structure_type", "N/A"),
            "1H Structure": data_1h.get("structure_type", "N/A"),
            "5M BOS": "✅" if data_5m.get("bullish_choch") else "❌" if data_5m.get("bearish_choch") else "−",
            "15M BOS": "✅" if data_15m.get("bullish_choch") else "❌" if data_15m.get("bearish_choch") else "−",
            "1H BOS": "✅" if data_1h.get("bullish_choch") else "❌" if data_1h.get("bearish_choch") else "−",
            "5M CHoCH": "✅" if data_5m.get("bullish_choch") else "❌" if data_5m.get("bearish_choch") else "−",
            "15M CHoCH": "✅" if data_15m.get("bullish_choch") else "❌" if data_15m.get("bearish_choch") else "−",
            "1H CHoCH": "✅" if data_1h.get("bullish_choch") else "❌" if data_1h.get("bearish_choch") else "−",
            "5M MSS": "✅" if data_5m.get("bullish_mss") else "❌" if data_5m.get("bearish_mss") else "−",
            "15M MSS": "✅" if data_15m.get("bullish_mss") else "❌" if data_15m.get("bearish_mss") else "−",
            "1H MSS": "✅" if data_1h.get("bullish_mss") else "❌" if data_1h.get("bearish_mss") else "−",
            "5M CISD": "✅" if data_5m.get("bullish_cisd") else "❌" if data_5m.get("bearish_cisd") else "−",
            "VWAP": round(data_5m.get("vwap", 0), 2) if data_5m.get("vwap") else "N/A",
            "RSI": round(data_5m.get("rsi", 50), 1),
            "MACD": "🟢" if data_5m.get("macd_bullish") else "🔴",
            "ATR": data_5m.get("atr", "N/A"),
            "Volume": data_5m.get("rvol", "N/A"),
            "RVOL": data_5m.get("rvol", 0),
            "🟢 BUY PRESSURE %": data_5m.get("buying_pressure", "N/A"),
            "🔴 SELL PRESSURE %": data_5m.get("selling_pressure", "N/A"),
            "PRESSURE SIGNAL": data_5m.get("pressure_trend", "N/A"),
            "NEXT CANDLE BIAS": next_bias.get("bias", "NEUTRAL"),
            "NEXT CANDLE CONFIDENCE %": next_bias.get("confidence", 0.0),
            "ATM STRIKE": round(options_data.get("atm_strike", 0), 2) if options_data.get("atm_strike") else "N/A",
            "PCR": options_data.get("pcr", "N/A"),
            "CE OI": options_data.get("ce_oi", "N/A"),
            "PE OI": options_data.get("pe_oi", "N/A"),
            "CE OI CHANGE": options_data.get("ce_oi_change", "N/A"),
            "PE OI CHANGE": options_data.get("pe_oi_change", "N/A"),
            "OPTIONS BIAS": options_data.get("options_bias", "N/A"),
            "AI SIGNAL": master["final_signal"],
            "AI CONFIDENCE %": master["confidence"],
            "SIGNAL REASON": master["signal_reason"],
            "ENTRY": master["entry"],
            "STOP LOSS": master["stop_loss"],
            "TARGET 1": master["target1"],
            "TARGET 2": master["target2"],
            "RISK:REWARD": master["rr_ratio"],
        }, None
    
    except Exception as e:
        return None, f"{symbol}: error ({type(e).__name__})"

# ════════════════════════════════════════════════════════════════════════════════
# MOMENTUM SCANNER WORKER (NEW)
# ════════════════════════════════════════════════════════════════════════════════
def detect_block_order_activity(df: pd.DataFrame) -> Dict[str, Any]:
    """Estimate institutional/block-order activity from OHLCV only.

    This is a probability score, NOT a true exchange order-book/block-trade feed.
    It uses abnormal volume, candle body/wicks, price acceptance/rejection and
    repeated high-volume levels to flag possible large-player activity.
    """
    out = {
        "block_score": 0.0,
        "block_signal": "NONE",
        "block_side": "NONE",
        "block_level": None,
        "block_rvol": 0.0,
        "block_reason": "Insufficient data",
    }
    if df is None or len(df) < 25:
        return out
    try:
        d = df.reset_index(drop=True).copy()
        cols = ["Open", "High", "Low", "Close", "Volume"]
        if any(c not in d.columns for c in cols):
            out["block_reason"] = "Missing OHLCV columns"
            return out

        last = d.iloc[-1]
        o,h,l,c,v = [float(last[x]) for x in cols]
        prev = float(d["Close"].iloc[-2])
        if min(o,h,l,c,prev) <= 0:
            return out

        rng = max(h-l, 1e-9)
        body = abs(c-o)
        body_pct = body/rng*100
        upper_wick = h-max(o,c)
        lower_wick = min(o,c)-l
        close_pos = (c-l)/rng

        base = float(d["Volume"].iloc[-21:-1].mean())
        rvol = v/base if base > 0 else 0.0
        recent = d.tail(20).copy()
        avg_ranges = (recent["High"]-recent["Low"]).astype(float).replace(0,np.nan).mean()
        range_ratio = rng/float(avg_ranges) if avg_ranges and pd.notna(avg_ranges) else 0.0

        # High-volume candle close behavior.
        buy_points = sell_points = 0.0
        reasons=[]
        if rvol >= 2.0:
            buy_points += 20; sell_points += 20; reasons.append(f"RVOL {rvol:.1f}x")
        elif rvol >= 1.5:
            buy_points += 12; sell_points += 12; reasons.append(f"RVOL {rvol:.1f}x")
        if range_ratio >= 1.5:
            buy_points += 10; sell_points += 10; reasons.append("large range")

        # Acceptance near high = possible buy absorption/accumulation.
        if close_pos >= 0.75 and c > o:
            buy_points += 25
        if close_pos <= 0.25 and c < o:
            sell_points += 25

        # Rejection/absorption: large wick with high volume.
        if lower_wick/rng >= 0.45 and rvol >= 1.5:
            buy_points += 20; reasons.append("lower-wick absorption")
        if upper_wick/rng >= 0.45 and rvol >= 1.5:
            sell_points += 20; reasons.append("upper-wick rejection")

        # Repeated volume concentration around current price.
        lo, hi = float(recent["Low"].min()), float(recent["High"].max())
        if hi > lo:
            bins = np.linspace(lo, hi, 11)
            mids=(bins[:-1]+bins[1:])/2
            idx=np.clip(np.digitize(recent["Close"].astype(float), bins)-1,0,9)
            vol_by_bin=recent.groupby(idx)["Volume"].sum()
            if len(vol_by_bin):
                peak_bin=int(vol_by_bin.idxmax())
                peak_level=float(mids[peak_bin])
                peak_share=float(vol_by_bin.max()/max(float(recent["Volume"].sum()),1.0))
                if peak_share >= 0.20 and abs(c-peak_level)/c <= 0.02:
                    buy_points += 15 if c >= peak_level else 5
                    sell_points += 15 if c <= peak_level else 5
                    reasons.append("volume concentration")
                else:
                    peak_level=c
        else:
            peak_level=c

        # Directional confirmation from recent closes.
        ret3 = (c/float(d["Close"].iloc[-4])-1)*100 if len(d)>=4 else 0
        if ret3 > 0.4: buy_points += 10
        elif ret3 < -0.4: sell_points += 10

        buy_points=min(100,buy_points); sell_points=min(100,sell_points)
        side = "BUY" if buy_points > sell_points else "SELL" if sell_points > buy_points else "NONE"
        score=max(buy_points,sell_points)
        if score >= 75:
            signal="🔥 VERY HIGH"
        elif score >= 60:
            signal="🟢 HIGH BUY" if side=="BUY" else "🔴 HIGH SELL" if side=="SELL" else "🟡 HIGH"
        elif score >= 45:
            signal="🟡 POSSIBLE"
        else:
            signal="NONE"

        direction_reason = "BUY-side" if side=="BUY" else "SELL-side" if side=="SELL" else "mixed"
        out.update({
            "block_score": round(score,1),
            "block_signal": signal,
            "block_side": side,
            "block_level": round(float(peak_level),2) if peak_level is not None else None,
            "block_rvol": round(rvol,2),
            "block_reason": f"{direction_reason} | " + (" + ".join(reasons) if reasons else "normal volume/price behavior"),
        })
        return out
    except Exception as e:
        out["block_reason"] = f"ERROR: {str(e)[:100]}"
        return out


def detect_live_sudden_move(df: pd.DataFrame) -> Dict[str, Any]:
    """Detect CURRENT sudden 5M BUY/SELL movement. No consolidation required."""
    result = {
        "signal": "NO MOVE", "direction": "NONE", "score": 0.0,
        "reason": "", "move_pct": 0.0, "body_pct": 0.0,
        "body_atr": 0.0, "rvol": 0.0, "structure": "NONE",
        "hh_hl": False, "lh_ll": False, "price_acceleration": 0.0,
        "volume_spike": False,
    }
    if df is None or len(df) < 12:
        result["reason"] = "Insufficient recent 5M candles"
        return result
    try:
        d = df.reset_index(drop=True).copy()
        last = d.iloc[-1]
        o, h, l, c, v = map(float, [last["Open"], last["High"], last["Low"], last["Close"], last["Volume"]])
        prev = float(d["Close"].iloc[-2])
        if min(o, c, prev) <= 0:
            result["reason"] = "Invalid price data"
            return result

        move_pct = ((c - prev) / prev) * 100.0
        candle_range = max(h - l, 1e-9)
        body = abs(c - o)
        body_pct = body / candle_range * 100.0
        atr_s = calculate_atr(d, 14)
        atr = float(atr_s.iloc[-1]) if len(atr_s) and pd.notna(atr_s.iloc[-1]) else max(c * 0.003, 0.01)
        body_atr = body / atr if atr > 0 else 0.0

        vol_base = float(d["Volume"].iloc[-11:-1].mean()) if len(d) >= 11 else float(d["Volume"].iloc[:-1].mean())
        rvol = v / vol_base if vol_base > 0 else 0.0

        # Recent acceleration: current candle move versus average of last few candle moves.
        moves = []
        for i in range(max(1, len(d)-6), len(d)-1):
            pc = float(d["Close"].iloc[i-1]); cc = float(d["Close"].iloc[i])
            if pc > 0:
                moves.append(abs((cc-pc)/pc)*100.0)
        avg_move = float(np.mean(moves)) if moves else 0.0
        acceleration = abs(move_pct) / avg_move if avg_move > 0 else 0.0

        # Recent market structure, not old daily/history setup.
        ph, pl = _confirmed_pivots(d.tail(20), left=1, right=1)
        hh_hl = lh_ll = False
        if len(ph) >= 2 and len(pl) >= 2:
            hh_hl = ph[-1][1] > ph[-2][1] and pl[-1][1] > pl[-2][1]
            lh_ll = ph[-1][1] < ph[-2][1] and pl[-1][1] < pl[-2][1]
        structure = "HH/HL" if hh_hl else "LH/LL" if lh_ll else "NONE"

        bullish = c > o
        bearish = c < o
        volume_spike = rvol >= LIVE_MOVE_MIN_RVOL
        strong_volume = rvol >= LIVE_MOVE_STRONG_RVOL
        strong_body = body_pct >= LIVE_MOVE_MIN_BODY_PCT
        very_strong_body = body_pct >= 65.0
        acceleration_ok = acceleration >= 1.30

        buy = 0.0
        sell = 0.0
        if bullish: buy += 20
        if move_pct >= LIVE_MOVE_MIN_PCT: buy += 20
        if move_pct >= LIVE_MOVE_BIG_PCT: buy += 10
        if bullish and strong_body: buy += 10
        if bullish and very_strong_body: buy += 5
        if volume_spike: buy += 10
        if strong_volume: buy += 5
        if move_pct > 0 and acceleration_ok: buy += 10
        if hh_hl: buy += 10

        if bearish: sell += 20
        if move_pct <= -LIVE_MOVE_MIN_PCT: sell += 20
        if move_pct <= -LIVE_MOVE_BIG_PCT: sell += 10
        if bearish and strong_body: sell += 10
        if bearish and very_strong_body: sell += 5
        if volume_spike: sell += 10
        if strong_volume: sell += 5
        if move_pct < 0 and acceleration_ok: sell += 10
        if lh_ll: sell += 10

        buy = min(100.0, buy); sell = min(100.0, sell)
        score = max(buy, sell)

        if buy >= LIVE_MOVE_MIN_SCORE and buy > sell:
            direction = "BUY"
            signal = "🔥 BIG BUY" if buy >= LIVE_MOVE_STRONG_SCORE else "🟢 BUY"
            reasons = [f"Move +{move_pct:.2f}%", f"RVOL {rvol:.2f}x", f"Body {body_pct:.0f}%"]
            if hh_hl: reasons.append("HH/HL")
            if acceleration_ok: reasons.append("Acceleration")
            reason = " + ".join(reasons)
            score = buy
        elif sell >= LIVE_MOVE_MIN_SCORE and sell > buy:
            direction = "SELL"
            signal = "🔥 BIG SELL" if sell >= LIVE_MOVE_STRONG_SCORE else "🔴 SELL"
            reasons = [f"Move {move_pct:.2f}%", f"RVOL {rvol:.2f}x", f"Body {body_pct:.0f}%"]
            if lh_ll: reasons.append("LH/LL")
            if acceleration_ok: reasons.append("Acceleration")
            reason = " + ".join(reasons)
            score = sell
        else:
            direction = "NONE"
            signal = "NO MOVE"
            reason = f"Move {move_pct:.2f}% | RVOL {rvol:.2f}x | Body {body_pct:.0f}% | Structure {structure}"

        result.update({
            "signal": signal, "direction": direction, "score": round(score, 1),
            "reason": reason, "move_pct": round(move_pct, 3),
            "body_pct": round(body_pct, 1), "body_atr": round(body_atr, 2),
            "rvol": round(rvol, 2), "structure": structure,
            "hh_hl": hh_hl, "lh_ll": lh_ll,
            "price_acceleration": round(acceleration, 2),
            "volume_spike": volume_spike,
        })
        return result
    except Exception as e:
        result["reason"] = f"ERROR: {str(e)[:120]}"
        return result


def _fetch_momentum_signal(fyers, symbol: str, is_fo: bool = False):
    """LIVE SUDDEN MOVEMENT worker: recent 5M candles only."""
    stock_ticker = symbol.replace("NSE:", "").replace("-EQ", "") if isinstance(symbol, str) else str(symbol)
    if not isinstance(symbol, str) or not _VALID_EQ_SYMBOL_RE.match(symbol):
        return None, f"{symbol}: invalid format"
    try:
        # Only recent 5M data is fetched. No 15M/1H confirmation and no consolidation scan.
        df5 = _fetch_timeframe_data(fyers, symbol, "5", lookback_days=LIVE_MOVE_LOOKBACK_DAYS)
        if df5 is None or len(df5) < 12:
            return None, f"{symbol}: insufficient recent 5M data"

        move = detect_live_sudden_move(df5)
        block = detect_block_order_activity(df5)
        ltp = float(df5["Close"].iloc[-1])
        result = {
            "Symbol": stock_ticker,
            "LTP": round(ltp, 2),
            "SIGNAL": move["signal"],
            "DIRECTION": move["direction"],
            "MOVE %": move["move_pct"],
            "BODY %": move["body_pct"],
            "BODY / ATR": move["body_atr"],
            "RVOL": move["rvol"],
            "STRUCTURE": move["structure"],
            "HH/HL": "✅" if move["hh_hl"] else "−",
            "LH/LL": "✅" if move["lh_ll"] else "−",
            "ACCELERATION": move["price_acceleration"],
            "VOLUME SPIKE": "🔥" if move["volume_spike"] else "−",
            "SCORE": move["score"],
            "BLOCK ORDER SCORE": block["block_score"],
            "BLOCK ACTIVITY": block["block_signal"],
            "BLOCK SIDE": block["block_side"],
            "BLOCK LEVEL": block["block_level"],
            "BLOCK RVOL": block["block_rvol"],
            "BLOCK REASON": block["block_reason"],
            "REASON": move["reason"],
        }
        if is_fo and move["direction"] in ("BUY", "SELL"):
            try:
                options_data = fetch_options_chain_data(fyers, symbol)
                result["PCR"] = options_data.get("pcr", "N/A")
                result["OPTIONS BIAS"] = options_data.get("options_bias", "N/A")
            except Exception:
                result["PCR"] = "N/A"
                result["OPTIONS BIAS"] = "N/A"
        return result, None
    except Exception as e:
        logger.exception("LIVE MOMENTUM worker failed for %s", symbol)
        return None, f"{symbol}: error ({type(e).__name__}: {str(e)[:120]})"

# ════════════════════════════════════════════════════════════════════════════════
# THREADED SCAN FUNCTIONS
# ════════════════════════════════════════════════════════════════════════════════
def run_nse_scan(fyers, symbols):
    """Threaded scan for NSE stocks."""
    symbols = _validate_symbols(symbols)
    results, errors = [], []
    stats = ScanStats(total=len(symbols))
    progress = st.progress(0.0, text=f"Scanning NSE Stocks 0 / {len(symbols)}")
    done = 0
    
    for i in range(0, len(symbols), BATCH_SIZE):
        batch = symbols[i:i + BATCH_SIZE]
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {executor.submit(_fetch_nse_signal, fyers, s): s for s in batch}
            for future in as_completed(futures):
                try:
                    res, err = future.result()
                except Exception as e:
                    res, err = None, f"{futures[future]}: worker error"
                
                if res:
                    results.append(res)
                if err:
                    errors.append(err)
                stats.record(has_result=bool(res), has_error=bool(err))
                done += 1
                progress.progress(done / max(len(symbols), 1), text=f"Scanning NSE {done} / {len(symbols)}")
        
        if i + BATCH_SIZE < len(symbols):
            time.sleep(BATCH_PAUSE_SECONDS)
    
    progress.empty()
    gc.collect()
    return results, errors, stats

def run_fo_scan(fyers, symbols):
    """Threaded scan for F&O stocks."""
    symbols = _validate_symbols(symbols)
    results, errors = [], []
    stats = ScanStats(total=len(symbols))
    progress = st.progress(0.0, text=f"Scanning F&O Stocks 0 / {len(symbols)}")
    done = 0
    
    for i in range(0, len(symbols), BATCH_SIZE):
        batch = symbols[i:i + BATCH_SIZE]
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {executor.submit(_fetch_fo_signal, fyers, s): s for s in batch}
            for future in as_completed(futures):
                try:
                    res, err = future.result()
                except Exception as e:
                    res, err = None, f"{futures[future]}: worker error"
                
                if res:
                    results.append(res)
                if err:
                    errors.append(err)
                stats.record(has_result=bool(res), has_error=bool(err))
                done += 1
                progress.progress(done / max(len(symbols), 1), text=f"Scanning F&O {done} / {len(symbols)}")
        
        if i + BATCH_SIZE < len(symbols):
            time.sleep(BATCH_PAUSE_SECONDS)
    
    progress.empty()
    gc.collect()
    return results, errors, stats

def run_momentum_scan(fyers, symbols, is_fo: bool = False):
    """Threaded LIVE sudden movement scan. Returns BUY/SELL only."""
    symbols = _validate_symbols(symbols)
    results, errors = [], []
    stats = ScanStats(total=len(symbols))
    progress = st.progress(0.0, text=f"Scanning Live Movement 0 / {len(symbols)}")
    done = 0

    for i in range(0, len(symbols), BATCH_SIZE):
        batch = symbols[i:i + BATCH_SIZE]
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {executor.submit(_fetch_momentum_signal, fyers, s, is_fo): s for s in batch}
            for future in as_completed(futures):
                symbol = futures[future]
                try:
                    res, err = future.result()
                except Exception as e:
                    res, err = None, f"{symbol}: worker error: {str(e)[:100]}"
                if res and res.get("DIRECTION") in ("BUY", "SELL"):
                    results.append(res)
                if err:
                    errors.append(err)
                stats.record(has_result=bool(res), has_error=bool(err))
                done += 1
                progress.progress(done / max(len(symbols), 1), text=f"Live Movement {done} / {len(symbols)}")
        if i + BATCH_SIZE < len(symbols):
            time.sleep(BATCH_PAUSE_SECONDS)

    progress.empty()
    gc.collect()
    results.sort(key=lambda x: (float(x.get("SCORE", 0)), abs(float(x.get("MOVE %", 0))), float(x.get("RVOL", 0))), reverse=True)
    return results, errors, stats


# ════════════════════════════════════════════════════════════════════════════════
# PIN RULES — ADDITIONAL LIQUIDITY / REVERSAL / BIG-MOVE ANALYSIS
# Existing scanner logic is intentionally untouched. This tab runs only when used.
# ════════════════════════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════════════════════════
# AMD SUPPORT HELPER — COMPLETED CANDLES
# Added only for AMD/PIN modules; existing scanner logic is unchanged.
def _completed_candles(df: pd.DataFrame, resolution_minutes: int = 5) -> pd.DataFrame:
    """Return completed OHLCV candles only. Safe when Time is missing/invalid."""
    if df is None or len(df) == 0:
        return df
    d = df.copy()
    if "Time" not in d.columns:
        return d.reset_index(drop=True)
    try:
        t = pd.to_datetime(d["Time"], errors="coerce", utc=True)
        now_ist = _now_ist()
        cutoff = pd.Timestamp(now_ist)
        if cutoff.tzinfo is None:
            cutoff = cutoff.tz_localize("Asia/Kolkata")
        cutoff = cutoff.tz_convert("UTC")
        mask = t.notna() & ((t + pd.Timedelta(minutes=resolution_minutes)) <= cutoff)
        out = d.loc[mask].copy()
        out["Time"] = t.loc[mask].values
        return out.reset_index(drop=True)
    except Exception:
        return d.reset_index(drop=True)


# AMD — ACCUMULATION / MANIPULATION / DISTRIBUTION ENGINE
# Rule-based inference from completed OHLCV candles.
# IMPORTANT: this is a market-structure/volume heuristic, not proof of intent.
# ════════════════════════════════════════════════════════════════════════════════
AMD_LOOKBACK = 24
AMD_MIN_BARS = 30
AMD_RANGE_MAX_PCT = 3.0
AMD_RVOL_HIGH = 1.50
AMD_RVOL_EXTREME = 2.00
AMD_SWEEP_TOL_PCT = 0.15
AMD_MIN_SIGNAL_SCORE = 65.0


def calculate_amd_signal(df_5m: pd.DataFrame, df_15m: Optional[pd.DataFrame] = None) -> Dict[str, Any]:
    """Detect Accumulation / Manipulation / Distribution from completed candles.

    Accumulation: tight range + relatively strong/steady volume + bullish acceptance.
    Manipulation: sweep of a recent high/low followed by rejection back inside range.
    Distribution: high-volume weakness + bearish acceptance / range breakdown.

    The result is an inference from OHLCV, not a claim about actual institutional intent.
    """
    out = {
        "AMD PHASE": "NEUTRAL",
        "AMD SIGNAL": "WAIT",
        "AMD SCORE": 0.0,
        "AMD BUY SCORE": 0.0,
        "AMD SELL SCORE": 0.0,
        "AMD CONFIDENCE %": 0.0,
        "AMD RANGE HIGH": None,
        "AMD RANGE LOW": None,
        "AMD SWEEP": "NONE",
        "AMD RVOL": 0.0,
        "AMD REASON": "Insufficient completed 5M data",
    }
    if df_5m is None or len(df_5m) < AMD_MIN_BARS:
        return out

    try:
        d = _completed_candles(df_5m, 5)
        if d is None or len(d) < AMD_MIN_BARS:
            out["AMD REASON"] = "Waiting for completed 5M candles"
            return out

        d = d.copy()
        for col in ["Open", "High", "Low", "Close", "Volume"]:
            d[col] = pd.to_numeric(d[col], errors="coerce")
        d = d.dropna(subset=["Open", "High", "Low", "Close", "Volume"]).reset_index(drop=True)
        if len(d) < AMD_MIN_BARS:
            return out

        last = d.iloc[-1]
        o, h, l, c, v = [float(last[x]) for x in ["Open", "High", "Low", "Close", "Volume"]]
        if min(o, h, l, c) <= 0:
            out["AMD REASON"] = "Invalid price data"
            return out

        # Recent range excludes the current candle so the current candle can sweep it.
        w = d.iloc[-AMD_LOOKBACK-1:-1]
        range_high = float(w["High"].max())
        range_low = float(w["Low"].min())
        mid = (range_high + range_low) / 2.0
        range_pct = ((range_high - range_low) / mid * 100.0) if mid > 0 else 999.0

        atr_s = calculate_atr(d, 14)
        atr = float(atr_s.iloc[-1]) if len(atr_s) and pd.notna(atr_s.iloc[-1]) else max(c * 0.005, 0.01)
        base_vol = float(d["Volume"].iloc[-21:-1].mean()) if len(d) >= 22 else float(d["Volume"].iloc[:-1].mean())
        rvol = v / base_vol if base_vol > 0 else 0.0

        rng = max(h - l, 1e-9)
        body = abs(c - o)
        body_pct = body / rng * 100.0
        close_pos = (c - l) / rng
        upper_wick = h - max(o, c)
        lower_wick = min(o, c) - l

        # Sweep = price temporarily breaks a prior range extreme but closes back inside.
        high_sweep = h > range_high * (1.0 + AMD_SWEEP_TOL_PCT / 100.0) and c < range_high
        low_sweep = l < range_low * (1.0 - AMD_SWEEP_TOL_PCT / 100.0) and c > range_low
        sweep = "HIGH SWEEP" if high_sweep else "LOW SWEEP" if low_sweep else "NONE"

        # Recent directional acceptance.
        recent = d.tail(5)
        recent_change = (float(recent["Close"].iloc[-1]) - float(recent["Close"].iloc[0])) / float(recent["Close"].iloc[0]) * 100.0
        above_mid = c > mid
        below_mid = c < mid

        # Volume behavior inside the prior range.
        vol_recent = float(d["Volume"].iloc[-6:-1].mean()) if len(d) >= 7 else base_vol
        volume_stable = (vol_recent / base_vol) if base_vol > 0 else 1.0

        buy = 0.0
        sell = 0.0
        manipulation = 0.0
        buy_reasons = []
        sell_reasons = []
        manip_reasons = []

        # ACCUMULATION evidence
        if range_pct <= AMD_RANGE_MAX_PCT:
            buy += 25; buy_reasons.append(f"tight range {range_pct:.2f}%")
        if rvol >= 1.10 and close_pos >= 0.55:
            buy += 15; buy_reasons.append(f"volume acceptance {rvol:.2f}x")
        if recent_change >= 0.20:
            buy += 15; buy_reasons.append(f"recent +{recent_change:.2f}%")
        if lower_wick > upper_wick * 1.20:
            buy += 10; buy_reasons.append("lower-wick rejection")
        if c >= mid:
            buy += 10
        if volume_stable >= 1.0:
            buy += 5

        # DISTRIBUTION evidence
        if range_pct <= AMD_RANGE_MAX_PCT:
            sell += 15
        if rvol >= 1.10 and close_pos <= 0.45:
            sell += 15; sell_reasons.append(f"selling volume {rvol:.2f}x")
        if recent_change <= -0.20:
            sell += 15; sell_reasons.append(f"recent {recent_change:.2f}%")
        if upper_wick > lower_wick * 1.20:
            sell += 10; sell_reasons.append("upper-wick rejection")
        if c <= mid:
            sell += 10
        if c < range_low:
            sell += 20; sell_reasons.append("range breakdown")

        # MANIPULATION / liquidity sweep evidence.
        if high_sweep:
            manipulation += 55
            sell += 20
            manip_reasons.append("high liquidity sweep + close back inside")
            if rvol >= AMD_RVOL_HIGH:
                manipulation += 15; manip_reasons.append(f"high RVOL {rvol:.2f}x")
            if upper_wick >= max(body * 1.20, atr * 0.25):
                manipulation += 15; manip_reasons.append("upper rejection")
        elif low_sweep:
            manipulation += 55
            buy += 20
            manip_reasons.append("low liquidity sweep + close back inside")
            if rvol >= AMD_RVOL_HIGH:
                manipulation += 15; manip_reasons.append(f"high RVOL {rvol:.2f}x")
            if lower_wick >= max(body * 1.20, atr * 0.25):
                manipulation += 15; manip_reasons.append("lower rejection")

        # 15M confirmation, when available.
        tf15 = "NEUTRAL"
        if df_15m is not None and len(df_15m) >= 10:
            try:
                p = df_15m.copy()
                for col in ["Open", "Close"]:
                    p[col] = pd.to_numeric(p[col], errors="coerce")
                p = p.dropna(subset=["Open", "Close"])
                if len(p) >= 5:
                    pchg = (float(p["Close"].iloc[-1]) - float(p["Close"].iloc[-4])) / float(p["Close"].iloc[-4]) * 100.0
                    tf15 = "BULLISH" if pchg > 0.25 else "BEARISH" if pchg < -0.25 else "NEUTRAL"
                    if tf15 == "BULLISH": buy += 5
                    elif tf15 == "BEARISH": sell += 5
            except Exception:
                tf15 = "NEUTRAL"

        buy = min(100.0, buy)
        sell = min(100.0, sell)
        manipulation = min(100.0, manipulation)

        # Priority: a confirmed sweep is classified as manipulation; otherwise compare A vs D.
        if manipulation >= 65 and sweep != "NONE":
            phase = "MANIPULATION"
            if sweep == "LOW SWEEP" and buy >= sell:
                signal = "🟢 AMD BUY AFTER SWEEP"
            elif sweep == "HIGH SWEEP" and sell >= buy:
                signal = "🔴 AMD SELL AFTER SWEEP"
            else:
                signal = "🟠 AMD SWEEP — WAIT"
            score = manipulation
            reason = " + ".join(manip_reasons) or sweep
        elif buy >= 65 and buy > sell + 8:
            phase = "ACCUMULATION"
            signal = "🟢 AMD ACCUMULATION BUY WATCH"
            score = buy
            reason = " + ".join(buy_reasons) or "Bullish accumulation evidence"
        elif sell >= 65 and sell > buy + 8:
            phase = "DISTRIBUTION"
            signal = "🔴 AMD DISTRIBUTION SELL WATCH"
            score = sell
            reason = " + ".join(sell_reasons) or "Bearish distribution evidence"
        else:
            phase = "TRANSITION" if max(buy, sell, manipulation) >= 50 else "NEUTRAL"
            signal = "🟡 AMD TRANSITION — WAIT" if phase == "TRANSITION" else "⚪ AMD WAIT"
            score = max(buy, sell, manipulation)
            reason = "Mixed AMD evidence"

        confidence = min(100.0, round(max(0.0, score * 0.85), 1))
        out.update({
            "AMD PHASE": phase,
            "AMD SIGNAL": signal,
            "AMD SCORE": round(score, 1),
            "AMD BUY SCORE": round(buy, 1),
            "AMD SELL SCORE": round(sell, 1),
            "AMD CONFIDENCE %": confidence,
            "AMD RANGE HIGH": round(range_high, 2),
            "AMD RANGE LOW": round(range_low, 2),
            "AMD SWEEP": sweep,
            "AMD RVOL": round(rvol, 2),
            "AMD REASON": f"{reason} | 15M {tf15}",
        })
        return out
    except Exception as e:
        out["AMD REASON"] = f"AMD error: {type(e).__name__}"
        return out



PIN_MIN_CONFIDENCE = 70
PIN_STRONG_CONFIDENCE = 82
PIN_PIVOT_LEN = 5
PIN_EQUAL_ATR_TOL = 0.15
PIN_BIGMOVE_MIN_SCORE = 70
PIN_MAX_SCAN = 100


def calculate_pin_rules(df_5m: pd.DataFrame, data_5m: Dict[str, Any], data_15m: Dict[str, Any], data_1h: Dict[str, Any]) -> Dict[str, Any]:
    """Rule-based implementation of the supplied Pine 'AI PRO v3' ideas.
    This is NOT machine-learning AI and it does not access the exchange order book.
    """
    out = {
        "PIN SIGNAL": "WAIT", "PIN SCORE": 0.0, "LIQUIDITY": "NONE",
        "SWEEP": "NONE", "REVERSAL": "NONE", "EQUAL HIGH": "NO", "EQUAL LOW": "NO",
        "BIG MOVEMENT": "NO", "BIG MOVE SCORE": 0.0, "STRUCTURE": "NONE",
        "5M TREND": data_5m.get("structure_trend", "N/A"),
        "15M TREND": data_15m.get("structure_trend", "N/A"),
        "1H TREND": data_1h.get("structure_trend", "N/A"),
        "RVOL": data_5m.get("rvol", 0), "RSI": data_5m.get("rsi", 50),
        "PRESSURE": data_5m.get("pressure_trend", "N/A"), "REASON": ""
    }
    if df_5m is None or len(df_5m) < 30:
        out["REASON"] = "Insufficient 5M candles"
        return out
    try:
        d = df_5m.reset_index(drop=True).copy()
        last = d.iloc[-1]
        o, h, l, c, v = [float(last[x]) for x in ["Open", "High", "Low", "Close", "Volume"]]
        body = abs(c-o)
        rng = max(h-l, 1e-9)
        upper_wick = h-max(o,c)
        lower_wick = min(o,c)-l
        atr_s = calculate_atr(d, 14)
        atr = float(atr_s.iloc[-1]) if pd.notna(atr_s.iloc[-1]) else max(c*0.005, 0.01)
        vwap_s = calculate_vwap(d)
        vwap = float(vwap_s.iloc[-1]) if len(vwap_s) else c
        rsi = float(data_5m.get("rsi", 50) or 50)
        rvol = float(data_5m.get("rvol", 0) or 0)
        macd_bull = bool(data_5m.get("macd_bullish", False))
        ema_trend = data_5m.get("ema_trend", "NEUTRAL")
        structure_trend = data_5m.get("structure_trend", "NEUTRAL")
        bull = c > o; bear = c < o
        strong_bull = bull and body/rng*100 >= 55
        strong_bear = bear and body/rng*100 >= 55

        # Confirmed liquidity pivots. Current candle is never used as a pivot.
        ph, pl = _confirmed_pivots(d, left=PIN_PIVOT_LEN, right=PIN_PIVOT_LEN)
        last_hi = ph[-1][1] if ph else None
        prev_hi = ph[-2][1] if len(ph) >= 2 else None
        last_lo = pl[-1][1] if pl else None
        prev_lo = pl[-2][1] if len(pl) >= 2 else None
        eq_hi = last_hi is not None and prev_hi is not None and abs(last_hi-prev_hi) <= atr*PIN_EQUAL_ATR_TOL
        eq_lo = last_lo is not None and prev_lo is not None and abs(last_lo-prev_lo) <= atr*PIN_EQUAL_ATR_TOL
        sweep_buy = last_hi is not None and h > last_hi and c < last_hi and upper_wick > body
        sweep_sell = last_lo is not None and l < last_lo and c > last_lo and lower_wick > body
        bullish_sweep = sweep_sell
        bearish_sweep = sweep_buy

        bullish_reversal = bullish_sweep and bull and c > vwap and rsi > 45
        bearish_reversal = bearish_sweep and bear and c < vwap and rsi < 55

        # Pine-style confluence score.
        buy = 0.0; sell = 0.0
        buy += 25 if ema_trend == "BULLISH" and structure_trend == "BULLISH" else 15 if structure_trend == "BULLISH" else 0
        sell += 25 if ema_trend == "BEARISH" and structure_trend == "BEARISH" else 15 if structure_trend == "BEARISH" else 0
        buy += 15 if rsi >= 55 else 7 if rsi >= 50 else 0
        sell += 15 if rsi <= 45 else 7 if rsi <= 50 else 0
        buy += 15 if c > vwap else 0; sell += 15 if c < vwap else 0
        buy += 20 if macd_bull and float(data_5m.get("macd_hist", 0) or 0) > 0 else 10 if macd_bull else 0
        sell += 20 if (not macd_bull) and float(data_5m.get("macd_hist", 0) or 0) < 0 else 10 if not macd_bull else 0
        buy += 10 if rvol >= 1.5 and bull else 0
        sell += 10 if rvol >= 1.5 and bear else 0
        buy += 5 if strong_bull else 0; sell += 5 if strong_bear else 0
        buy += 10 if bullish_sweep else 0; sell += 10 if bearish_sweep else 0
        buy += 10 if bullish_reversal else 0; sell += 10 if bearish_reversal else 0
        pin_score = min(100.0, max(buy, sell))
        direction = "BUY" if buy > sell else "SELL" if sell > buy else "WAIT"
        if pin_score >= PIN_STRONG_CONFIDENCE and direction != "WAIT":
            pin_signal = f"{'🟢 STRONG BUY' if direction=='BUY' else '🔴 STRONG SELL'}"
        elif pin_score >= PIN_MIN_CONFIDENCE and direction != "WAIT":
            pin_signal = f"{'🟢 BUY' if direction=='BUY' else '🔴 SELL'}"
        else:
            pin_signal = "🟡 WAIT"

        # Existing BIG MOVE engine is reused; no duplicate scan logic.
        bm = detect_big_move_setup(d)
        structure = bm.get("structure", "NONE")
        liquidity = "EQ HIGH" if eq_hi else "EQ LOW" if eq_lo else "HIGH" if last_hi is not None else "LOW" if last_lo is not None else "NONE"
        sweep = "🟢 LOW SWEPT" if bullish_sweep else "🔴 HIGH SWEPT" if bearish_sweep else "NONE"
        reversal = "🟢 BULL REVERSAL" if bullish_reversal else "🔴 BEAR REVERSAL" if bearish_reversal else "NONE"
        out.update({
            "PIN SIGNAL": pin_signal, "PIN SCORE": round(pin_score, 1),
            "LIQUIDITY": liquidity, "SWEEP": sweep, "REVERSAL": reversal,
            "EQUAL HIGH": "YES" if eq_hi else "NO", "EQUAL LOW": "YES" if eq_lo else "NO",
            "BIG MOVEMENT": bm.get("signal", "NO BIG MOVE"),
            "BIG MOVE SCORE": bm.get("score", 0.0), "STRUCTURE": structure,
            "REASON": " | ".join([x for x in [
                "EQ HIGH" if eq_hi else "", "EQ LOW" if eq_lo else "",
                "LOW SWEEP" if bullish_sweep else "HIGH SWEEP" if bearish_sweep else "",
                "BULL REVERSAL" if bullish_reversal else "BEAR REVERSAL" if bearish_reversal else "",
                "BIG MOVE" if bm.get("direction") in ["UP", "DOWN"] else ""
            ] if x]) or "No PIN confirmation"
        })
        return out
    except Exception as e:
        out["REASON"] = f"PIN error: {type(e).__name__}"
        return out


def _show_pin_rules_tab(fyers) -> None:
    st.markdown("### 📌 PIN RULES — Liquidity + Reversal + Big Movement")
    st.caption("Additional analysis only. Existing Scanner tabs and scanner logic are not modified.")

    source = st.selectbox("Source", ["NSE Stocks", "F&O Stocks"], key="pin_source")
    source_key = "nse_df" if source == "NSE Stocks" else "fo_df"
    base_df = st.session_state.get(source_key)

    # PIN RULES can run independently. If the main NSE/F&O scanner has not
    # been run yet, build candidates directly from the loaded symbol universe.
    if base_df is None or base_df.empty:
        raw_symbols = (
            st.session_state.get("all_symbols", [])
            if source == "NSE Stocks"
            else st.session_state.get("fo_symbols", [])
        )
        if raw_symbols:
            base_df = pd.DataFrame({
                "Symbol": [str(x).replace("NSE:", "").replace("-EQ", "") for x in raw_symbols],
                "LTP": ["N/A"] * len(raw_symbols),
            })
        else:
            base_df = pd.DataFrame(columns=["Symbol", "LTP"])

    with st.expander("📖 PIN Rules", expanded=False):
        st.markdown("""
        **Liquidity:** confirmed Pivot High/Low → Equal High/Low → liquidity level.
        **Sweep:** High breaks and closes back below = bearish; Low breaks and closes back above = bullish.
        **Reversal:** Sweep + candle direction + VWAP + RSI confirmation.
        **Confluence:** Trend + RSI + VWAP + MACD + RVOL + candle + sweep + reversal.
        **Big Movement:** existing consolidation-breakout + candle + RVOL + structure engine.
        """)

    if base_df is None or base_df.empty:
        st.warning(f"⚠️ No {source} symbols are available. Check the symbol master.")
        return

    max_scan = min(PIN_MAX_SCAN, len(base_df))
    c1, c2, c3 = st.columns(3)
    with c1:
        pin_limit = st.number_input("PIN scan limit", 1, max_scan, min(30, max_scan), 1, key="pin_limit")
    with c2:
        pin_min = st.slider("Minimum PIN score", 50, 100, PIN_MIN_CONFIDENCE, 1, key="pin_min_score")
    with c3:
        pin_mode = st.selectbox("Show", ["ALL", "BUY ONLY", "SELL ONLY", "STRONG ONLY"], key="pin_mode")

    candidates = base_df.copy()
    if "AI CONFIDENCE %" in candidates.columns:
        candidates["__conf"] = pd.to_numeric(candidates["AI CONFIDENCE %"], errors="coerce").fillna(0)
        candidates = candidates.sort_values("__conf", ascending=False)
    candidates = candidates.head(int(pin_limit))

    if st.button("📌 RUN PIN RULES", key="pin_run", use_container_width=True):
        rows = []; errors = []
        progress = st.progress(0.0)
        total = len(candidates)
        for n, (_, row) in enumerate(candidates.iterrows(), 1):
            symbol = str(row.get("Symbol", "")).strip()
            fyers_symbol = symbol if symbol.startswith("NSE:") else f"NSE:{symbol}-EQ"
            try:
                a5 = analyze_timeframe(fyers, fyers_symbol, "5")
                a15 = analyze_timeframe(fyers, fyers_symbol, "15")
                a1h = analyze_timeframe(fyers, fyers_symbol, "60")
                if a5.get("status") != "OK" or a5.get("df") is None:
                    errors.append(f"{symbol}: 5M data unavailable")
                    progress.progress(n / max(total, 1))
                    continue
                pin = calculate_pin_rules(a5.get("df"), a5.get("data", {}), a15.get("data", {}), a1h.get("data", {}))
                pin["Symbol"] = symbol.replace("NSE:", "").replace("-EQ", "")
                pin["LTP"] = row.get("LTP", "N/A")
                pin["AI CONFIDENCE %"] = row.get("AI CONFIDENCE %", "N/A")
                pin["AI SIGNAL"] = row.get("AI SIGNAL", "N/A")
                rows.append(pin)
            except Exception as e:
                errors.append(f"{symbol}: {type(e).__name__}")
            progress.progress(n / max(total, 1))
        progress.empty()
        result = pd.DataFrame(rows)
        if not result.empty:
            result["__score"] = pd.to_numeric(result["PIN SCORE"], errors="coerce").fillna(0)
            result = result[result["__score"] >= pin_min]
            if pin_mode == "BUY ONLY":
                result = result[result["PIN SIGNAL"].astype(str).str.contains("BUY", na=False)]
            elif pin_mode == "SELL ONLY":
                result = result[result["PIN SIGNAL"].astype(str).str.contains("SELL", na=False)]
            elif pin_mode == "STRONG ONLY":
                result = result[result["PIN SIGNAL"].astype(str).str.contains("STRONG", na=False)]
            result = result.drop(columns=["__score"], errors="ignore")
        st.session_state["pin_df"] = result
        st.session_state["pin_errors"] = errors

    pin_df = st.session_state.get("pin_df")
    if pin_df is not None and not pin_df.empty:
        pc1, pc2, pc3, pc4 = st.columns(4)
        pc1.metric("📌 PIN SETUPS", len(pin_df))
        pc2.metric("🟢 BUY", int(pin_df["PIN SIGNAL"].astype(str).str.contains("BUY", na=False).sum()))
        pc3.metric("🔴 SELL", int(pin_df["PIN SIGNAL"].astype(str).str.contains("SELL", na=False).sum()))
        pc4.metric("💧 SWEEPS", int((pin_df["SWEEP"].astype(str) != "NONE").sum()))
        display_cols = [c for c in ["Symbol","LTP","PIN SIGNAL","PIN SCORE","LIQUIDITY","SWEEP","REVERSAL","EQUAL HIGH","EQUAL LOW","BIG MOVEMENT","BIG MOVE SCORE","STRUCTURE","5M TREND","15M TREND","1H TREND","RVOL","RSI","PRESSURE","AI CONFIDENCE %","AI SIGNAL","REASON"] if c in pin_df.columns]
        st.dataframe(pin_df[display_cols], use_container_width=True, height=500)
        st.download_button("📥 DOWNLOAD PIN RULES EXCEL", _format_excel_output(pin_df, "PIN_RULES"), f"PIN_RULES_{_now_ist().strftime('%Y%m%d_%H%M')}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="pin_excel")
    elif "pin_df" in st.session_state:
        st.warning("No stocks matched the selected PIN rules.")
    if st.session_state.get("pin_errors"):
        st.caption(f"⚠️ {len(st.session_state['pin_errors'])} symbols could not be analyzed.")


# ════════════════════════════════════════════════════════════════════════════════
# ADDITIONAL FULL-UNIVERSE PIN + AMD SCANNERS
# These are additive only. Existing NSE/F&O/Momentum/PIN Rules code is retained.
# ════════════════════════════════════════════════════════════════════════════════

def _scan_universe_map(all_symbols, fo_symbols, source):
    """Build a de-duplicated NSE/F&O universe without changing the original lists."""
    nse = _validate_symbols(all_symbols or [])
    fo = _validate_symbols(fo_symbols or [])
    if source == "NSE Stocks":
        return [(s, "NSE") for s in nse]
    if source == "F&O Stocks":
        return [(s, "F&O") for s in fo]

    # ALL = union, with F&O label taking priority for overlapping symbols.
    fo_set = set(fo)
    ordered = []
    seen = set()
    for s in fo:
        if s not in seen:
            seen.add(s)
            ordered.append((s, "F&O"))
    for s in nse:
        if s not in seen:
            seen.add(s)
            ordered.append((s, "NSE"))
    return ordered


def _fetch_full_pin_signal(fyers, symbol: str, source: str):
    """Worker: fresh 5M/15M/1H data -> existing PIN rules engine."""
    ticker = symbol.replace("NSE:", "").replace("-EQ", "")
    try:
        a5 = analyze_timeframe(fyers, symbol, "5")
        a15 = analyze_timeframe(fyers, symbol, "15")
        a1h = analyze_timeframe(fyers, symbol, "60")
        if a5.get("status") != "OK" or a5.get("df") is None:
            return None, f"{ticker}: 5M data unavailable"
        pin = calculate_pin_rules(
            a5.get("df"), a5.get("data", {}),
            a15.get("data", {}), a1h.get("data", {})
        )
        d5 = a5.get("data", {}) or {}
        pin["Symbol"] = ticker
        pin["SOURCE"] = source
        pin["LTP"] = d5.get("last_close", "N/A")
        pin["5M TREND"] = d5.get("structure_trend", "N/A")
        pin["15M TREND"] = (a15.get("data", {}) or {}).get("structure_trend", "N/A")
        pin["1H TREND"] = (a1h.get("data", {}) or {}).get("structure_trend", "N/A")
        pin["RVOL"] = d5.get("rvol", 0)
        pin["RSI"] = d5.get("rsi", 50)
        pin["PRESSURE"] = d5.get("pressure_trend", "N/A")
        return pin, None
    except Exception as e:
        return None, f"{ticker}: {type(e).__name__}: {str(e)[:100]}"


def _run_full_pin_scan(fyers, universe, pin_min=70, pin_mode="ALL"):
    """Threaded full-universe PIN scanner. Existing PIN calculation is reused."""
    pairs = list(universe or [])
    stats = ScanStats(total=len(pairs))
    results, errors = [], []
    if not pairs:
        return results, errors, stats

    progress = st.progress(0.0, text=f"PIN Scan 0 / {len(pairs)}")
    done = 0
    for i in range(0, len(pairs), BATCH_SIZE):
        batch = pairs[i:i + BATCH_SIZE]
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {
                executor.submit(_fetch_full_pin_signal, fyers, symbol, source): (symbol, source)
                for symbol, source in batch
            }
            for future in as_completed(futures):
                res, err = future.result()
                if res:
                    results.append(res)
                if err:
                    errors.append(err)
                stats.record(has_result=bool(res), has_error=bool(err))
                done += 1
                progress.progress(done / max(len(pairs), 1), text=f"PIN Scan {done} / {len(pairs)}")
        if i + BATCH_SIZE < len(pairs):
            time.sleep(BATCH_PAUSE_SECONDS)
    progress.empty()

    df = pd.DataFrame(results)
    if not df.empty:
        df["__score"] = pd.to_numeric(df.get("PIN SCORE", 0), errors="coerce").fillna(0)
        df = df[df["__score"] >= float(pin_min)].copy()
        sig = df.get("PIN SIGNAL", pd.Series("", index=df.index)).astype(str)
        if pin_mode == "BUY ONLY":
            df = df[sig.str.contains("BUY", na=False)]
        elif pin_mode == "SELL ONLY":
            df = df[sig.str.contains("SELL", na=False)]
        elif pin_mode == "STRONG ONLY":
            df = df[sig.str.contains("STRONG", na=False)]
        df = df.drop(columns=["__score"], errors="ignore")
        if "PIN SCORE" in df.columns:
            df = df.sort_values("PIN SCORE", ascending=False)
    return df.to_dict("records") if not df.empty else [], errors, stats


def _fetch_amd_signal_full(fyers, symbol: str, source: str):
    """Worker: fresh completed 5M + 15M data -> AMD inference engine."""
    ticker = symbol.replace("NSE:", "").replace("-EQ", "")
    try:
        a5 = analyze_timeframe(fyers, symbol, "5")
        if a5.get("status") != "OK" or a5.get("df") is None:
            return None, f"{ticker}: 5M data unavailable"
        a15 = analyze_timeframe(fyers, symbol, "15")
        amd = calculate_amd_signal(a5.get("df"), a15.get("df"))
        d5 = a5.get("data", {}) or {}
        d15 = a15.get("data", {}) or {}
        row = {
            "Symbol": ticker,
            "SOURCE": source,
            "LTP": d5.get("last_close", "N/A"),
            "AMD PHASE": amd.get("AMD PHASE", "NEUTRAL"),
            "AMD SIGNAL": amd.get("AMD SIGNAL", "WAIT"),
            "AMD SCORE": amd.get("AMD SCORE", 0),
            "AMD BUY SCORE": amd.get("AMD BUY SCORE", 0),
            "AMD SELL SCORE": amd.get("AMD SELL SCORE", 0),
            "AMD CONFIDENCE %": amd.get("AMD CONFIDENCE %", 0),
            "AMD SWEEP": amd.get("AMD SWEEP", "NONE"),
            "AMD RANGE HIGH": amd.get("AMD RANGE HIGH"),
            "AMD RANGE LOW": amd.get("AMD RANGE LOW"),
            "AMD RVOL": amd.get("AMD RVOL", 0),
            "AMD REASON": amd.get("AMD REASON", ""),
            "5M TREND": d5.get("structure_trend", "N/A"),
            "15M TREND": d15.get("structure_trend", "N/A"),
            "5M RSI": d5.get("rsi", 50),
            "5M RVOL": d5.get("rvol", 0),
            "BUY PRESSURE %": d5.get("buying_pressure", "N/A"),
            "SELL PRESSURE %": d5.get("selling_pressure", "N/A"),
        }
        return row, None
    except Exception as e:
        return None, f"{ticker}: {type(e).__name__}: {str(e)[:100]}"


def _run_amd_scan(fyers, universe):
    """Threaded full-universe AMD scanner. Does not alter existing scanners."""
    pairs = list(universe or [])
    stats = ScanStats(total=len(pairs))
    results, errors = [], []
    if not pairs:
        return results, errors, stats

    progress = st.progress(0.0, text=f"AMD Scan 0 / {len(pairs)}")
    done = 0
    for i in range(0, len(pairs), BATCH_SIZE):
        batch = pairs[i:i + BATCH_SIZE]
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {
                executor.submit(_fetch_amd_signal_full, fyers, symbol, source): (symbol, source)
                for symbol, source in batch
            }
            for future in as_completed(futures):
                res, err = future.result()
                if res:
                    results.append(res)
                if err:
                    errors.append(err)
                stats.record(has_result=bool(res), has_error=bool(err))
                done += 1
                progress.progress(done / max(len(pairs), 1), text=f"AMD Scan {done} / {len(pairs)}")
        if i + BATCH_SIZE < len(pairs):
            time.sleep(BATCH_PAUSE_SECONDS)
    progress.empty()

    results.sort(key=lambda x: (
        float(x.get("AMD SCORE", 0) or 0),
        float(x.get("AMD CONFIDENCE %", 0) or 0)
    ), reverse=True)
    return results, errors, stats


def _show_pin_full_scan_tab(fyers, all_symbols, fo_symbols):
    st.markdown("### 📌 PIN SCANNER — FULL NSE + F&O UNIVERSE")
    st.caption("Independent full-universe PIN scan. Existing scanner tabs and rules are unchanged.")

    source = st.radio(
        "PIN Universe",
        ["ALL NSE + F&O", "NSE Stocks", "F&O Stocks"],
        horizontal=True,
        key="pin_full_source",
    )
    universe = _scan_universe_map(all_symbols, fo_symbols, "NSE Stocks" if source == "NSE Stocks" else "F&O Stocks" if source == "F&O Stocks" else "ALL")
    total = len(universe)
    st.metric("AVAILABLE STOCKS", f"{total:,}")
    st.info("📌 This scanner uses fresh 5M + 15M + 1H candles and the existing PIN rules engine.")

    c1, c2, c3 = st.columns(3)
    with c1:
        pin_limit = st.number_input("PIN scan limit (0 = ALL)", 0, max(total, 1), min(total, 300), 25, key="pin_full_limit") if total else 0
    with c2:
        pin_min = st.slider("Minimum PIN score", 50, 100, PIN_MIN_CONFIDENCE, 1, key="pin_full_min")
    with c3:
        pin_mode = st.selectbox("Show", ["ALL", "BUY ONLY", "SELL ONLY", "STRONG ONLY"], key="pin_full_mode")

    scan_pairs = universe if pin_limit == 0 else universe[:int(pin_limit)]
    if st.button(f"📌 RUN PIN SCANNER ({len(scan_pairs):,} STOCKS)", key="pin_full_run", type="primary", use_container_width=True):
        with st.spinner(f"Running PIN scanner on {len(scan_pairs):,} stocks…"):
            rows, errors, stats = _run_full_pin_scan(fyers, scan_pairs, pin_min=pin_min, pin_mode=pin_mode)
        st.session_state["pin_full_df"] = _add_signal_time_columns(pd.DataFrame(rows), "PIN SIGNAL")
        st.session_state["pin_full_errors"] = errors
        st.session_state["pin_full_stats"] = stats
        st.session_state["pin_full_time"] = _generated_timestamp()

    if "pin_full_stats" in st.session_state:
        _display_scan_summary(st.session_state["pin_full_stats"])
        st.caption(f"Last PIN scan: {st.session_state.get('pin_full_time', 'N/A')}")

    df = st.session_state.get("pin_full_df")
    if df is not None and not df.empty:
        df = _add_signal_time_columns(df, "PIN SIGNAL")
        st.session_state["pin_full_df"] = df
        buy_n = int(df.get("PIN SIGNAL", pd.Series(dtype=str)).astype(str).str.contains("BUY", na=False).sum())
        sell_n = int(df.get("PIN SIGNAL", pd.Series(dtype=str)).astype(str).str.contains("SELL", na=False).sum())
        sweep_n = int((df.get("SWEEP", pd.Series(dtype=str)).astype(str) != "NONE").sum()) if "SWEEP" in df.columns else 0
        c1, c2, c3 = st.columns(3)
        c1.metric("📌 PIN SETUPS", len(df))
        c2.metric("🟢 BUY", buy_n)
        c3.metric("🔴 SELL", sell_n)
        st.dataframe(df, use_container_width=True, height=550)
        _excel_download_button(df, "PIN_FULL_SCAN", "pin_full_excel", label="📥 DOWNLOAD PIN EXCEL")
    elif "pin_full_stats" in st.session_state:
        st.warning("PIN scan completed — no rows matched the selected PIN score/filter.")

    errors = st.session_state.get("pin_full_errors", [])
    if errors:
        with st.expander(f"⚠️ PIN scan errors ({len(errors)})", expanded=False):
            st.dataframe(pd.DataFrame({"Error": errors}), use_container_width=True)


def _show_amd_scan_tab(fyers, all_symbols, fo_symbols):
    st.markdown("### 🧠 AMD SCANNER — ACCUMULATION / MANIPULATION / DISTRIBUTION")
    st.caption("Fresh completed-candle AMD inference for NSE equities and F&O stocks. This is a rule-based market-structure heuristic, not proof of institutional intent.")

    source = st.radio(
        "AMD Universe",
        ["ALL NSE + F&O", "NSE Stocks", "F&O Stocks"],
        horizontal=True,
        key="amd_source",
    )
    universe = _scan_universe_map(all_symbols, fo_symbols, "NSE Stocks" if source == "NSE Stocks" else "F&O Stocks" if source == "F&O Stocks" else "ALL")
    total = len(universe)
    st.metric("AVAILABLE STOCKS", f"{total:,}")

    c1, c2 = st.columns(2)
    with c1:
        amd_limit = st.number_input("AMD scan limit (0 = ALL)", 0, max(total, 1), min(total, 300), 25, key="amd_limit") if total else 0
    with c2:
        amd_show = st.selectbox("Show AMD", ["ALL", "ACCUMULATION", "MANIPULATION", "DISTRIBUTION", "BUY SIGNALS", "SELL SIGNALS"], key="amd_show")

    scan_pairs = universe if amd_limit == 0 else universe[:int(amd_limit)]
    if st.button(f"🧠 RUN AMD SCANNER ({len(scan_pairs):,} STOCKS)", key="amd_run", type="primary", use_container_width=True):
        with st.spinner(f"Running AMD scanner on {len(scan_pairs):,} stocks…"):
            rows, errors, stats = _run_amd_scan(fyers, scan_pairs)
        st.session_state["amd_df"] = _add_signal_time_columns(pd.DataFrame(rows), "AMD SIGNAL")
        st.session_state["amd_errors"] = errors
        st.session_state["amd_stats"] = stats
        st.session_state["amd_time"] = _generated_timestamp()

    if "amd_stats" in st.session_state:
        _display_scan_summary(st.session_state["amd_stats"])
        st.caption(f"Last AMD scan: {st.session_state.get('amd_time', 'N/A')}")

    df = st.session_state.get("amd_df")
    if df is not None and not df.empty:
        df = _add_signal_time_columns(df, "AMD SIGNAL")
        st.session_state["amd_df"] = df
        out = df.copy()
        phase = out.get("AMD PHASE", pd.Series("", index=out.index)).astype(str)
        sig = out.get("AMD SIGNAL", pd.Series("", index=out.index)).astype(str)
        if amd_show in ["ACCUMULATION", "MANIPULATION", "DISTRIBUTION"]:
            out = out[phase == amd_show]
        elif amd_show == "BUY SIGNALS":
            out = out[sig.str.contains("BUY", na=False)]
        elif amd_show == "SELL SIGNALS":
            out = out[sig.str.contains("SELL", na=False)]

        buy_n = int(sig.str.contains("BUY", na=False).sum())
        sell_n = int(sig.str.contains("SELL", na=False).sum())
        acc_n = int((phase == "ACCUMULATION").sum())
        manip_n = int((phase == "MANIPULATION").sum())
        dist_n = int((phase == "DISTRIBUTION").sum())
        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("TOTAL", len(df))
        c2.metric("🟢 BUY", buy_n)
        c3.metric("🔴 SELL", sell_n)
        c4.metric("📥 ACCUMULATION", acc_n)
        c5.metric("📤 DISTRIBUTION", dist_n)
        if manip_n:
            st.info(f"🟠 MANIPULATION / SWEEP: {manip_n}")

        st.caption(f"📄 AMD REPORT: {len(out)} rows shown")
        st.dataframe(out, use_container_width=True, height=550)
        _excel_download_button(out, "AMD_SCAN", "amd_excel", label="📥 DOWNLOAD AMD EXCEL")
    elif "amd_stats" in st.session_state:
        st.warning("AMD scan completed — no analyzable rows returned.")

    errors = st.session_state.get("amd_errors", [])
    if errors:
        with st.expander(f"⚠️ AMD scan errors ({len(errors)})", expanded=False):
            st.dataframe(pd.DataFrame({"Error": errors}), use_container_width=True)

# ════════════════════════════════════════════════════════════════════════════════
# MAIN APP - V17 WITH NEW MOMENTUM MOVERS TAB
# ════════════════════════════════════════════════════════════════════════════════
def show_scanner(fyers) -> None:
    """Streamlit main app - NSE AI PRO V17 with MOMENTUM MOVERS"""
    
    try:
        st.set_page_config(page_title="NSE AI PRO V17", layout="wide")
    except:
        pass
    
    st.title("🚀 NSE AI PRO V17 — Professional Intraday + Swing Scanner")
    st.caption(f"🕒 Current Time (IST): {_now_ist().strftime('%d-%b-%Y %H:%M:%S')} IST | Multi-Timeframe + Momentum Engine")
    
    try:
        all_symbols = load_nse_equity_symbols()
        fo_symbols = load_fo_stocks()
        # Keep universes available to independent PIN/AMD tabs without
        # requiring the main scanner to be executed first.
        st.session_state["all_symbols"] = all_symbols
        st.session_state["fo_symbols"] = fo_symbols
    except Exception as e:
        st.error(f"❌ Error loading symbols: {e}")
        logger.error(f"Symbol loading error: {e}")
        return
    
    if not all_symbols:
        st.error("❌ No symbols loaded — check FYERS API access.")
        return
    
    st.caption(f"📊 NSE Equities: {len(all_symbols)} | 📈 F&O Stocks: {len(fo_symbols)}")
    
    tabs = st.tabs([
        "🇮🇳 NSE STOCKS",
        "📊 F&O STOCKS",
        "⚡ MOMENTUM MOVERS",
        "⚡ LIVE INTRADAY",
        "🔥 STRONG SIGNALS",
        "📈 SWING (GOLDEN/DEATH CROSS)",
        "🧠 ADDITIONAL ANALYSIS",
        "📊 MARKET DASHBOARD",
        "⚙️ SETTINGS",
        "📌 PIN RULES",
        "📌 PIN FULL SCAN",
        "🧠 AMD SCAN"
    ])
    
    # ════════════════════════════════════════════════════════════════════════════════
    # TAB 0: NSE STOCKS
    # ════════════════════════════════════════════════════════════════════════════════
    with tabs[0]:
        st.markdown("### NSE Equity Stocks Scanner\n✅ Strict validation - only high-quality signals")
        
        col1, col2 = st.columns([3, 1])
        with col1:
            nse_limit = st.number_input("Scan limit (0=all)", min_value=0, max_value=len(all_symbols),
                                       value=min(500, len(all_symbols)), step=50, key="nse_limit")
        with col2:
            st.metric("Available", len(all_symbols))
        
        nse_universe = all_symbols if nse_limit == 0 else all_symbols[:nse_limit]
        
        if st.button(f"🔍 SCAN NSE ({len(nse_universe)} stocks)", key="nse_run"):
            with st.spinner("Analyzing NSE stocks…"):
                nse_results, nse_errors, nse_stats = run_nse_scan(fyers, nse_universe)
                st.session_state["nse_df"] = pd.DataFrame(nse_results) if nse_results else pd.DataFrame()
                st.session_state["nse_errors"] = nse_errors
                st.session_state["nse_stats"] = nse_stats
        
        if "nse_stats" in st.session_state:
            _display_scan_summary(st.session_state["nse_stats"])
        
        nse_df = st.session_state.get("nse_df")
        if nse_df is not None and not nse_df.empty:
            nse_df = _add_signal_time_columns(nse_df, "AI SIGNAL")
            st.session_state["nse_df"] = nse_df
            try:
                st.info(f"📊 Loaded: {len(nse_df)} signals")
                
                col_f1, col_f2, col_f3 = st.columns(3)
                
                with col_f1:
                    nse_min_conf = st.slider("Min Confidence %", 0, 100, 70, step=5, key="nse_conf_filter")
                with col_f2:
                    nse_signal = st.selectbox("Signal", ["ALL", "BUY", "SELL", "STRONG ONLY"], key="nse_signal_filter")
                with col_f3:
                    nse_sort = st.selectbox("Sort By", ["AI CONFIDENCE %", "LTP", "RVOL"], key="nse_sort_filter")
                
                nse_filtered = nse_df.copy()
                try:
                    confidence_col = pd.to_numeric(nse_filtered["AI CONFIDENCE %"], errors='coerce')
                    nse_filtered = nse_filtered[confidence_col >= nse_min_conf]
                except:
                    pass
                
                if nse_signal != "ALL":
                    try:
                        normalized = nse_filtered["AI SIGNAL"].apply(normalize_signal)
                        if nse_signal == "BUY":
                            nse_filtered = nse_filtered[normalized == "BUY"]
                        elif nse_signal == "SELL":
                            nse_filtered = nse_filtered[normalized == "SELL"]
                        elif nse_signal == "STRONG ONLY":
                            nse_filtered = nse_filtered[nse_filtered["AI SIGNAL"].astype(str).str.contains("STRONG", na=False)]
                    except:
                        pass
                
                try:
                    if nse_sort == "LTP":
                        nse_filtered = nse_filtered.sort_values("LTP", ascending=False)
                    elif nse_sort == "RVOL":
                        nse_filtered = nse_filtered.sort_values("RVOL", ascending=False)
                    else:
                        nse_filtered = nse_filtered.sort_values("AI CONFIDENCE %", ascending=False)
                except:
                    pass
                
                st.caption(f"📄 NSE REPORT: {len(nse_filtered)} rows shown")
                st.dataframe(nse_filtered, use_container_width=True, height=500)
                
                st.markdown("### 📥 Download")
                col_d1, col_d2, col_d3 = st.columns(3)
                
                with col_d1:
                    try:
                        excel_data = _format_excel_output(nse_filtered, "NSE")
                        st.download_button("📊 Excel", excel_data, f"NSE_{_now_ist().strftime('%Y%m%d_%H%M')}.xlsx",
                                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="nse_xls")
                    except Exception as e:
                        st.error(f"❌ Excel: {str(e)[:50]}")
                
                with col_d2:
                    try:
                        csv_data = to_csv_bytes(nse_filtered)
                        st.download_button("📄 CSV", csv_data, f"NSE_{_now_ist().strftime('%Y%m%d_%H%M')}.csv",
                                         "text/csv", key="nse_csv")
                    except Exception as e:
                        st.error(f"❌ CSV: {str(e)[:50]}")
                
                with col_d3:
                    try:
                        json_data = to_json_bytes(nse_filtered)
                        st.download_button("📋 JSON", json_data, f"NSE_{_now_ist().strftime('%Y%m%d_%H%M')}.json",
                                         "application/json", key="nse_json")
                    except Exception as e:
                        st.error(f"❌ JSON: {str(e)[:50]}")
            
            except Exception as e:
                st.error(f"❌ NSE Tab Error: {str(e)[:100]}")
        else:
            st.info("👈 Click 'SCAN NSE' to start")
    
    # ════════════════════════════════════════════════════════════════════════════════
    # TAB 1: F&O STOCKS
    # ════════════════════════════════════════════════════════════════════════════════
    with tabs[1]:
        st.markdown("### F&O Stocks Scanner\n✅ Strict validation + options analysis")
        
        col1, col2 = st.columns([3, 1])
        with col1:
            fo_limit = st.number_input("Scan limit (0=all)", min_value=0, max_value=len(fo_symbols),
                                      value=min(200, len(fo_symbols)), step=25, key="fo_limit")
        with col2:
            st.metric("Available", len(fo_symbols))
        
        fo_universe = fo_symbols if fo_limit == 0 else fo_symbols[:fo_limit]
        
        # F&O scan button: keep the scan result in session_state so it
        # survives Streamlit reruns and can be downloaded immediately.
        if st.button(
            f"🔍 SCAN F&O ({len(fo_universe)} stocks)",
            key="fo_run",
            type="primary",
            use_container_width=True,
        ):
            with st.spinner(f"Analyzing F&O stocks ({len(fo_universe)} symbols)…"):
                try:
                    fo_results, fo_errors, fo_stats = run_fo_scan(fyers, fo_universe)
                    st.session_state["fo_df"] = (
                        pd.DataFrame(fo_results) if fo_results else pd.DataFrame()
                    )
                    st.session_state["fo_errors"] = fo_errors or []
                    st.session_state["fo_stats"] = fo_stats
                except Exception as e:
                    st.session_state["fo_df"] = pd.DataFrame()
                    st.session_state["fo_errors"] = [str(e)]
                    st.error(f"❌ F&O scan failed: {type(e).__name__}: {str(e)[:180]}")
        
        if "fo_stats" in st.session_state:
            _display_scan_summary(st.session_state["fo_stats"])
        
        fo_df = st.session_state.get("fo_df")
        if fo_df is not None and not fo_df.empty:
            fo_df = _add_signal_time_columns(fo_df, "AI SIGNAL")
            st.session_state["fo_df"] = fo_df
            try:
                st.info(f"📊 Loaded: {len(fo_df)} signals")
                
                col_f1, col_f2, col_f3, col_f4 = st.columns(4)
                with col_f1:
                    fo_min_conf = st.slider("Min Confidence %", 0, 100, 70, step=5, key="fo_conf_filter")
                with col_f2:
                    fo_signal = st.selectbox("Signal", ["ALL", "BUY", "SELL", "STRONG ONLY"], key="fo_signal_filter")
                with col_f3:
                    fo_opt = st.selectbox("Options", ["ALL", "BULLISH", "BEARISH"], key="fo_opt_filter")
                with col_f4:
                    fo_sort = st.selectbox("Sort By", ["AI CONFIDENCE %", "LTP", "PCR"], key="fo_sort_filter")
                
                fo_filtered = fo_df.copy()
                try:
                    confidence_col = pd.to_numeric(fo_filtered["AI CONFIDENCE %"], errors='coerce')
                    fo_filtered = fo_filtered[confidence_col >= fo_min_conf]
                except:
                    pass
                
                if fo_signal != "ALL":
                    try:
                        normalized = fo_filtered["AI SIGNAL"].apply(normalize_signal)
                        if fo_signal == "BUY":
                            fo_filtered = fo_filtered[normalized == "BUY"]
                        elif fo_signal == "SELL":
                            fo_filtered = fo_filtered[normalized == "SELL"]
                        elif fo_signal == "STRONG ONLY":
                            fo_filtered = fo_filtered[fo_filtered["AI SIGNAL"].astype(str).str.contains("STRONG", na=False)]
                    except:
                        pass
                
                if fo_opt != "ALL":
                    try:
                        fo_filtered = fo_filtered[fo_filtered["OPTIONS BIAS"].astype(str).str.contains(fo_opt, na=False)]
                    except:
                        pass
                
                try:
                    if fo_sort == "PCR":
                        fo_filtered = fo_filtered.sort_values("PCR", ascending=False)
                    elif fo_sort == "LTP":
                        fo_filtered = fo_filtered.sort_values("LTP", ascending=False)
                    else:
                        fo_filtered = fo_filtered.sort_values("AI CONFIDENCE %", ascending=False)
                except:
                    pass
                
                st.caption(f"📄 F&O REPORT: {len(fo_filtered)} rows shown")
                st.dataframe(fo_filtered, use_container_width=True, height=500)
                
                st.markdown("### 📥 Download")
                col_d1, col_d2, col_d3 = st.columns(3)
                
                with col_d1:
                    _excel_download_button(
                        fo_filtered,
                        "FO",
                        "fo_xls_filtered",
                        label="📊 Excel",
                    )
                
                with col_d2:
                    try:
                        csv_data = to_csv_bytes(fo_filtered)
                        st.download_button("📄 CSV", csv_data, f"FO_{_now_ist().strftime('%Y%m%d_%H%M')}.csv",
                                         "text/csv", key="fo_csv")
                    except:
                        pass
                
                with col_d3:
                    try:
                        json_data = to_json_bytes(fo_filtered)
                        st.download_button("📋 JSON", json_data, f"FO_{_now_ist().strftime('%Y%m%d_%H%M')}.json",
                                         "application/json", key="fo_json")
                    except:
                        pass
            
            except Exception as e:
                st.error(f"❌ F&O Tab Error: {str(e)[:100]}")
        else:
            st.info("👈 Click 'SCAN F&O' to start")
    
    # ════════════════════════════════════════════════════════════════════════════════
    # TAB 2: MOMENTUM MOVERS — LIVE SUDDEN BUY / SELL
    # ════════════════════════════════════════════════════════════════════════════════
    with tabs[2]:
        st.markdown("### ⚡ LIVE SUDDEN MOVEMENT — BUY / SELL")
        st.caption("Only recent 5M price action + volume + acceleration + HH/HL or LH/LL. No previous consolidation, 15M, 1H or EMA confirmation.")

        col_m1, col_m2, col_m3 = st.columns([2, 2, 1])
        with col_m1:
            momentum_type = st.radio("Select Universe", ["NSE Stocks", "F&O Stocks"], horizontal=True, key="momentum_type")
            momentum_universe = all_symbols if momentum_type == "NSE Stocks" else fo_symbols
        with col_m2:
            momentum_limit = st.number_input("Scan limit", min_value=50, max_value=len(momentum_universe), value=min(500, len(momentum_universe)), step=50, key="momentum_limit")
        with col_m3:
            st.metric("Available", len(momentum_universe))

        momentum_symbols = momentum_universe[:momentum_limit]
        if st.button(f"⚡ SCAN LIVE MOVEMENT ({len(momentum_symbols)} stocks)", key="momentum_run", type="primary"):
            with st.spinner("Scanning current 5M sudden BUY / SELL movement…"):
                is_fo = momentum_type == "F&O Stocks"
                momentum_results, momentum_errors, momentum_stats = run_momentum_scan(fyers, momentum_symbols, is_fo=is_fo)
                mdf = pd.DataFrame(momentum_results) if momentum_results else pd.DataFrame()
                st.session_state["momentum_df"] = mdf
                st.session_state["momentum_errors"] = momentum_errors
                st.session_state["momentum_stats"] = momentum_stats
                st.session_state["momentum_scanned_at"] = _generated_timestamp()

        if "momentum_stats" in st.session_state:
            _display_scan_summary(st.session_state["momentum_stats"])
            st.caption(f"Last scan: {st.session_state.get('momentum_scanned_at', 'N/A')}")

        mdf = st.session_state.get("momentum_df")
        if mdf is not None and not mdf.empty:
            buy = mdf[mdf["DIRECTION"] == "BUY"].copy().sort_values(["SCORE", "RVOL"], ascending=False)
            sell = mdf[mdf["DIRECTION"] == "SELL"].copy().sort_values(["SCORE", "RVOL"], ascending=False)

            c1, c2 = st.columns(2)
            with c1:
                st.markdown(f"### 🟢 BUY — {len(buy)}")
                if not buy.empty:
                    st.dataframe(buy, use_container_width=True, height=430)
                else:
                    st.info("No current BUY movement found.")
            with c2:
                st.markdown(f"### 🔴 SELL — {len(sell)}")
                if not sell.empty:
                    st.dataframe(sell, use_container_width=True, height=430)
                else:
                    st.info("No current SELL movement found.")

            st.markdown("### 📥 Download LIVE MOVEMENT Report")
            try:
                st.download_button(
                    "📊 Excel",
                    _format_excel_output(mdf, "LIVE_MOVEMENT"),
                    f"LIVE_MOVEMENT_{_now_ist().strftime('%Y%m%d_%H%M')}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="momentum_xls",
                )
            except Exception as e:
                st.error(f"Excel export error: {e}")
        elif "momentum_stats" in st.session_state:
            st.warning("Scan completed — no current BUY/SELL movement matched the live thresholds.")
        else:
            st.info("👈 Click 'SCAN LIVE MOVEMENT' to find stocks moving NOW.")

        if st.session_state.get("momentum_errors"):
            with st.expander(f"⚠️ API / scan errors ({len(st.session_state['momentum_errors'])})"):
                st.dataframe(pd.DataFrame({"Error": st.session_state["momentum_errors"]}), use_container_width=True)
    # ════════════════════════════════════════════════════════════════════════════════
    # TAB 3: LIVE INTRADAY
    # ════════════════════════════════════════════════════════════════════════════════
    with tabs[3]:
        st.markdown("### ⚡ Live Intraday Scanner\nReal-time multi-timeframe analysis (5M, 15M, 1H)")
        
        col1, col2 = st.columns(2)
        with col1:
            intraday_type = st.radio("Select Universe", ["NSE Stocks", "F&O Stocks"], horizontal=True, key="intraday_type")
            intraday_universe = all_symbols if intraday_type == "NSE Stocks" else fo_symbols
        
        with col2:
            intraday_limit = st.number_input("Scan limit", min_value=10, max_value=len(intraday_universe),
                                            value=min(100, len(intraday_universe)), step=10, key="intraday_limit")
        
        intraday_symbols = intraday_universe[:intraday_limit]
        
        if st.button(f"⚡ SCAN LIVE INTRADAY ({len(intraday_symbols)} stocks)", key="intraday_run"):
            with st.spinner("Fetching live intraday data…"):
                if intraday_type == "NSE Stocks":
                    intraday_results, _, _ = run_nse_scan(fyers, intraday_symbols)
                else:
                    intraday_results, _, _ = run_fo_scan(fyers, intraday_symbols)
                
                if intraday_results:
                    intraday_df = _add_signal_time_columns(pd.DataFrame(intraday_results), "AI SIGNAL")
                    st.session_state["intraday_df"] = intraday_df
        
        intraday_df = st.session_state.get("intraday_df")
        if intraday_df is not None and not intraday_df.empty:
            intraday_df = _add_signal_time_columns(intraday_df, "AI SIGNAL")
            st.session_state["intraday_df"] = intraday_df
            st.success(f"✅ Live data: {len(intraday_df)} stocks")
            
            col_if1, col_if2, col_if3 = st.columns(3)
            with col_if1:
                intraday_min_rvol = st.slider("Min RVOL", 0.5, 3.0, 1.2, 0.1, key="intraday_rvol")
            with col_if2:
                intraday_signal_filter = st.selectbox("Signal", ["ALL", "BUY", "SELL"], key="intraday_sig_filter")
            with col_if3:
                intraday_show_cols = st.multiselect("Show Columns", intraday_df.columns, 
                                                   default=[c for c in ["Symbol", "LTP", "AI SIGNAL", "AI CONFIDENCE %", "SIGNAL TIME", "LAST SEEN", "SIGNAL AGE", "RVOL", "🟢 BUY PRESSURE %", "🔴 SELL PRESSURE %"] if c in intraday_df.columns],
                                                   key="intraday_cols")
            
            intraday_filtered = intraday_df.copy()
            try:
                intraday_filtered = intraday_filtered[pd.to_numeric(intraday_filtered["RVOL"], errors='coerce') >= intraday_min_rvol]
            except:
                pass
            
            if intraday_signal_filter != "ALL":
                try:
                    normalized = intraday_filtered["AI SIGNAL"].apply(normalize_signal)
                    if intraday_signal_filter == "BUY":
                        intraday_filtered = intraday_filtered[normalized == "BUY"]
                    elif intraday_signal_filter == "SELL":
                        intraday_filtered = intraday_filtered[normalized == "SELL"]
                except:
                    pass
            
            if intraday_show_cols:
                st.dataframe(intraday_filtered[intraday_show_cols], use_container_width=True, height=400)
            else:
                st.dataframe(intraday_filtered, use_container_width=True, height=400)
        else:
            st.info("👈 Click 'SCAN LIVE INTRADAY' to fetch data")
    
    # ════════════════════════════════════════════════════════════════════════════════
    # TAB 4: STRONG SIGNALS
    # ════════════════════════════════════════════════════════════════════════════════
    with tabs[4]:
        st.markdown("### 🔥 Strong Signals Only\nHigh-confidence setups (≥70%)")
        strong_source = st.radio("Source", ["NSE Stocks", "F&O Stocks"], horizontal=True, key="strong_source")

        strong_universe_all = all_symbols if strong_source == "NSE Stocks" else fo_symbols
        strong_default = min(500 if strong_source == "NSE Stocks" else 200, len(strong_universe_all))
        strong_limit = st.number_input(
            "Strong Signals scan limit (0 = all)",
            min_value=0, max_value=len(strong_universe_all),
            value=strong_default,
            step=50 if strong_source == "NSE Stocks" else 25,
            key="strong_limit"
        )

        if st.button("🔥 RUN STRONG SIGNALS", key="run_strong_signals", type="primary", use_container_width=True):
            with st.spinner(f"Scanning {strong_source} for strong signals…"):
                universe = strong_universe_all if strong_limit == 0 else strong_universe_all[:strong_limit]
                if strong_source == "NSE Stocks":
                    r, e, s = run_nse_scan(fyers, universe)
                    st.session_state["nse_df"] = _add_signal_time_columns(pd.DataFrame(r), "AI SIGNAL") if r else pd.DataFrame()
                    st.session_state["nse_errors"] = e
                    st.session_state["nse_stats"] = s
                else:
                    r, e, s = run_fo_scan(fyers, universe)
                    st.session_state["fo_df"] = _add_signal_time_columns(pd.DataFrame(r), "AI SIGNAL") if r else pd.DataFrame()
                    st.session_state["fo_errors"] = e
                    st.session_state["fo_stats"] = s

        strong_df = st.session_state.get("nse_df" if strong_source == "NSE Stocks" else "fo_df")
        
        if strong_df is not None and not strong_df.empty:
            try:
                confidence_col = pd.to_numeric(strong_df.get("AI CONFIDENCE %", pd.Series([])), errors='coerce')
                strong_filtered = strong_df[confidence_col >= 75].copy()
                
                normalized = strong_filtered["AI SIGNAL"].apply(normalize_signal)
                strong_filtered = strong_filtered[normalized != "NEUTRAL"]
                
                st.subheader(f"💪 {len(strong_filtered)} Strong Signals")
                
                if len(strong_filtered) > 0:
                    st.dataframe(strong_filtered.sort_values("AI CONFIDENCE %", ascending=False), 
                               use_container_width=True, height=400)
                    _excel_download_button(
                        strong_filtered.sort_values("AI CONFIDENCE %", ascending=False),
                        "STRONG_SIGNALS",
                        "download_strong_signals_excel"
                    )
                else:
                    st.warning("No strong signals (≥75% confidence) found. Lower the threshold in Settings tab.")
            
            except Exception as e:
                st.error(f"❌ Error: {str(e)[:100]}")
        else:
            st.info(f"👈 Run '{strong_source}' scanner first")
    
    # ════════════════════════════════════════════════════════════════════════════════
    # TAB 5: SWING (GOLDEN CROSS / DEATH CROSS)
    # ════════════════════════════════════════════════════════════════════════════════
    with tabs[5]:
        st.markdown("### 📈 Swing Analysis - Golden Cross & Death Cross Detection\nDaily EMA50/EMA200 crossovers")
        
        col1, col2 = st.columns([3, 1])
        with col1:
            swing_limit = st.number_input("Scan limit", min_value=10, max_value=len(all_symbols),
                                         value=min(100, len(all_symbols)), step=25, key="swing_limit")
        with col2:
            st.metric("Available", len(all_symbols))
        
        swing_symbols = all_symbols[:swing_limit]
        
        if st.button(f"📈 DETECT CROSSOVERS ({len(swing_symbols)} stocks)", key="swing_run"):
            with st.spinner("Analyzing daily charts for Golden/Death Cross…"):
                swing_results = []
                swing_progress = st.progress(0)
                
                for idx, symbol in enumerate(swing_symbols):
                    try:
                        cc_data = detect_golden_death_cross(fyers, symbol)
                        if cc_data.get("reason") == "OK":
                            ticker = symbol.replace("NSE:", "").replace("-EQ", "")
                            swing_results.append({
                                "Symbol": ticker,
                                "LTP": cc_data.get("ltp", "N/A"),
                                "EMA50": cc_data.get("ema50", "N/A"),
                                "EMA200": cc_data.get("ema200", "N/A"),
                                "EMA Trend": cc_data.get("ema_trend", "N/A"),
                                "Signal": cc_data.get("signal", "NONE"),
                                "Signal Date": cc_data.get("signal_date", "N/A"),
                            })
                    except:
                        pass
                    
                    swing_progress.progress((idx + 1) / len(swing_symbols))
                
                swing_progress.empty()
                
                if swing_results:
                    swing_df = pd.DataFrame(swing_results)
                    st.session_state["swing_df"] = swing_df
        
        swing_df = st.session_state.get("swing_df")
        if swing_df is not None and not swing_df.empty:
            st.success(f"✅ Analysis Complete: {len(swing_df)} stocks analyzed")
            
            col_s1, col_s2 = st.columns(2)
            with col_s1:
                swing_signal_filter = st.selectbox("Signal Type", ["ALL", "🟢 GOLDEN CROSS", "🔴 DEATH CROSS"], key="swing_sig_filter")
            with col_s2:
                swing_trend_filter = st.selectbox("EMA Trend", ["ALL", "BULLISH", "BEARISH", "NEUTRAL"], key="swing_trend_filter")
            
            swing_filtered = swing_df.copy()
            
            if swing_signal_filter != "ALL":
                swing_filtered = swing_filtered[swing_filtered["Signal"] == swing_signal_filter]
            
            if swing_trend_filter != "ALL":
                swing_filtered = swing_filtered[swing_filtered["EMA Trend"] == swing_trend_filter]
            
            if len(swing_filtered) > 0:
                golden = swing_filtered[swing_filtered["Signal"] == "🟢 GOLDEN CROSS"]
                death = swing_filtered[swing_filtered["Signal"] == "🔴 DEATH CROSS"]
                other = swing_filtered[swing_filtered["Signal"] == "NONE"]
                
                if len(golden) > 0:
                    st.subheader(f"🟢 Golden Cross ({len(golden)})")
                    st.dataframe(golden, use_container_width=True, height=250)
                
                if len(death) > 0:
                    st.subheader(f"🔴 Death Cross ({len(death)})")
                    st.dataframe(death, use_container_width=True, height=250)
                
                if len(other) > 0:
                    with st.expander(f"📊 Other ({len(other)})"):
                        st.dataframe(other, use_container_width=True, height=200)
            else:
                st.warning("No signals match current filters")
        else:
            st.info("👈 Click 'DETECT CROSSOVERS' to start swing analysis")
    
    # ════════════════════════════════════════════════════════════════════════════════
    # TAB 6: ADDITIONAL ANALYSIS
    # ════════════════════════════════════════════════════════════════════════════════
    with tabs[6]:
        st.markdown("### 🧠 Additional Analysis - Deep Dive on Single Stock")
        
        col1, col2 = st.columns(2)
        with col1:
            aa_universe = st.radio("Select Universe", ["NSE", "F&O"], horizontal=True, key="aa_universe")
            aa_symbols = all_symbols if aa_universe == "NSE" else fo_symbols
        
        with col2:
            aa_symbol_input = st.selectbox("Choose Stock", aa_symbols, key="aa_stock_select")
        
        if st.button("🔍 ANALYZE", key="aa_run"):
            with st.spinner(f"Analyzing {aa_symbol_input}…"):
                try:
                    analysis_5m = analyze_timeframe(fyers, aa_symbol_input, "5")
                    analysis_15m = analyze_timeframe(fyers, aa_symbol_input, "15")
                    analysis_1h = analyze_timeframe(fyers, aa_symbol_input, "60")
                    
                    if aa_universe == "F&O":
                        options_data = fetch_options_chain_data(fyers, aa_symbol_input)
                    else:
                        options_data = None
                    
                    master_signal = calculate_master_signal(aa_symbol_input, analysis_5m, analysis_15m, analysis_1h, options_data)
                    
                    st.session_state["aa_analysis"] = {
                        "5m": analysis_5m,
                        "15m": analysis_15m,
                        "1h": analysis_1h,
                        "master": master_signal,
                        "options": options_data,
                    }
                
                except Exception as e:
                    st.error(f"❌ Analysis failed: {str(e)}")
        
        aa_analysis = st.session_state.get("aa_analysis")
        if aa_analysis:
            ticker = aa_symbol_input.replace("NSE:", "").replace("-EQ", "")
            
            st.markdown(f"## {ticker} Analysis")
            master = aa_analysis.get("master", {})
            signal = master.get("final_signal", "NEUTRAL")
            conf = master.get("confidence", 0)
            
            if "BUY" in signal:
                st.success(f"{signal} — {conf:.0f}% Confidence")
            elif "SELL" in signal:
                st.error(f"{signal} — {conf:.0f}% Confidence")
            else:
                st.warning(f"{signal} — {conf:.0f}% Confidence")
            
            st.markdown("### ⏱️ Timeframe Analysis")
            col_5m, col_15m, col_1h = st.columns(3)
            
            with col_5m:
                d5 = aa_analysis["5m"].get("data", {})
                if d5:
                    st.write("**5 MINUTE**")
                    st.write(f"Trend: {d5.get('structure_trend', 'N/A')}")
                    st.write(f"Structure: {d5.get('structure_type', 'N/A')}")
                    st.write(f"LTP: {d5.get('last_close', 'N/A')}")
                    st.write(f"RSI: {d5.get('rsi', 'N/A')}")
                    st.write(f"RVOL: {d5.get('rvol', 'N/A')}x")
            
            with col_15m:
                d15 = aa_analysis["15m"].get("data", {})
                if d15:
                    st.write("**15 MINUTE**")
                    st.write(f"Trend: {d15.get('structure_trend', 'N/A')}")
                    st.write(f"Structure: {d15.get('structure_type', 'N/A')}")
                    st.write(f"VWAP: {d15.get('vwap', 'N/A')}")
                    st.write(f"EMA Trend: {d15.get('ema_trend', 'N/A')}")
            
            with col_1h:
                d1h = aa_analysis["1h"].get("data", {})
                if d1h:
                    st.write("**1 HOUR**")
                    st.write(f"Trend: {d1h.get('structure_trend', 'N/A')}")
                    st.write(f"Structure: {d1h.get('structure_type', 'N/A')}")
                    st.write(f"EMA50: {d1h.get('ema50', 'N/A')}")
                    st.write(f"EMA200: {d1h.get('ema200', 'N/A')}")
            
            st.markdown("### 📊 Pressure Analysis")
            if d5:
                bp = d5.get("buying_pressure", 0)
                sp = d5.get("selling_pressure", 0)
                col_bp1, col_bp2 = st.columns(2)
                with col_bp1:
                    st.metric("🟢 Buying Pressure", f"{bp}%")
                with col_bp2:
                    st.metric("🔴 Selling Pressure", f"{sp}%")
            
            st.markdown("### 🏗️ Market Structure")
            col_struct1, col_struct2, col_struct3 = st.columns(3)
            
            with col_struct1:
                if d5 and d5.get("bullish_choch"):
                    st.success("✅ Bullish CHoCH")
                elif d5 and d5.get("bearish_choch"):
                    st.error("❌ Bearish CHoCH")
            
            with col_struct2:
                if d5 and d5.get("bullish_mss"):
                    st.success("✅ Bullish MSS")
                elif d5 and d5.get("bearish_mss"):
                    st.error("❌ Bearish MSS")
            
            with col_struct3:
                if d5 and d5.get("bullish_cisd"):
                    st.success("✅ Bullish CISD")
                elif d5 and d5.get("bearish_cisd"):
                    st.error("❌ Bearish CISD")
            
            st.markdown("### 💹 Trade Plan")
            col_tp1, col_tp2, col_tp3, col_tp4 = st.columns(4)
            
            entry = master.get("entry", "N/A")
            sl = master.get("stop_loss", "N/A")
            t1 = master.get("target1", "N/A")
            t2 = master.get("target2", "N/A")
            rr = master.get("rr_ratio", "N/A")
            
            with col_tp1:
                st.metric("Entry", f"{entry}")
            with col_tp2:
                st.metric("Stop Loss", f"{sl}")
            with col_tp3:
                st.metric("Target 1", f"{t1}")
            with col_tp4:
                st.metric("Target 2 / R:R", f"{t2} / {rr}")
            
            st.markdown("### 📝 Signal Explanation")
            reason = master.get("signal_reason", "N/A")
            st.info(reason)
            
            if aa_universe == "F&O" and aa_analysis.get("options"):
                st.markdown("### 📊 Options Analysis")
                opt = aa_analysis["options"]
                col_opt1, col_opt2, col_opt3, col_opt4 = st.columns(4)
                
                with col_opt1:
                    st.metric("ATM Strike", opt.get("atm_strike", "N/A"))
                with col_opt2:
                    st.metric("PCR", opt.get("pcr", "N/A"))
                with col_opt3:
                    st.metric("CE OI", opt.get("ce_oi", "N/A"))
                with col_opt4:
                    st.metric("Options Bias", opt.get("options_bias", "N/A"))
    
    # ════════════════════════════════════════════════════════════════════════════════
    # TAB 7: MARKET DASHBOARD
    # ════════════════════════════════════════════════════════════════════════════════
    with tabs[7]:
        st.markdown("### 📊 Market Dashboard - Statistics & Sentiment")
        dashboard_source = st.radio("Data Source", ["NSE Stocks", "F&O Stocks"], horizontal=True, key="dash_source")

        dash_universe_all = all_symbols if dashboard_source == "NSE Stocks" else fo_symbols
        dash_default = min(500 if dashboard_source == "NSE Stocks" else 200, len(dash_universe_all))
        dash_limit = st.number_input(
            "Dashboard scan limit (0 = all)",
            min_value=0, max_value=len(dash_universe_all),
            value=dash_default,
            step=50 if dashboard_source == "NSE Stocks" else 25,
            key="dash_limit"
        )

        if st.button("📊 RUN MARKET DASHBOARD", key="run_market_dashboard", type="primary", use_container_width=True):
            with st.spinner(f"Scanning {dashboard_source} for market dashboard…"):
                universe = dash_universe_all if dash_limit == 0 else dash_universe_all[:dash_limit]
                if dashboard_source == "NSE Stocks":
                    r, e, s = run_nse_scan(fyers, universe)
                    st.session_state["nse_df"] = _add_signal_time_columns(pd.DataFrame(r), "AI SIGNAL") if r else pd.DataFrame()
                    st.session_state["nse_errors"] = e
                    st.session_state["nse_stats"] = s
                else:
                    r, e, s = run_fo_scan(fyers, universe)
                    st.session_state["fo_df"] = _add_signal_time_columns(pd.DataFrame(r), "AI SIGNAL") if r else pd.DataFrame()
                    st.session_state["fo_errors"] = e
                    st.session_state["fo_stats"] = s

        dash_df = st.session_state.get("nse_df" if dashboard_source == "NSE Stocks" else "fo_df")
        
        if dash_df is not None and not dash_df.empty:
            stats = calculate_market_stats(dash_df)
            
            st.markdown("### 📈 Market Overview")
            col_ov1, col_ov2, col_ov3, col_ov4, col_ov5 = st.columns(5)
            
            with col_ov1:
                st.metric("Total Scanned", stats["total"])
            with col_ov2:
                st.metric("🟢 BUY", stats["buy"])
            with col_ov3:
                st.metric("🔴 SELL", stats["sell"])
            with col_ov4:
                st.metric("🟡 NEUTRAL", stats["neutral"])
            with col_ov5:
                st.metric("Avg Confidence", f"{stats['avg_confidence']:.1f}%")
            
            st.markdown("### 💪 Strong Signals")
            col_str1, col_str2 = st.columns(2)
            
            with col_str1:
                st.metric("💪 Strong BUY", stats["strong_buy"])
            with col_str2:
                st.metric("💪 Strong SELL", stats["strong_sell"])
            
            st.markdown("### 😊 Market Sentiment")
            col_sent1, col_sent2, col_sent3 = st.columns(3)
            
            with col_sent1:
                st.metric("Bullish %", f"{stats['buy_pct']:.1f}%")
            with col_sent2:
                st.metric("Bearish %", f"{stats['sell_pct']:.1f}%")
            with col_sent3:
                st.metric("Neutral %", f"{stats['neutral_pct']:.1f}%")

            _excel_download_button(
                dash_df,
                "MARKET_DASHBOARD",
                "download_market_dashboard_excel"
            )
        else:
            st.info(f"👈 Run '{dashboard_source}' scanner first")
    
    # ════════════════════════════════════════════════════════════════════════════════
    # TAB 8: SETTINGS
    # ════════════════════════════════════════════════════════════════════════════════
    with tabs[8]:
        st.markdown("### ⚙️ Scanner Settings & Configuration")
        
        st.markdown("#### 🎯 Signal Filtering")
        col_set1, col_set2, col_set3 = st.columns(3)
        
        with col_set1:
            default_conf = st.number_input("Default Min Confidence %", 0, 100, DEFAULT_CONFIDENCE_THRESHOLD, 5, key="set_conf")
        with col_set2:
            default_rvol = st.slider("Default Min RVOL", 0.5, 3.0, DEFAULT_RVOL_THRESHOLD, 0.1, key="set_rvol")
        with col_set3:
            default_strong_rvol = st.slider("Strong Signal RVOL", 1.0, 3.0, DEFAULT_STRONG_RVOL, 0.1, key="set_strong_rvol")
        
        st.markdown("#### ⚡ Live Movement Settings")
        col_mom1, col_mom2, col_mom3 = st.columns(3)
        
        with col_mom1:
            momentum_min_score = st.slider("Min Live Movement Score", 50, 100, MOMENTUM_MIN_SCORE, 5, key="set_mom_score")
        with col_mom2:
            momentum_strong_score = st.slider("Strong BIG MOVE Score", 75, 100, MOMENTUM_STRONG_SCORE, 5, key="set_mom_strong")
        with col_mom3:
            momentum_min_rvol = st.slider("Min Live Movement RVOL", 1.0, 3.0, MOMENTUM_MIN_RVOL, 0.1, key="set_mom_rvol")
        
        st.markdown("#### ℹ️ Information")
        st.info("""
        **NSE AI PRO V17 — Features:**
        - ✅ Multi-timeframe analysis (5M, 15M, 1H, Daily)
        - ✅ Strict signal validation engine
        - ✅ Pressure-based confirmation
        - ✅ VWAP, EMA, RSI, MACD indicators
        - ✅ Market structure (CHoCH, MSS, CISD)
        - ✅ Options chain analysis (F&O)
        - ✅ Golden Cross / Death Cross detection
        - ✅ **⚡ NEW: MOMENTUM MOVERS Scanner**
        - ✅ Next Candle Bias prediction
        
        **Data Source:** Fyers Live API
        **Timeframes:** 5M, 15M, 1H, Daily
        **Universes:** NSE Equities + F&O Stocks
        """)
    
    # ════════════════════════════════════════════════════════════════════════════════
    # TAB 9: PIN RULES — ADDITIONAL ONLY
    # ════════════════════════════════════════════════════════════════════════════════
    with tabs[9]:
        _show_pin_rules_tab(fyers)

    # ════════════════════════════════════════════════════════════════════════════════
    # TAB 10: PIN FULL SCAN — ADDITIONAL ONLY
    # ════════════════════════════════════════════════════════════════════════════════
    with tabs[10]:
        _show_pin_full_scan_tab(fyers, all_symbols, fo_symbols)

    # ════════════════════════════════════════════════════════════════════════════════
    # TAB 11: AMD FULL SCAN — ADDITIONAL ONLY
    # ════════════════════════════════════════════════════════════════════════════════
    with tabs[11]:
        _show_amd_scan_tab(fyers, all_symbols, fo_symbols)
    
    gc.collect()

# ════════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ════════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    try:
        access_token = os.environ.get("FYERS_ACCESS_TOKEN")
        if not access_token:
            st.error("❌ FYERS_ACCESS_TOKEN not set in environment variables")
            st.info("Please set: export FYERS_ACCESS_TOKEN='your_token_here'")
            st.stop()
        
        try:
            from fyers_api import fyersModel
        except ImportError as ie:
            st.error("❌ fyers-api not installed")
            st.code("pip install fyers-api", language="bash")
            st.stop()
        
        app_id = os.environ.get("FYERS_APP_ID", "DEMO")
        
        try:
            fyers = fyersModel.FyersModel(client_id=app_id, token=access_token, log_path="")
            logger.info("Fyers client initialized successfully")
            show_scanner(fyers)
        except Exception as init_error:
            st.error(f"❌ Failed to initialize Fyers API: {str(init_error)}")
            logger.error(f"Fyers initialization error: {init_error}", exc_info=True)
            
            with st.expander("Debug Information"):
                st.write(f"App ID: {app_id}")
                st.write(f"Token set: {bool(access_token)}")
                st.write(f"Error: {init_error}")
    
    except Exception as e:
        st.error(f"❌ Unexpected Error: {e}")
        logger.error(f"Unexpected error: {e}", exc_info=True)
